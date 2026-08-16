"""File card photos so they can be used as eBay picture URLs.

A bulk upload carries image *links*, not image files, so every photo needs a
public https address before it is any use. These get published with the Pages
site, which is a legitimate free host for them.

    python add_photos.py --sku CRH-0001 front.jpg back.jpg
    python add_photos.py --inbox          # everything in photos/inbox
    python add_photos.py --status         # which cards still have no photo
    python add_photos.py --publish        # commit and push, so they go live

EXIF is stripped, which is not cosmetic: a phone photo carries the GPS
coordinates of wherever it was taken, and these land in a public repo.
"""

import argparse
import os
import re
import subprocess
import sys

import inuse

from PIL import Image, ImageOps

PHOTOS = "photos"
INBOX = os.path.join(PHOTOS, "inbox")
WORKBOOK = "Card Run HQ - Master.xlsx"

# eBay wants 500px minimum on the longest side and recommends 1600px, which is
# also the point where its zoom viewer switches on. Bigger just costs upload.
LONG_EDGE = 1600
QUALITY = 88
SKU_RE = re.compile(r"(CRH-\d{4})", re.I)

# Suffix in the saved name -> words in a source filename that mean it.
SLOTS = [
    ("", ("front", "f")),
    ("-back", ("back", "rear", "b", "reverse")),
    ("-2", ("2", "detail")),
    ("-3", ("3", "corner")),
]


def slot_from_name(name):
    """Work out front/back/extra from whatever the file happens to be called."""
    words = set(re.split(r"[^a-z0-9]+", name.lower()))
    for suffix, hints in SLOTS:
        if suffix and words & set(hints):
            return suffix
    return ""


def free_slot(sku):
    for suffix, _hints in SLOTS:
        if not os.path.exists(dest_for(sku, suffix)):
            return suffix
    return None


def dest_for(sku, suffix):
    return os.path.join(PHOTOS, "%s%s.jpg" % (sku, suffix))


def convert(src, dest):
    """Resize, flatten and re-encode without any metadata."""
    with Image.open(src) as im:
        im = ImageOps.exif_transpose(im)          # honour the rotation flag
        if im.mode not in ("RGB", "L"):
            im = im.convert("RGB")
        w, h = im.size
        if max(w, h) > LONG_EDGE:                 # never upscale
            scale = LONG_EDGE / float(max(w, h))
            im = im.resize((max(1, round(w * scale)), max(1, round(h * scale))),
                           Image.LANCZOS)
        if min(im.size) < 500:
            print("   warning: %dx%d is under eBay's 500px minimum"
                  % im.size)
        # Rebuild from raw pixels: the new image has no info dict and no EXIF
        # block, so nothing can survive into the file by accident.
        clean = Image.frombytes(im.mode, im.size, im.tobytes())
        clean.save(dest, "JPEG", quality=QUALITY, optimize=True,
                   progressive=True)
    return os.path.getsize(dest)


def place(src, sku, suffix=None):
    if suffix is None:
        suffix = slot_from_name(os.path.basename(src))
        if os.path.exists(dest_for(sku, suffix)):
            suffix = free_slot(sku)
            if suffix is None:
                print("   %s: all four slots taken, skipped" % sku)
                return False
    dest = dest_for(sku, suffix)
    size = convert(src, dest)
    print("   %-28s -> %s  (%d KB)"
          % (os.path.basename(src), os.path.basename(dest), size // 1024))
    return True


def skus_from_workbook():
    """SKU -> player name, so --status can say what is missing a photo."""
    if not os.path.exists(WORKBOOK):
        return {}
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {}
    ws = load_workbook(WORKBOOK, data_only=True)["Inventory"]
    head = [str(c.value or "") for c in ws[1]]
    try:
        i_sku = head.index("SKU")
        i_who = head.index("Player or card name")
    except ValueError:
        return {}
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[i_sku] and row[i_who]:
            out[str(row[i_sku]).strip()] = str(row[i_who]).strip()
    return out


def photos_for(sku):
    return [dest_for(sku, s) for s, _h in SLOTS if os.path.exists(dest_for(sku, s))]


def do_status():
    cards = skus_from_workbook()
    if not cards:
        print("no cards in the workbook yet")
        return
    missing = []
    for sku, who in sorted(cards.items()):
        have = photos_for(sku)
        mark = "ok " if have else "-- "
        print("%s %-10s %-26s %d photo%s"
              % (mark, sku, who[:26], len(have), "" if len(have) == 1 else "s"))
        if not have:
            missing.append(sku)
    print()
    if missing:
        print("%d card%s with no photo. eBay will take the listing without one,"
              % (len(missing), "" if len(missing) == 1 else "s"))
        print("but a raw single with no picture does not sell.")
    else:
        print("every card has at least one photo")


EXTS = (".jpg", ".jpeg", ".png", ".heic", ".webp", ".tif", ".tiff", ".bmp")


def do_import(src_dir, assign=None, pairs=False, move=False):
    """Pull photos out of a folder -- our inbox, or whatever the scanner writes.

    A scanner names its output scan0001.jpg with no SKU in it, so --assign
    maps the files onto SKUs in filename order instead, and --pairs reads
    them as front, back, front, back.
    """
    if not os.path.isdir(src_dir):
        if src_dir == INBOX:
            os.makedirs(INBOX, exist_ok=True)
            print("made %s -- drop photos in there and run this again." % INBOX)
            print("Name them with the SKU (CRH-0001-front.jpg), or use "
                  "--assign to map them in order.")
        else:
            print("no such folder: %s" % src_dir)
        return

    files = sorted(f for f in os.listdir(src_dir)
                   if os.path.splitext(f)[1].lower() in EXTS)
    if not files:
        print("nothing to import from %s" % src_dir)
        return

    plan = []
    if assign:
        skus = [s.strip().upper() for s in assign.split(",") if s.strip()]
        if pairs:
            wanted = len(skus) * 2
            if len(files) != wanted:
                print("--pairs with %d SKUs expects %d files, found %d."
                      % (len(skus), wanted, len(files)))
                print("Nothing filed -- fix the count so fronts and backs "
                      "cannot end up on the wrong cards.")
                return
            for i, sku in enumerate(skus):
                plan.append((files[i * 2], sku, ""))
                plan.append((files[i * 2 + 1], sku, "-back"))
        else:
            if len(files) != len(skus):
                print("--assign lists %d SKUs but the folder holds %d files."
                      % (len(skus), len(files)))
                print("Nothing filed -- a silent off-by-one would put photos "
                      "on the wrong cards.")
                return
            plan = [(f, sku, None) for f, sku in zip(files, skus)]
        print("filing in filename order:")
        for f, sku, slot in plan:
            print("   %-28s -> %s%s" % (f, sku, slot or ""))
        print()
    else:
        for f in files:
            m = SKU_RE.search(f)
            plan.append((f, m.group(1).upper() if m else None, None))

    done, skipped = 0, []
    for f, sku, slot in plan:
        if not sku:
            skipped.append(f)
            continue
        src = os.path.join(src_dir, f)
        try:
            if place(src, sku, slot):
                if move:
                    os.remove(src)
                done += 1
        except Exception as e:            # a bad file should not stop the batch
            print("   %s: %s" % (f, e))
            skipped.append(f)

    print("\nfiled %d%s, skipped %d"
          % (done, ", removed from the source" if move else "", len(skipped)))
    for f in skipped:
        print("   no SKU in the name: %s" % f)
    if skipped:
        print("rename them to include the SKU, or use --assign to map them "
              "in order.")
    if done and not move:
        print("source files left where they were. Pass --move to clear them.")


def do_publish():
    if not os.path.isdir(PHOTOS):
        sys.exit("no photos folder")
    files = [f for f in os.listdir(PHOTOS) if f.lower().endswith(".jpg")]
    if not files:
        sys.exit("no photos to publish")
    print("publishing %d photo(s) to the public site" % len(files))
    cmds = [
        ["git", "add", "--", PHOTOS],
        ["git", "commit", "-m", "Add card photos for eBay picture URLs"],
        ["git", "push", "origin", "master"],
    ]
    for cmd in cmds:
        r = subprocess.run(cmd, capture_output=True, text=True)
        if r.returncode and "nothing to commit" not in (r.stdout + r.stderr):
            sys.exit("%s failed:\n%s%s" % (" ".join(cmd), r.stdout, r.stderr))
    print("pushed. The site rebuilds in a couple of minutes, then run "
          "make_ebay_csv.py so PicURL picks them up.")


def main():
    p = argparse.ArgumentParser()
    p.add_argument("files", nargs="*", help="image files to file")
    p.add_argument("--sku", help="SKU these files belong to")
    p.add_argument("--slot", choices=[s for s, _h in SLOTS if s] + ["front"],
                   help="force front/back/-2/-3 instead of guessing")
    p.add_argument("--inbox", action="store_true",
                   help="process everything in photos/inbox")
    p.add_argument("--from", dest="src", metavar="DIR",
                   help="import from any folder, e.g. the scanner's output")
    p.add_argument("--assign", metavar="SKU,SKU,...",
                   help="map files to these SKUs in filename order, for "
                        "scanner output with no SKU in the name")
    p.add_argument("--pairs", action="store_true",
                   help="with --assign, read the files as front, back, "
                        "front, back")
    p.add_argument("--move", action="store_true",
                   help="delete the source files once filed (off by default, "
                        "so a scanner folder is never emptied by surprise)")
    p.add_argument("--status", action="store_true",
                   help="list which cards still have no photo")
    p.add_argument("--publish", action="store_true",
                   help="commit and push the photos so they go live")
    a = p.parse_args()

    inuse.refuse_if_open(a.workbook)

    os.makedirs(PHOTOS, exist_ok=True)

    if a.status:
        return do_status()
    if a.publish:
        return do_publish()
    if a.inbox or a.src or a.assign:
        return do_import(a.src or INBOX, a.assign, a.pairs,
                         a.move or (not a.src and a.inbox))

    if not a.files:
        return do_status()
    if not a.sku:
        sys.exit("--sku is required when naming files directly")
    sku = a.sku.upper()
    slot = None
    if a.slot:
        slot = "" if a.slot == "front" else a.slot
    print("filing %d photo(s) for %s" % (len(a.files), sku))
    for i, f in enumerate(a.files):
        if not os.path.exists(f):
            print("   missing: %s" % f)
            continue
        # An explicit slot applies to the first file; the rest fall in behind.
        place(f, sku, slot if (slot is not None and i == 0) else None)
    print("\n%s now has %d photo(s)" % (sku, len(photos_for(sku))))


if __name__ == "__main__":
    main()
