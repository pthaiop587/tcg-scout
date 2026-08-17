"""Put a clickable link to each card's photos in the Inventory tab.

    python link_photos.py            # say what it would do
    python link_photos.py --go       # write the links in

The photos are already named after the SKU they belong to -- CRH-0062.jpg and
CRH-0062-back.jpg -- so nothing needs matching up. What was missing is a way to
get at them from the row: when you are writing a listing you are looking at the
sheet, and the pictures were a folder away with sixty files in it.

So Inventory gets a **Photos** column holding a hyperlink. Click it and the
front opens; the back is the next file along. The cell says how many there
are, so a card with none is visible at a glance rather than something you find
out at the point of listing.

The link is RELATIVE -- photos\\CRH-0062.jpg, not G:\\Claude\\... -- so it
still works if the folder is moved or copied to another machine, which an
absolute path would not.

It is a real Excel hyperlink rather than a HYPERLINK() formula, because
make_ebay_csv.py and prices.py read the workbook with data_only=True: a formula
Excel has not recalculated reads as empty, and the column would look blank to
every script that opened it.
"""

import argparse
import os
import re
import sys

from openpyxl import load_workbook
from openpyxl.styles import Font

import inuse

WORKBOOK = "Card Run HQ - Master.xlsx"
PHOTOS = "photos"
COLUMN = "Photos"
SKU_RE = re.compile(r"^(CRH-\d{4})(-.*)?\.jpg$", re.I)

LINKFONT = Font(color="0563C1", underline="single")


def by_sku(folder):
    """Every filed photo, grouped by the card it belongs to."""
    out = {}
    if not os.path.isdir(folder):
        return out
    for f in sorted(os.listdir(folder)):
        m = SKU_RE.match(f)
        if m:
            out.setdefault(m.group(1).upper(), []).append(f)
    # front first: CRH-0062.jpg sorts before CRH-0062-back.jpg anyway, but be
    # explicit rather than relying on it
    for sku in out:
        out[sku].sort(key=lambda n: (len(n), n))
    return out


def ensure_column(ws):
    """Append Photos if it is not there. Appended, never inserted:
    workbook_extra.py addresses Inventory by column letter."""
    hdr = [c.value for c in ws[1]]
    if COLUMN in hdr:
        return hdr, False
    col = len(hdr) + 1
    c = ws.cell(row=1, column=col, value=COLUMN)
    base = ws.cell(row=1, column=1)
    from copy import copy
    c.font, c.fill, c.alignment = (copy(base.font), copy(base.fill),
                                   copy(base.alignment))
    from openpyxl.utils import get_column_letter
    ws.column_dimensions[get_column_letter(col)].width = 14
    hdr.append(COLUMN)
    return hdr, True


def main():
    p = argparse.ArgumentParser(
        description="Link each card's photos from its Inventory row.")
    p.add_argument("--workbook", default=WORKBOOK)
    p.add_argument("--photos", default=PHOTOS)
    p.add_argument("--go", action="store_true", help="write the links in")
    a = p.parse_args()

    inuse.refuse_if_open(a.workbook)

    if not os.path.exists(a.workbook):
        sys.exit("no workbook at %s" % a.workbook)

    shots = by_sku(a.photos)
    if not shots:
        print("No photos in %s yet. File some with add_photos.py first."
              % a.photos)
        return 0

    wb = load_workbook(a.workbook)
    ws = wb["Inventory"]
    hdr, added = ensure_column(ws)
    g = {n: i + 1 for i, n in enumerate(hdr) if n}
    if added:
        print("added a %r column at the end of Inventory" % COLUMN)

    linked = cleared = 0
    for r in range(2, ws.max_row + 1):
        sku = str(ws.cell(row=r, column=g["SKU"]).value or "").strip().upper()
        if not sku:
            continue
        cell = ws.cell(row=r, column=g[COLUMN])
        files = shots.get(sku)
        if not files:
            if cell.value not in (None, ""):
                if a.go:
                    cell.value, cell.hyperlink = None, None
                cleared += 1
            continue

        # Relative to the WORKBOOK, not to wherever this was run from, and
        # not simply whatever --photos was spelled as. An absolute link works
        # until the folder is copied, and then points at a machine that is not
        # this one.
        here = os.path.dirname(os.path.abspath(a.workbook)) or "."
        full = os.path.abspath(os.path.join(a.photos, files[0]))
        try:
            target = os.path.relpath(full, here).replace("\\", "/")
        except ValueError:          # different drive: nothing relative exists
            target = full.replace("\\", "/")
        label = "%d photo%s" % (len(files), "" if len(files) == 1 else "s")
        if a.go:
            cell.value = label
            cell.hyperlink = target
            cell.font = LINKFONT
        linked += 1
        if linked <= 6:
            print("   %-9s %-10s -> %s" % (sku, label, target))

    if linked > 6:
        print("   ... and %d more" % (linked - 6))
    print("\n%d card(s) linked%s"
          % (linked, ", %d cleared (photos gone)" % cleared if cleared else ""))

    have = {s for s in shots}
    print("%d photo file(s) across %d card(s) in %s"
          % (sum(len(v) for v in shots.values()), len(have), a.photos))

    if a.go:
        wb.save(a.workbook)
        print("\nSaved %s. Click a Photos cell and the front opens."
              % a.workbook)
    else:
        print("\nNothing written. Add --go to put the links in.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
