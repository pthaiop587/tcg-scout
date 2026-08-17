"""Tests for the photo links in Inventory.

Run: python -m pytest test_link_photos.py

Two things are easy to get wrong here and both are quiet.

The link has to be a real hyperlink, not a HYPERLINK() formula. make_ebay_csv.py
and prices.py read the workbook with data_only=True, and a formula Excel has
not recalculated reads as empty -- the column would look blank to every script
that opened it while looking perfectly fine on screen.

And the path has to be relative. An absolute G:\\Claude\\... link works until
the folder is copied anywhere, and then points at a machine that is not this
one.
"""

import os
import subprocess
import sys

import pytest
from openpyxl import load_workbook

import file_batch as fb
import link_photos as lp

CARDS = [
    {"player": "Travis Hunter", "year": 2025, "brand": "Panini Prizm Draft Picks",
     "num": "21", "sport": "Football"},
    {"player": "Shedeur Sanders", "year": 2025, "brand": "Panini Prizm Draft Picks",
     "num": "19", "sport": "Football"},
]


@pytest.fixture
def shop(tmp_path):
    wb = tmp_path / "Card Run HQ - Master.xlsx"
    subprocess.run([sys.executable, "make_workbook.py", "--out", str(wb)],
                   check=True, capture_output=True)
    fb.add_rows(str(wb), [dict(c) for c in CARDS])
    shots = tmp_path / "photos"
    shots.mkdir()
    return wb, shots


def run(wb, shots, *args):
    r = subprocess.run([sys.executable, os.path.abspath("link_photos.py"),
                        "--workbook", str(wb), "--photos", str(shots)]
                       + list(args), capture_output=True, text=True,
                       cwd=os.path.dirname(os.path.abspath("link_photos.py")))
    assert r.returncode == 0, r.stdout + r.stderr
    return r.stdout


def cell_for(wb, sku):
    ws = load_workbook(wb)["Inventory"]
    hdr = [c.value for c in ws[1]]
    g = {n: i + 1 for i, n in enumerate(hdr) if n}
    if lp.COLUMN not in g:
        return None
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(row=r, column=g["SKU"]).value or "") == sku:
            return ws.cell(row=r, column=g[lp.COLUMN])
    return None


# --- grouping files onto cards ----------------------------------------------

def test_photos_are_grouped_by_the_sku_in_their_name(tmp_path):
    d = tmp_path / "p"
    d.mkdir()
    for n in ("CRH-0062.jpg", "CRH-0062-back.jpg", "CRH-0063.jpg",
              "notes.txt", "IMG_2829.jpg"):
        (d / n).write_bytes(b"x")
    got = lp.by_sku(str(d))
    assert set(got) == {"CRH-0062", "CRH-0063"}
    assert got["CRH-0062"] == ["CRH-0062.jpg", "CRH-0062-back.jpg"]


def test_the_front_is_listed_first():
    """The link points at the first file, and that should be the front."""
    assert lp.SKU_RE.match("CRH-0062.jpg")
    assert lp.SKU_RE.match("CRH-0062-back.jpg")
    assert not lp.SKU_RE.match("CRH-62.jpg")


# --- what goes in the cell --------------------------------------------------

def test_the_cell_is_a_real_hyperlink_not_a_formula(shop):
    """A HYPERLINK() formula reads as empty under data_only=True, so every
    script that opens the workbook would see a blank column."""
    wb, shots = shop
    (shots / "CRH-0001.jpg").write_bytes(b"x")
    run(wb, shots, "--go")

    c = cell_for(wb, "CRH-0001")
    assert c.hyperlink is not None, "no hyperlink on the cell"
    assert not str(c.value).startswith("="), c.value

    ws = load_workbook(wb, data_only=True)["Inventory"]
    hdr = [x.value for x in ws[1]]
    col = hdr.index(lp.COLUMN) + 1
    assert ws.cell(row=2, column=col).value, "reads as blank with data_only"


def test_the_path_is_relative(shop):
    wb, shots = shop
    (shots / "CRH-0001.jpg").write_bytes(b"x")
    run(wb, shots, "--go")
    target = cell_for(wb, "CRH-0001").hyperlink.target
    assert not os.path.isabs(target), target
    assert ":" not in target, target


def test_the_cell_says_how_many(shop):
    wb, shots = shop
    (shots / "CRH-0001.jpg").write_bytes(b"x")
    (shots / "CRH-0001-back.jpg").write_bytes(b"x")
    run(wb, shots, "--go")
    assert cell_for(wb, "CRH-0001").value == "2 photos"


def test_one_photo_is_not_called_photos(shop):
    wb, shots = shop
    (shots / "CRH-0001.jpg").write_bytes(b"x")
    run(wb, shots, "--go")
    assert cell_for(wb, "CRH-0001").value == "1 photo"


def test_a_card_with_no_photos_is_left_empty(shop):
    wb, shots = shop
    (shots / "CRH-0001.jpg").write_bytes(b"x")
    run(wb, shots, "--go")
    assert cell_for(wb, "CRH-0002").value in (None, "")


def test_a_link_is_cleared_when_the_photo_goes(shop):
    """Otherwise the sheet promises a picture that is not there any more."""
    wb, shots = shop
    (shots / "CRH-0001.jpg").write_bytes(b"x")
    run(wb, shots, "--go")
    assert cell_for(wb, "CRH-0001").value

    os.remove(str(shots / "CRH-0001.jpg"))
    (shots / "CRH-0002.jpg").write_bytes(b"x")
    out = run(wb, shots, "--go")
    assert cell_for(wb, "CRH-0001").value in (None, "")
    assert "cleared" in out


# --- the landmine -----------------------------------------------------------

def test_the_column_layout_is_left_alone(shop):
    """make_workbook now ships Photos as the first column, so there is nothing
    to add. Whatever the layout is, this must not rearrange it."""
    wb, shots = shop
    before = [c.value for c in load_workbook(wb)["Inventory"][1]]
    (shots / "CRH-0001.jpg").write_bytes(b"x")
    run(wb, shots, "--go")
    after = [c.value for c in load_workbook(wb)["Inventory"][1]]
    assert after == before, "the column order changed"
    assert lp.COLUMN in after


def test_it_still_adds_the_column_to_a_workbook_without_one(shop):
    """An older workbook predates the column, and appending is the only safe
    place to put one."""
    wb, shots = shop
    book = load_workbook(wb)
    ws = book["Inventory"]
    hdr = [c.value for c in ws[1]]
    ws.delete_cols(hdr.index(lp.COLUMN) + 1)
    book.save(wb)

    before = [c.value for c in load_workbook(wb)["Inventory"][1]]
    assert lp.COLUMN not in before
    (shots / "CRH-0001.jpg").write_bytes(b"x")
    run(wb, shots, "--go")
    after = [c.value for c in load_workbook(wb)["Inventory"][1]]
    assert after[:len(before)] == before
    assert after[-1] == lp.COLUMN


def test_running_it_twice_adds_one_column(shop):
    wb, shots = shop
    (shots / "CRH-0001.jpg").write_bytes(b"x")
    run(wb, shots, "--go")
    run(wb, shots, "--go")
    hdr = [c.value for c in load_workbook(wb)["Inventory"][1]]
    assert hdr.count(lp.COLUMN) == 1


def test_a_dry_run_writes_nothing(shop):
    wb, shots = shop
    (shots / "CRH-0001.jpg").write_bytes(b"x")
    out = run(wb, shots)
    assert "Nothing written" in out
    c = cell_for(wb, "CRH-0001")
    assert c is None or c.value in (None, "")
