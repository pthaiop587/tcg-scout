"""Upload card photos to eBay Picture Services and remember the URLs it gives back.

    python ebay_pics.py --check                 # credentials only, no upload
    python ebay_pics.py --sku CRH-0122          # one card
    python ebay_pics.py                         # every card ready to list
    python ebay_pics.py --go                    # actually upload

WHY UPLOAD TO EBAY AT ALL.

eBay copies a self-hosted picture onto its own servers during the upload
anyway. So any host you put in PicURL is a delivery truck: it has to be
reachable at that moment, over HTTPS, and then never again. Skipping the truck
and handing the picture straight to eBay removes the whole problem -- nothing
public, no bandwidth, no repository swelling with 500KB scans that git will
keep for ever.

GitHub Pages was the right answer for five listings. It is the wrong one for
nine hundred: a published Pages site may be no larger than 1GB, these photos
run about 1MB a card, and deleting them later does not shrink the repository
because history keeps every byte that was ever committed.

WHAT IT NEEDS FROM YOU.

An eBay developer account (free) at developer.ebay.com, then production keys
and a user token. Put them in ebay-credentials.json beside this script:

    {"dev_id": "...", "app_id": "...", "cert_id": "...", "token": "..."}

or in the environment as EBAY_DEV_ID / EBAY_APP_ID / EBAY_CERT_ID /
EBAY_TOKEN. The file is gitignored: a user token is a password.

The URLs come back into photos/eps-urls.json, and make_ebay_csv.py prefers
them over the Pages ones. eBay purges a picture that no listing uses within
about 30 days, so upload when you are ready to list rather than in advance.
"""

import argparse
import json
import os
import re
import sys
import urllib.error
import urllib.request

from openpyxl import load_workbook

import inuse

WORKBOOK = "Card Run HQ - Master.xlsx"
CREDS = "ebay-credentials.json"
URLS = os.path.join("photos", "eps-urls.json")
PHOTOS = "photos"

ENDPOINT = "https://api.ebay.com/ws/api.dll"
COMPAT = "967"
SITE_ID = "0"                      # US
BOUNDARY = "MIME_boundary_cardrun"

FIELDS = ("dev_id", "app_id", "cert_id", "token")


def load_creds(path=CREDS):
    """Credentials from the file, or the environment, or a clear complaint."""
    got = {}
    if os.path.exists(path):
        with open(path, encoding="utf-8") as fh:
            got.update(json.load(fh))
    for f in FIELDS:
        got.setdefault(f, os.environ.get("EBAY_" + f.upper(), ""))
    missing = [f for f in FIELDS if not str(got.get(f) or "").strip()]
    if missing:
        raise SystemExit(
            "missing eBay credential(s): %s\n"
            "Put them in %s as {\"dev_id\": ..., \"app_id\": ..., "
            "\"cert_id\": ..., \"token\": ...}\n"
            "or set EBAY_DEV_ID / EBAY_APP_ID / EBAY_CERT_ID / EBAY_TOKEN.\n"
            "Get them from developer.ebay.com -> your keys (production)."
            % (", ".join(missing), path))
    return got


def payload(name, token):
    """The XML half of the request. Kept separate so it can be read."""
    return (
        '<?xml version="1.0" encoding="utf-8"?>'
        '<UploadSiteHostedPicturesRequest '
        'xmlns="urn:ebay:apis:eBLBaseComponents">'
        '<PictureName>%s</PictureName>'
        '<RequesterCredentials><eBayAuthToken>%s</eBayAuthToken>'
        '</RequesterCredentials>'
        '</UploadSiteHostedPicturesRequest>' % (xml_escape(name), token))


def xml_escape(s):
    return (str(s).replace("&", "&amp;").replace("<", "&lt;")
            .replace(">", "&gt;"))


def multipart(xml, image_bytes, filename):
    """eBay wants the XML first and the image second, in that order.

    Reversed, the call fails with an unhelpful error about the payload, so
    the order is the point of this function existing.
    """
    b = ("--%s" % BOUNDARY).encode()
    parts = [
        b,
        b'Content-Disposition: form-data; name="XML Payload"',
        b'Content-Type: text/xml;charset=utf-8',
        b'',
        xml.encode("utf-8"),
        b,
        ('Content-Disposition: form-data; name="dummy"; filename="%s"'
         % filename).encode(),
        b'Content-Type: application/octet-stream',
        b'',
        image_bytes,
        ("--%s--" % BOUNDARY).encode(),
        b'',
    ]
    return b"\r\n".join(parts)


def full_url(xml_text):
    """The hosted URL out of a response, or None with the error eBay gave.

    eBay answers HTTP 200 for failures too, with the reason in the body, so
    "it did not raise" is not the same as "it worked".
    """
    m = re.search(r"<FullURL>(.*?)</FullURL>", xml_text, re.S)
    if m:
        return m.group(1).strip(), None
    ack = re.search(r"<Ack>(.*?)</Ack>", xml_text, re.S)
    msg = (re.search(r"<LongMessage>(.*?)</LongMessage>", xml_text, re.S)
           or re.search(r"<ShortMessage>(.*?)</ShortMessage>", xml_text,
                        re.S))
    return None, "%s: %s" % (ack.group(1) if ack else "no FullURL",
                             msg.group(1).strip() if msg
                             else xml_text[:200])


def upload(path, name, creds, timeout=120):
    """One picture to eBay. Returns (url, error)."""
    with open(path, "rb") as fh:
        blob = fh.read()
    body = multipart(payload(name, creds["token"]), blob,
                     os.path.basename(path))
    req = urllib.request.Request(ENDPOINT, data=body, method="POST")
    req.add_header("Content-Type",
                   "multipart/form-data; boundary=%s" % BOUNDARY)
    req.add_header("X-EBAY-API-COMPATIBILITY-LEVEL", COMPAT)
    req.add_header("X-EBAY-API-DEV-NAME", creds["dev_id"])
    req.add_header("X-EBAY-API-APP-NAME", creds["app_id"])
    req.add_header("X-EBAY-API-CERT-NAME", creds["cert_id"])
    req.add_header("X-EBAY-API-CALL-NAME", "UploadSiteHostedPictures")
    req.add_header("X-EBAY-API-SITEID", SITE_ID)
    try:
        with urllib.request.urlopen(req, timeout=timeout) as r:
            return full_url(r.read().decode("utf-8", "replace"))
    except urllib.error.HTTPError as e:
        return None, "HTTP %s: %s" % (e.code,
                                      e.read()[:200].decode("utf-8", "replace"))
    except Exception as e:                      # network, DNS, timeout
        return None, str(e)[:200]


def ready(ws, g, want=None):
    """Cards that are going to be listed: Unlisted, with an ask price."""
    out = []
    for r in range(2, ws.max_row + 1):
        sku = ws.cell(row=r, column=g["SKU"]).value
        if not sku:
            continue
        if want and str(sku).upper() not in want:
            continue
        status = str(ws.cell(row=r, column=g["Status"]).value or "Unlisted")
        ask = ws.cell(row=r, column=g["Ask price"]).value
        if not want and (status.lower() != "unlisted" or ask in (None, "")):
            continue
        out.append(str(sku))
    return out


def photos_for(sku, folder=PHOTOS):
    got = []
    for suffix in (".jpg", "-back.jpg"):
        p = os.path.join(folder, sku + suffix)
        if os.path.exists(p):
            got.append(p)
    return got


def main():
    p = argparse.ArgumentParser(
        description="Upload listing photos to eBay Picture Services.")
    p.add_argument("--workbook", default=WORKBOOK)
    p.add_argument("--creds", default=CREDS)
    p.add_argument("--urls", default=URLS)
    p.add_argument("--photos", default=PHOTOS)
    p.add_argument("--sku", action="append", help="only these SKUs")
    p.add_argument("--check", action="store_true",
                   help="verify the credentials load, upload nothing")
    p.add_argument("--again", action="store_true",
                   help="re-upload cards that already have eBay URLs")
    p.add_argument("--go", action="store_true", help="actually upload")
    a = p.parse_args()

    creds = load_creds(a.creds)
    if a.check:
        print("credentials load: %s"
              % ", ".join("%s=%s..." % (f, str(creds[f])[:6]) for f in FIELDS))
        print("(this only proves they are present; eBay checks them on the "
              "first upload)")
        return

    inuse.refuse_if_open(a.workbook)
    ws = load_workbook(a.workbook, data_only=True)["Inventory"]
    hdr = [c.value for c in ws[1]]
    g = {n: i + 1 for i, n in enumerate(hdr) if n}
    want = {s.strip().upper() for v in (a.sku or []) for s in v.split(",")
            if s.strip()}
    skus = ready(ws, g, want or None)

    have = {}
    if os.path.exists(a.urls):
        with open(a.urls, encoding="utf-8") as fh:
            have = json.load(fh)

    todo = [(s, photos_for(s, a.photos)) for s in skus]
    todo = [(s, ps) for s, ps in todo if ps and (a.again or s not in have)]
    skipped = [s for s in skus if s in have and not a.again]
    nophoto = [s for s in skus if not photos_for(s, a.photos)]

    n_pics = sum(len(ps) for _s, ps in todo)
    print("%d card(s) ready to list; %d to upload (%d picture(s))"
          % (len(skus), len(todo), n_pics))
    if skipped:
        print("   %d already uploaded -- --again to redo them" % len(skipped))
    if nophoto:
        print("   %d with no photo on disk: %s"
              % (len(nophoto), ", ".join(nophoto[:6])))
    if not a.go:
        print("\nNothing uploaded. Add --go.")
        return

    done, failed = 0, []
    for sku, paths in todo:
        urls = []
        for i, path in enumerate(paths):
            side = "back" if "-back" in os.path.basename(path) else "front"
            url, err = upload(path, "%s %s" % (sku, side), creds)
            if err:
                failed.append((sku, side, err))
                break
            urls.append(url)
        if urls and len(urls) == len(paths):
            have[sku] = urls
            done += 1
            print("   %-9s %d picture(s)  %s" % (sku, len(urls), urls[0]))

    os.makedirs(os.path.dirname(a.urls) or ".", exist_ok=True)
    with open(a.urls, "w", encoding="utf-8") as fh:
        json.dump(have, fh, indent=2, sort_keys=True)
    print("\n%d card(s) uploaded; %s now holds %d card(s)"
          % (done, a.urls, len(have)))
    if failed:
        print("failed:")
        for sku, side, err in failed[:8]:
            print("   %-9s %-5s %s" % (sku, side, err[:90]))
        sys.exit(1)


if __name__ == "__main__":
    main()
