"""Tests for filing a scanned batch into the workbook.

Run: python -m pytest test_file_batch.py

The one that matters is the Review gate. A parallel is read off colour and
pattern rather than off print, and a sports market value has no free feed to
look up at all -- so those are judgements, and a judgement must not be able
to walk into a live eBay listing on its own. Any card with an unsure field is
filed as Review, and make_ebay_csv.py exports only Unlisted. If that ever
stops holding, somebody's guess ends up on a real listing.
"""

import json
import shutil
import subprocess
import sys

import pytest
from openpyxl import load_workbook

import file_batch as fb

CARD = {"player": "Shedeur Sanders", "year": 2025,
        "brand": "Panini Prizm Draft Picks", "insert": "Student Orientation",
        "parallel": "Gold Cracked Ice", "num": "8", "rc": True,
        "team": "Colorado Buffaloes", "league": "NCAA"}


@pytest.fixture(scope="module")
def blank(tmp_path_factory):
    """One real workbook, built once by the tool that makes the real one."""
    d = tmp_path_factory.mktemp("wb")
    out = d / "Card Run HQ - Master.xlsx"
    subprocess.run([sys.executable, "make_workbook.py", "--out", str(out)],
                   check=True, capture_output=True)
    return out


@pytest.fixture
def wb(blank, tmp_path):
    dest = tmp_path / "wb.xlsx"
    shutil.copy(blank, dest)
    return dest


def rows(path):
    ws = load_workbook(path)["Inventory"]
    hdr = [c.value for c in ws[1]]
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r[hdr.index("SKU")]:
            break
        out.append(dict(zip(hdr, r)))
    return out


def write_batch(tmp_path, cards, pairs=False):
    p = tmp_path / "batch.json"
    p.write_text(json.dumps({"pairs": pairs, "cards": cards}), encoding="utf-8")
    return p


# --- the gate ---------------------------------------------------------------

def test_a_certain_card_is_filed_unlisted(wb, tmp_path):
    fb.add_rows(str(wb), [dict(CARD)])
    assert rows(wb)[0]["Status"] == "Unlisted"


def test_an_unsure_card_is_held_back_as_review(wb, tmp_path):
    card = dict(CARD, unsure=["parallel"])
    fb.add_rows(str(wb), [card])
    got = rows(wb)[0]
    assert got["Status"] == "Review", "an unsure card must not be Unlisted"
    assert "parallel" in got["Notes"], "the note should say what to check"


def test_review_rows_do_not_reach_the_ebay_export(wb, tmp_path):
    """The gate itself: make_ebay_csv only ever takes Unlisted rows."""
    fb.add_rows(str(wb), [dict(CARD, unsure=["market"]), dict(CARD, player="Jonah Coleman")])
    r = subprocess.run([sys.executable, "make_ebay_csv.py", "--workbook", str(wb),
                        "--out", str(tmp_path / "out.csv")],
                       capture_output=True, text=True, cwd=".")
    body = r.stdout + r.stderr
    assert "Jonah Coleman" in body or "1 listing" in body, body
    assert "Shedeur" not in body.split("left blank")[0], \
        "the Review card reached the export:\n" + body


def test_settling_a_review_card_lets_it_export(wb, tmp_path):
    fb.add_rows(str(wb), [dict(CARD, unsure=["market"])])
    book = load_workbook(wb)
    ws = book["Inventory"]
    hdr = [c.value for c in ws[1]]
    ws.cell(row=2, column=hdr.index("Status") + 1, value="Unlisted")
    book.save(wb)
    r = subprocess.run([sys.executable, "make_ebay_csv.py", "--workbook", str(wb),
                        "--out", str(tmp_path / "out.csv")],
                       capture_output=True, text=True, cwd=".")
    assert "1 listing" in (r.stdout + r.stderr), r.stdout + r.stderr


# --- fields and SKUs --------------------------------------------------------

def test_every_field_lands_in_its_column(wb):
    fb.add_rows(str(wb), [dict(CARD)])
    got = rows(wb)[0]
    assert got["Player or card name"] == "Shedeur Sanders"
    assert got["Brand / set"] == "Panini Prizm Draft Picks"
    assert got["Insert set"] == "Student Orientation"
    assert got["Parallel"] == "Gold Cracked Ice"
    assert str(got["Card #"]) == "8"
    assert got["Team"] == "Colorado Buffaloes"
    assert got["League"] == "NCAA"
    assert got["RC"] == "Yes"


def test_defaults_fill_the_fields_a_scan_cannot_show(wb):
    fb.add_rows(str(wb), [dict(CARD)])
    got = rows(wb)[0]
    assert got["Sport or game"] == "Football"
    assert got["Category"] == "Sports"
    assert got["Card condition"] == "Near Mint or Better"
    assert got["Qty"] == 1


def test_skus_run_on_from_what_is_already_there(wb):
    fb.add_rows(str(wb), [dict(CARD)])
    second = fb.add_rows(str(wb), [dict(CARD, player="Jonah Coleman")])
    assert [e["sku"] for e in second] == ["CRH-0002"]
    assert [r["SKU"] for r in rows(wb)] == ["CRH-0001", "CRH-0002"]


def test_a_batch_is_numbered_in_the_order_given(wb):
    added = fb.add_rows(str(wb), [dict(CARD), dict(CARD, player="Jonah Coleman"),
                                  dict(CARD, player="Travis Hunter")])
    assert [e["sku"] for e in added] == ["CRH-0001", "CRH-0002", "CRH-0003"]
    assert [r["Player or card name"] for r in rows(wb)] == \
        ["Shedeur Sanders", "Jonah Coleman", "Travis Hunter"]


def test_dry_run_changes_nothing(wb):
    fb.add_rows(str(wb), [dict(CARD)], dry_run=True)
    assert rows(wb) == []


# --- a bad batch file is loud, not quietly wrong ----------------------------

def test_a_typo_in_a_field_name_is_rejected(tmp_path):
    p = write_batch(tmp_path, [dict(CARD, paralel="Gold")])
    with pytest.raises(SystemExit) as e:
        fb.load_batch(str(p))
    assert "paralel" in str(e.value)


def test_a_card_with_no_player_is_rejected(tmp_path):
    p = write_batch(tmp_path, [{"year": 2025}])
    with pytest.raises(SystemExit) as e:
        fb.load_batch(str(p))
    assert "player" in str(e.value)


def test_marking_a_field_unsure_that_does_not_exist_is_rejected(tmp_path):
    p = write_batch(tmp_path, [dict(CARD, unsure=["colour"])])
    with pytest.raises(SystemExit):
        fb.load_batch(str(p))


def test_an_empty_batch_is_rejected(tmp_path):
    p = write_batch(tmp_path, [])
    with pytest.raises(SystemExit):
        fb.load_batch(str(p))


def test_a_bare_list_of_cards_is_accepted(tmp_path):
    p = tmp_path / "b.json"
    p.write_text(json.dumps([dict(CARD)]), encoding="utf-8")
    pairs, cards = fb.load_batch(str(p))
    assert pairs is False and len(cards) == 1


def test_pairs_comes_through(tmp_path):
    p = write_batch(tmp_path, [dict(CARD)], pairs=True)
    pairs, _ = fb.load_batch(str(p))
    assert pairs is True
