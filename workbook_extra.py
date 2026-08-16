"""The tabs that turn the workbook from a card list into the whole record.

Imported by make_workbook.py. Kept apart from it only because that file was
already 500 lines of Inventory and eBay codes, and these are a different
subject: where the money went, what the photos are, and what is wrong.

    Money in   -> Purchases   every buy, with its receipt
                  Expenses    supplies, postage, anything deductible
    Money out  -> Sales       (in make_workbook.py)
    The stock  -> Inventory   (in make_workbook.py)
                  Photos      which pictures exist, with the pictures in it
    The truth  -> Summary     what it all adds up to
                  Audit       what is wrong before it costs you

Every tab is formulas over the raw sheets, so nothing here has to be kept up
to date by hand -- type a purchase in and the summary moves.
"""

from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

MONEY = '"$"#,##0.00'
PCT = "0.0%"

PURCHASE_ROWS = 400
EXPENSE_ROWS = 300
PHOTO_ROWS = 400

# ------------------------------------------------------------------ Purchases
# Everything bought, whether it was a sealed box, a lot off eBay, a single at
# a show, or a stack of toploaders. The Box log stays for what came OUT of a
# box; this is what went out of your pocket, and it is what a receipt matches.
PURCHASE_COLS = [
    ("Purchase ID", 13), ("Date", 11), ("Vendor / store", 22),
    ("Bought", 12), ("What", 40), ("Type", 15), ("Lot ID", 11),
    ("Qty", 6), ("Unit price", 11), ("Subtotal", 11), ("Tax", 9),
    ("Shipping", 10), ("Total paid", 12), ("Paid with", 14),
    ("Order / receipt #", 20), ("Receipt file", 24), ("Notes", 34),
]

PURCHASE_TYPES = ["Sealed box", "Sealed case", "Pack", "Single card",
                  "Bulk lot", "Supplies", "Postage", "Other"]
BOUGHT_WHERE = ["In store", "Online", "Show", "Private"]
PAID_WITH = ["Card", "Cash", "PayPal", "Gift card", "Other"]


def build_purchases(wb, head, note, dv, NOTEFILL, BOX, google=False):
    ws = wb.create_sheet("Purchases")
    head(ws, PURCHASE_COLS)
    c = {n: get_column_letter(i) for i, (n, _w) in enumerate(PURCHASE_COLS, 1)}

    if google:
        ws["%s2" % c["Subtotal"]] = (
            '=ARRAYFORMULA(IF({q}2:{q}="","",{q}2:{q}*{u}2:{u}))'
            .format(q=c["Qty"], u=c["Unit price"]))
        ws["%s2" % c["Total paid"]] = (
            '=ARRAYFORMULA(IF({s}2:{s}="","",N({s}2:{s})+N({t}2:{t})+N({sh}2:{sh})))'
            .format(s=c["Subtotal"], t=c["Tax"], sh=c["Shipping"]))

    for r in range(2, PURCHASE_ROWS + 2):
        if google:
            for col in ("Unit price", "Subtotal", "Tax", "Shipping", "Total paid"):
                ws["%s%d" % (c[col], r)].number_format = MONEY
            ws["%s%d" % (c["Date"], r)].number_format = "yyyy-mm-dd"
            continue
        ws["%s%d" % (c["Subtotal"], r)] = (
            '=IF(N({q}{r})=0,"",{q}{r}*{u}{r})'
            .format(q=c["Qty"], u=c["Unit price"], r=r))
        ws["%s%d" % (c["Total paid"], r)] = (
            '=IF(AND(N({s}{r})=0,N({t}{r})=0,N({sh}{r})=0),"",'
            'N({s}{r})+N({t}{r})+N({sh}{r}))'
            .format(s=c["Subtotal"], t=c["Tax"], sh=c["Shipping"], r=r))
        for col in ("Unit price", "Subtotal", "Tax", "Shipping", "Total paid"):
            ws["%s%d" % (c[col], r)].number_format = MONEY
        ws["%s%d" % (c["Date"], r)].number_format = "yyyy-mm-dd"

    dv(ws, PURCHASE_TYPES, "%s2:%s%d" % (c["Type"], c["Type"], PURCHASE_ROWS + 1))
    dv(ws, BOUGHT_WHERE, "%s2:%s%d" % (c["Bought"], c["Bought"], PURCHASE_ROWS + 1))
    dv(ws, PAID_WITH, "%s2:%s%d" % (c["Paid with"], c["Paid with"], PURCHASE_ROWS + 1))

    r = PURCHASE_ROWS + 3
    ws.cell(row=r, column=1, value="Everything you paid for, in one place.")\
        .font = Font(bold=True)
    note(ws, "A%d" % (r + 1),
         "Subtotal and Total paid work themselves out -- type Qty, Unit price, "
         "Tax and Shipping.\n"
         "Lot ID ties a purchase to the cards that came out of it: put the same "
         "code on the Box log row and on every Inventory row from that box, and "
         "cost per card stops being a guess.\n"
         "Receipt file is just the filename you saved the photo or PDF under, "
         "e.g. receipts/2026-08-15-target.jpg. Keep the file; the tax year will "
         "want it and eBay will want it if a buyer disputes a sale.\n"
         "Supplies and postage go here too if you bought them as part of a "
         "trip; anything ongoing belongs on Expenses.")
    ws["A%d" % (r + 1)].fill = NOTEFILL
    ws.merge_cells("A%d:H%d" % (r + 1, r + 6))
    return ws


# ------------------------------------------------------------------- Expenses
EXPENSE_COLS = [
    ("Date", 11), ("Category", 18), ("What", 40), ("Vendor", 22),
    ("Amount", 11), ("Paid with", 14), ("Receipt file", 24),
    ("Deductible", 12), ("Notes", 34),
]

EXPENSE_CATS = ["Supplies", "Shipping", "Postage", "Fees", "Software",
                "Subscription", "Mileage", "Equipment", "Other"]


def build_expenses(wb, head, note, dv, NOTEFILL):
    ws = wb.create_sheet("Expenses")
    head(ws, EXPENSE_COLS)
    c = {n: get_column_letter(i) for i, (n, _w) in enumerate(EXPENSE_COLS, 1)}

    for r in range(2, EXPENSE_ROWS + 2):
        ws["%s%d" % (c["Amount"], r)].number_format = MONEY
        ws["%s%d" % (c["Date"], r)].number_format = "yyyy-mm-dd"

    dv(ws, EXPENSE_CATS, "%s2:%s%d" % (c["Category"], c["Category"], EXPENSE_ROWS + 1))
    dv(ws, ["Yes", "No"], "%s2:%s%d" % (c["Deductible"], c["Deductible"], EXPENSE_ROWS + 1))

    r = EXPENSE_ROWS + 3
    ws.cell(row=r, column=1, value="The costs that are not a card.")\
        .font = Font(bold=True)
    note(ws, "A%d" % (r + 1),
         "Toploaders, sleeves, bubble mailers, stamps, the scanner, a "
         "subscription. These are what turn a gross profit into a real one, "
         "and they are the first thing forgotten.\n"
         "Anything bought as part of a buying trip can go on Purchases "
         "instead -- the Summary adds both, so it is counted once either way.\n"
         "Deductible is your call and worth marking as you go rather than "
         "reconstructing it in April.")
    ws["A%d" % (r + 1)].fill = NOTEFILL
    ws.merge_cells("A%d:E%d" % (r + 1, r + 5))
    return ws


# --------------------------------------------------------------------- Photos
# The picture side of inventory control. embed_photos.py fills this in and
# drops the actual thumbnails into it, so you can scroll the stock and see it
# rather than reading SKUs.
PHOTO_COLS = [
    ("SKU", 12), ("Card", 30), ("Picture", 22), ("Front", 22), ("Back", 22),
    ("Extras", 20), ("How many", 10), ("Picture URL for eBay", 52),
    ("Checked", 10), ("Notes", 30),
]


PAGES = "https://pthaiop587.github.io/tcg-scout/photos"


def build_photos(wb, head, note, dv, NOTEFILL, google=False):
    ws = wb.create_sheet("Photos")
    head(ws, PHOTO_COLS)
    c = {n: get_column_letter(i) for i, (n, _w) in enumerate(PHOTO_COLS, 1)}
    dv(ws, ["Yes", "No"], "%s2:%s%d" % (c["Checked"], c["Checked"], PHOTO_ROWS + 1))

    if google:
        # Google Sheets draws a picture from a formula, so the tab fills itself:
        # type a SKU and the photograph appears, live, straight off the
        # published site. No embedding, no file size, and it can never be a
        # stale copy of the picture eBay is actually serving.
        ws["L1"] = PAGES + "/"          # the base address, written once
        for r in range(2, min(PHOTO_ROWS, 40) + 2):
            sku = "%s%d" % (c["SKU"], r)
            url = '$L$1&%s&".jpg"' % sku
            back = '$L$1&%s&"-back.jpg"' % sku
            ws["%s%d" % (c["Picture"], r)] = (
                '=IF({s}="","",IFERROR(IMAGE({u},1),"no photo yet"))'
                .format(s=sku, u=url))
            ws["%s%d" % (c["Picture URL for eBay"], r)] = (
                '=IF({s}="","",{u}&IF({b}="","","|"&{bu}))'
                .format(s=sku, u=url, b="%s%d" % (c["Back"], r), bu=back))
            ws["%s%d" % (c["Front"], r)] = (
                '=IF({s}="","",{s}&".jpg")'.format(s=sku))
            ws.row_dimensions[r].height = 96

    r = PHOTO_ROWS + 3
    ws.cell(row=r, column=1, value="What each card actually has a picture of.")\
        .font = Font(bold=True)
    note(ws, "A%d" % (r + 1),
         "In the Google copy this tab fills itself: type a SKU in column A and "
         "the photograph appears, live, off the published site. In the Excel "
         "copy it is filled in by  python embed_photos.py  -- it reads the photos folder, "
         "writes a row per SKU and puts the thumbnail in the Picture column, so "
         "this tab is the visual stock check. Re-run it whenever you add "
         "photos; it rewrites the tab from what is on disk.\n"
         "The URL column is the address eBay will fetch. It only resolves once "
         "the photos have been pushed and published -- the workbook carries "
         "links, never the image files themselves, because eBay needs a public "
         "https address rather than an embedded picture.\n"
         "A card with no row here has no photograph yet.")
    ws["A%d" % (r + 1)].fill = NOTEFILL
    ws.merge_cells("A%d:F%d" % (r + 1, r + 7))
    return ws


# -------------------------------------------------------------------- Summary
def build_summary(wb, head, note, TITLEFONT, SUBFILL, HEADFONT, NOTEFILL, INV_ROWS):
    """Everything as one number each, all of it formulas over the other tabs."""
    ws = wb.create_sheet("Summary", 1)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 52

    ws["A1"] = "Card Run HQ — where it stands"
    ws["A1"].font = TITLEFONT
    ws.merge_cells("A1:C1")

    inv = "Inventory!"
    end = INV_ROWS + 1
    rows = [
        ("MONEY OUT", None, None),
        ("Spent on stock", "=N(SUM(Purchases!M2:M%d))" % (PURCHASE_ROWS + 1),
         "Every row on Purchases, tax and shipping included."),
        ("Spent on everything else", "=N(SUM(Expenses!E2:E%d))" % (EXPENSE_ROWS + 1),
         "Supplies, postage, subscriptions — the Expenses tab."),
        ("Total out of pocket", "={Spent on stock}+{Spent on everything else}",
         "The real number, not the card cost."),

        ("THE STOCK", None, None),
        ("Cards on hand", '=N(SUMIFS({i}X2:X{e},{i}B2:B{e},"Unlisted"))+'
                          'N(SUMIFS({i}X2:X{e},{i}B2:B{e},"Listed"))+'
                          'N(SUMIFS({i}X2:X{e},{i}B2:B{e},"Review"))'.format(i=inv, e=end),
         "Quantity, not rows — Unlisted, Listed and Review."),
        ("What they cost", '=N(SUMPRODUCT({i}X2:X{e},{i}Y2:Y{e}))'.format(i=inv, e=end),
         "Qty times cost each, across everything logged."),
        ("What they are worth", '=N(SUMPRODUCT({i}X2:X{e},{i}Z2:Z{e}))'.format(i=inv, e=end),
         "Qty times market value. A guess for sports; there is no free feed."),
        ("Unrealised gain", "={What they are worth}-{What they cost}",
         "Worth minus cost. Not money until it sells."),

        ("WAITING ON YOU", None, None),
        ("Held for review", '=COUNTIF({i}B2:B{e},"Review")'.format(i=inv, e=end),
         "Something on the card was uncertain. These cannot be exported."),
        ("Ready to list", '=COUNTIF({i}B2:B{e},"Unlisted")'.format(i=inv, e=end),
         "What make_ebay_csv.py will pick up next run."),
        ("Listed now", '=COUNTIF({i}B2:B{e},"Listed")'.format(i=inv, e=end), ""),

        ("MONEY IN", None, None),
        ("Sold", "=COUNTA(Sales!B2:B301)", "Rows on the Sales tab."),
        ("Gross", "=N(SUM(Sales!G2:G301))", "Sale price plus what they paid to ship."),
        ("Fees and postage", "=N(SUM(Sales!H2:H301))+N(SUM(Sales!I2:I301))+"
                             "N(SUM(Sales!J2:J301))", ""),
        ("Net received", "=N(SUM(Sales!K2:K301))", ""),
        ("Profit on what sold", "=N(SUM(Sales!M2:M301))",
         "Net minus what those cards cost."),

        ("THE ANSWER", None, None),
        ("Profit against everything spent",
         "={Profit on what sold}-{Total out of pocket}",
         "Sold profit minus every pound out. Negative until the stock turns over."),
        ("Return on money spent",
         '=IF(N({Total out of pocket})=0,"",'
         '{Profit on what sold}/{Total out of pocket})',
         "Profit over spend. This is the number that says whether it works."),
    ]

    # Two passes. A formula that refers to another row of this sheet names it
    # -- "{Total out of pocket}" -- and the row number is filled in afterwards
    # from where the label actually landed. Writing B3+B4 by hand meant
    # recounting every time a section heading moved, and it was wrong: the
    # unrealised gain was reading the card count instead of the cost.
    at = {}
    rr = 2
    for label, formula, _why in rows:
        if formula is not None:
            at[label] = rr
        rr += 1

    r = 2
    money_rows, pct_rows = [], []
    for label, formula, why in rows:
        if formula is not None:
            for key, where in at.items():
                formula = formula.replace("{%s}" % key, "B%d" % where)
        if formula is None:
            cell = ws.cell(row=r, column=1, value=label)
            cell.fill = SUBFILL
            cell.font = HEADFONT
            for col in ("B", "C"):
                ws["%s%d" % (col, r)].fill = SUBFILL
        else:
            ws.cell(row=r, column=1, value=label)
            ws.cell(row=r, column=2, value=formula)
            if why:
                cell = ws.cell(row=r, column=3, value=why)
                cell.font = Font(size=9, color="55607A")
                cell.alignment = Alignment(wrap_text=True, vertical="center")
            if "Return on" in label:
                pct_rows.append(r)
            elif any(k in label for k in ("Spent", "cost", "worth", "gain",
                                          "Gross", "Fees", "Net", "Profit", "Total")):
                money_rows.append(r)
        r += 1

    for rr in money_rows:
        ws["B%d" % rr].number_format = MONEY
    for rr in pct_rows:
        ws["B%d" % rr].number_format = PCT
    for rr in range(2, r):
        ws["B%d" % rr].font = Font(bold=True)

    note(ws, "A%d" % (r + 1),
         "Every number here is a formula over the other tabs. Nothing on this "
         "sheet is typed, so it cannot go stale — if a figure looks wrong, the "
         "row it came from is wrong.")
    ws["A%d" % (r + 1)].fill = NOTEFILL
    ws.merge_cells("A%d:C%d" % (r + 1, r + 3))
    return ws


# ---------------------------------------------------------------------- Audit
def build_audit(wb, head, note, TITLEFONT, SUBFILL, HEADFONT, NOTEFILL, INV_ROWS):
    """The things that quietly cost money, counted."""
    ws = wb.create_sheet("Audit")
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 74

    ws["A1"] = "Audit — what to fix before it costs you"
    ws["A1"].font = TITLEFONT
    ws.merge_cells("A1:C1")

    i, e = "Inventory!", INV_ROWS + 1
    checks = [
        ("STOPS A LISTING", None, None),
        ("Cards held for review",
         '=COUNTIF({i}B2:B{e},"Review")'.format(i=i, e=e),
         "Something was uncertain when it was filed — usually the parallel or "
         "the value. make_ebay_csv.py skips these on purpose. Settle the CHECK "
         "note in Notes, then set Status to Unlisted."),
        ("Ready to list but worth nothing",
         '=SUMPRODUCT(({i}B2:B{e}="Unlisted")*(N({i}Z2:Z{e})=0))'.format(i=i, e=e),
         "It will export at a price of nothing. Put a market value on it, or "
         "move it to Review until you have one."),
        ("Ready to list with no photo",
         '=SUMPRODUCT(({i}B2:B{e}="Unlisted")*'
         '(COUNTIF(Photos!A2:A{p},{i}A2:A{e})=0))'.format(i=i, e=e, p=PHOTO_ROWS + 1),
         "eBay needs a picture. Run embed_photos.py after adding photos so this "
         "tab knows about them."),
        ("Titles over 80 characters",
         '=SUMPRODUCT((N({i}AF2:AF{e})>80)*1)'.format(i=i, e=e),
         "eBay truncates at 80. The Len column on Inventory shows each one."),

        ("COSTS YOU MONEY", None, None),
        ("Cards with no cost recorded",
         '=SUMPRODUCT((LEN({i}A2:A{e})>0)*(N({i}Y2:Y{e})=0))'.format(i=i, e=e),
         "Profit on a sale is meaningless without it, and so is the Summary. "
         "If a card came out of a box, put the box's Lot ID on it and use cost "
         "per card from the Box log."),
        ("Cards with no Lot ID",
         '=SUMPRODUCT((LEN({i}A2:A{e})>0)*(LEN({i}E2:E{e})=0))'.format(i=i, e=e),
         "Nothing ties them back to a purchase, so you cannot tell which box "
         "paid for itself."),
        ("Purchases with no receipt file",
         '=SUMPRODUCT((LEN(Purchases!A2:A{p})>0)*(LEN(Purchases!P2:P{p})=0))'
         .format(p=PURCHASE_ROWS + 1),
         "The tax year will want them, and so will eBay if a buyer disputes."),
        ("Sales with no cost basis",
         '=SUMPRODUCT((LEN(Sales!B2:B301)>0)*(N(Sales!L2:L301)=0))',
         "Profit on those rows is overstated by exactly what the card cost."),

        ("PROBABLY A MISTAKE", None, None),
        ("Duplicate SKUs",
         '=SUMPRODUCT((LEN({i}A2:A{e})>0)*'
         '(COUNTIF({i}A2:A{e},{i}A2:A{e}&"")>1))'.format(i=i, e=e),
         "A SKU is meant to be one physical card. Two rows sharing one means "
         "photos and sales will attach to whichever Excel finds first."),
        ("Serial numbered with quantity over 1",
         '=SUMPRODUCT((LEN({i}O2:O{e})>0)*(N({i}X2:X{e})>1))'.format(i=i, e=e),
         "A numbered card is one of one at that number. Two of them is two "
         "different cards and wants two rows."),
        ("Graded with no cert number",
         '=SUMPRODUCT((LEN({i}T2:T{e})>0)*(LEN({i}V2:V{e})=0))'.format(i=i, e=e),
         "Buyers check the cert. eBay has a field for it."),
        ("Sold but still Unlisted or Listed",
         '=SUMPRODUCT((COUNTIF(Sales!B2:B301,{i}A2:A{e})>0)*'
         '({i}B2:B{e}<>"Sold"))'.format(i=i, e=e),
         "It is on the Sales tab but the stock still thinks you have it. That "
         "is how a card gets listed twice."),
    ]

    r = 3
    for label, formula, why in checks:
        if formula is None:
            cell = ws.cell(row=r, column=1, value=label)
            cell.fill = SUBFILL
            cell.font = HEADFONT
            for col in ("B", "C"):
                ws["%s%d" % (col, r)].fill = SUBFILL
        else:
            ws.cell(row=r, column=1, value=label)
            v = ws.cell(row=r, column=2, value=formula)
            v.font = Font(bold=True)
            v.alignment = Alignment(horizontal="center")
            cell = ws.cell(row=r, column=3, value=why)
            cell.font = Font(size=9, color="55607A")
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            ws.row_dimensions[r].height = 30
        r += 1

    note(ws, "A%d" % (r + 1),
         "Every count is a formula, so this tab is never out of date. A zero "
         "down the whole column means the workbook is clean.\n"
         "None of it blocks anything — it is a list of things that are cheap "
         "to fix now and expensive to find out about from a buyer.")
    ws["A%d" % (r + 1)].fill = NOTEFILL
    ws.merge_cells("A%d:C%d" % (r + 1, r + 3))
    return ws
