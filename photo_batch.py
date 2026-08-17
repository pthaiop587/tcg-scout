"""A folder of card photos in; the whole batch filed onto its SKUs.

    python photo_batch.py prep G:/Scans
    python photo_batch.py prep G:/Scans --backs        # sheet the backs too
    python photo_batch.py prep G:/Scans --singles      # fronts only, no backs
    python photo_batch.py file batch-photos.json       # onto cards already logged
    python photo_batch.py add  batch-photos.json       # a box that is not logged yet

Six cards was fine done by hand. Sixty is not: every photo has to be looked at
to know which card it is, and looking at them one at a time is the whole cost
of the job.

PREP does two things about that. It converts everything -- DNG included, which
a phone shooting RAW writes -- into working JPEGs with the rotation baked in.
Then it lays the FRONTS out in contact sheets, nine to a page with an index
number under each, so nine cards can be identified in one look instead of
nine. The backs are converted but not sheeted; they are only needed to settle
a card number, and most cards are already identifiable from the front.

It assumes the shots alternate front, back, front, back -- which is how a stack
gets photographed -- and prints the pairing so a slip is visible before
anything is filed. --singles says there are no backs.

--backs sheets the backs as well, numbered to match. That is for a box being
logged for the first time: a card NUMBER is printed on the back and nowhere
else, and prices.py needs it to find the card. For a box already in the
workbook the fronts are enough.

FILE takes the identifications back, matches each to a SKU, and files the
photos onto it. The match is by player and parallel against Inventory, because
that is what a front shows: a card number lives on the back and is not worth a
second photo per card just to confirm what the name already said. Anything
matching two cards, or none, is reported and skipped rather than guessed --
filing a picture onto the wrong SKU is a wrong picture in a listing.

ADD is for a box nobody has typed in yet. It creates the Inventory row first,
taking the SKU file_batch.py hands out, then files the photos onto it. What a
whole box shares -- the year, the set, the shop, the lot -- is written once in
a "defaults" block rather than sixty times, and anything a card is unsure of
still lands it on Review, so an uncertain card cannot reach an eBay export.
"""

import argparse
import glob
import json
import os
import re
import subprocess
import sys

from PIL import Image, ImageDraw, ImageOps

from openpyxl import load_workbook

WORKBOOK = "Card Run HQ - Master.xlsx"
# The helper scripts sit beside this one, and they all work against the folder
# the WORKBOOK is in -- photos/ is relative to it, and so is everything else.
# Resolving both explicitly means this can be run from anywhere, which matters
# because it usually is.
HERE = os.path.dirname(os.path.abspath(__file__))
WORK = "photos/_batch"          # converted working copies
SHEET_COLS, SHEET_ROWS = 3, 3   # nine to a contact sheet
CELL_W = 520                    # big enough to read a name band off
RAW = (".dng", ".jpg", ".jpeg", ".png", ".heic", ".tif", ".tiff", ".webp")


def norm(s):
    s = str(s or "").strip().lower()
    s = re.sub(r"[.'\u2019]", "", s)
    s = re.sub(r"[^\w\s]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


# --- prep -------------------------------------------------------------------

def convert(src, dest, long_edge=1700):
    im = ImageOps.exif_transpose(Image.open(src)).convert("RGB")
    im.thumbnail((long_edge, long_edge), Image.LANCZOS)
    im.save(dest, "JPEG", quality=90)
    return im.size


def card_box(im, pad=0.02):
    """Where the card is in the frame, so the sheet can show it big.

    A card photographed on a table is maybe a third of the picture; the rest is
    tabletop, and on a contact sheet that is two thirds of the space spent on
    nothing. The card is found the same way the parallel was measured -- by
    distance from the background colour, sampled from the frame's own edge --
    which works for a dark front and a white back alike.

    Only the CONTACT SHEET is cropped. What gets filed is the photo as shot.
    """
    try:
        import numpy as np
    except ImportError:
        return None
    a = np.asarray(im.convert("RGB").resize(
        (240, int(240 * im.size[1] / im.size[0])), Image.BILINEAR)
    ).astype(np.float32)
    h, w = a.shape[:2]
    b = max(3, min(h, w) // 20)
    ring = np.concatenate([a[:b].reshape(-1, 3), a[-b:].reshape(-1, 3),
                           a[:, :b].reshape(-1, 3), a[:, -b:].reshape(-1, 3)])
    bg = np.median(ring, axis=0)
    d = np.linalg.norm(a - bg, axis=2)
    mask = d > max(28.0, float(np.percentile(d, 82)))
    ys, xs = np.nonzero(mask)
    if len(xs) < mask.size * 0.02:
        return None
    sx, sy = im.size[0] / float(w), im.size[1] / float(h)
    px, py = im.size[0] * pad, im.size[1] * pad
    return (max(0, int(xs.min() * sx - px)), max(0, int(ys.min() * sy - py)),
            min(im.size[0], int(xs.max() * sx + px)),
            min(im.size[1], int(ys.max() * sy + py)))


def contact_sheet(paths, labels, out):
    """Fronts in a grid, numbered, so many can be read at once."""
    cell_h = int(CELL_W * 4032 / 3024.0)
    pad, strip = 10, 34
    W = SHEET_COLS * CELL_W + pad * (SHEET_COLS + 1)
    H = SHEET_ROWS * (cell_h + strip) + pad * (SHEET_ROWS + 1)
    sheet = Image.new("RGB", (W, H), (245, 245, 247))
    d = ImageDraw.Draw(sheet)

    for i, (p, label) in enumerate(zip(paths, labels)):
        r, c = divmod(i, SHEET_COLS)
        x = pad + c * (CELL_W + pad)
        y = pad + r * (cell_h + strip + pad)
        im = Image.open(p)
        box = card_box(im)
        if box and box[2] > box[0] and box[3] > box[1]:
            im = im.crop(box)
        im.thumbnail((CELL_W, cell_h), Image.LANCZOS)
        sheet.paste(im, (x + (CELL_W - im.size[0]) // 2,
                         y + (cell_h - im.size[1]) // 2))
        d.text((x + 6, y + cell_h + 6), label, fill=(20, 20, 30))
    sheet.save(out, "JPEG", quality=90)
    return sheet.size


def do_prep(a):
    files = []
    for ext in RAW:
        files.extend(glob.glob(os.path.join(a.src, "*" + ext)))
        files.extend(glob.glob(os.path.join(a.src, "*" + ext.upper())))
    files = sorted(set(files), key=lambda p: os.path.basename(p).lower())
    if a.skip:
        skip = {s.strip().lower() for s in a.skip.split(",")}
        files = [f for f in files
                 if os.path.splitext(os.path.basename(f))[0].lower() not in skip]
    if not files:
        sys.exit("no photos in %s" % a.src)

    os.makedirs(a.work, exist_ok=True)
    print("%d photo(s) in %s" % (len(files), a.src))

    converted = []
    for p in files:
        base = os.path.splitext(os.path.basename(p))[0]
        dest = os.path.join(a.work, base + ".jpg")
        size = convert(p, dest)
        converted.append((base, dest))
    print("converted into %s at %s" % (a.work, "%dx%d" % size))

    if a.singles:
        pairs = [(b, None) for b, _ in converted]
    else:
        if len(converted) % 2:
            print("\n!! %d photos is an odd number. If these are front/back "
                  "pairs one is missing,\n   and every pair after the gap is "
                  "wrong. Check before filing." % len(converted))
        pairs = []
        for i in range(0, len(converted), 2):
            f = converted[i][0]
            b = converted[i + 1][0] if i + 1 < len(converted) else None
            pairs.append((f, b))

    print("\n%d card(s), paired as:" % len(pairs))
    for i, (f, b) in enumerate(pairs, 1):
        print("   %2d  %-14s %s" % (i, f, b or "(no back)"))

    per = SHEET_COLS * SHEET_ROWS

    def sheets_of(which, prefix):
        paths, labels, out = [], [], []
        for i, (f, b) in enumerate(pairs, 1):
            name = f if which == "front" else b
            if not name:
                continue
            paths.append(os.path.join(a.work, name + ".jpg"))
            labels.append("%d  %s" % (i, name))
        for n, i in enumerate(range(0, len(paths), per), 1):
            dest = os.path.join(a.work, "%s-%02d.jpg" % (prefix, n))
            contact_sheet(paths[i:i + per], labels[i:i + per], dest)
            out.append(dest)
        return out

    made = sheets_of("front", "sheet")
    print("\n%d contact sheet(s) of the fronts:" % len(made))
    for x in made:
        print("   %s" % x)

    if a.backs:
        backs = sheets_of("back", "back")
        print("\n%d of the backs -- the card number is only printed there:"
              % len(backs))
        for x in backs:
            print("   %s" % x)

    defaults = {}
    for pair in (a.defaults or "").split(","):
        if "=" in pair:
            k, v = pair.split("=", 1)
            defaults[k.strip()] = v.strip()

    stub = [{"n": i, "front": f, "back": b, "name": "", "parallel": ""}
            for i, (f, b) in enumerate(pairs, 1)]
    body = {"work": a.work, "cards": stub}
    if defaults:
        body = {"work": a.work, "defaults": defaults, "cards": stub}
    with open(a.out, "w", encoding="utf-8") as fh:
        json.dump(body, fh, indent=2)
    print("\nwrote %s -- fill in name and parallel from the sheets, then:"
          % a.out)
    print("   python photo_batch.py file %s" % a.out)
    return 0


# --- file -------------------------------------------------------------------

def helper(workbook, script, *args):
    """Run a sibling script against the workbook's own folder."""
    return subprocess.run(
        [sys.executable, os.path.join(HERE, script)] + list(args),
        capture_output=True, text=True,
        cwd=os.path.dirname(os.path.abspath(workbook)) or ".")


def inventory(path):
    ws = load_workbook(path)["Inventory"]
    hdr = [c.value for c in ws[1]]
    g = {n: i for i, n in enumerate(hdr) if n}
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row = list(row) + [None] * (len(hdr) - len(row))
        if not row[g["SKU"]]:
            continue
        out.append({"sku": row[g["SKU"]],
                    "name": row[g["Player or card name"]],
                    "parallel": row[g["Parallel"]],
                    "num": row[g["Card #"]],
                    "insert": row[g.get("Insert set", 0)]})
    return out


def match(card, inv):
    """Every Inventory row a photographed card could be.

    A named SKU wins outright. That is not laziness -- two of the SAME card
    cannot be told apart by anything printed on either of them, and this
    inventory has a pair of Saquon Barkley #190 Silvers. No amount of looking
    at the photo resolves that, so somebody has to say which one, and this is
    where they say it."""
    if card.get("sku"):
        want = str(card["sku"]).strip().upper()
        return [c for c in inv if str(c["sku"] or "").strip().upper() == want]

    n, p = norm(card.get("name")), norm(card.get("parallel"))
    hits = [c for c in inv if norm(c["name"]) == n]
    if p:
        exact = [c for c in hits if norm(c["parallel"]) == p]
        if exact:
            hits = exact
    # The insert set, which the front announces in large letters -- STUDENT
    # ORIENTATION, INSTANT IMPACT, NEW RECRUITS. Two Shedeur Sanders Gold Ices
    # differ by nothing else, and it is far easier to read off a photo than a
    # card number, which lives on the back.
    ins = norm(card.get("insert"))
    if ins:
        exact = [c for c in hits if norm(c["insert"]) == ins]
        if exact:
            hits = exact
    elif len(hits) > 1:
        # no insert named: prefer the plain base card over an insert
        base = [c for c in hits if not norm(c["insert"])]
        if base:
            hits = base
    if card.get("num"):
        want = str(card["num"]).strip().lstrip("#")
        byn = [c for c in hits if str(c["num"] or "").strip() == want]
        if byn:
            hits = byn
    return hits


def do_file(a):
    with open(a.batch, encoding="utf-8") as fh:
        data = json.load(fh)
    work = data.get("work", WORK)
    cards = data["cards"]
    inv = inventory(a.workbook)
    print("%d card(s) in the batch, %d in Inventory" % (len(cards), len(inv)))

    plan, trouble = [], []
    for c in cards:
        if not str(c.get("name") or "").strip():
            trouble.append((c, "no name filled in"))
            continue
        hits = match(c, inv)
        if not hits:
            trouble.append((c, "nothing in Inventory matches"))
        elif len(hits) > 1:
            trouble.append((c, "matches %s" % ", ".join(h["sku"] for h in hits)))
        else:
            plan.append((hits[0], c))

    print("\n%d matched:" % len(plan))
    for row, c in plan:
        print("   %-9s %-22s %-10s #%-6s <- %s" %
              (row["sku"], row["name"], row["parallel"] or "", row["num"] or "",
               c["front"]))
    if trouble:
        print("\n%d NOT filed:" % len(trouble))
        for c, why in trouble:
            print("   %-3s %-22s %-12s %s"
                  % (c.get("n", "?"), c.get("name") or "(blank)",
                     c.get("parallel") or "", why))

    if not a.go:
        print("\nNothing filed. Add --go.")
        return 0 if not trouble else 1

    done = 0
    for row, c in plan:
        shots = [os.path.join(work, c["front"] + ".jpg")]
        if c.get("back"):
            shots.append(os.path.join(work, c["back"] + ".jpg"))
        shots = [s for s in shots if os.path.exists(s)]
        if not shots:
            print("   %s: no converted photo found" % row["sku"])
            continue
        r = helper(a.workbook, "add_photos.py", "--sku", row["sku"], *shots)
        if r.returncode:
            print("   %s FAILED\n%s" % (row["sku"],
                                        (r.stdout + r.stderr)[:200]))
        else:
            done += 1
    print("\nfiled %d card(s)" % done)

    helper(a.workbook, "link_photos.py", "--go")
    helper(a.workbook, "sport_tabs.py")
    print("links refreshed and the game tabs rebuilt.")
    return 0


def do_add(a):
    """A box nobody has typed in: make the rows, then file the photos."""
    import file_batch

    with open(a.batch, encoding="utf-8") as fh:
        data = json.load(fh)
    work = data.get("work", WORK)
    defaults = data.get("defaults", {})
    cards = data["cards"]

    named = [c for c in cards if str(c.get("name") or "").strip()]
    if len(named) != len(cards):
        print("%d card(s) have no name and will be skipped"
              % (len(cards) - len(named)))
    if not named:
        sys.exit("nothing to add")

    # what the whole box shares, written once
    rows, shots = [], []
    for c in named:
        # Everything the card says overrides the box default -- a list of
        # which fields may do so was a whitelist, and it silently dropped any
        # field that happened not to be on it. A Pokemon card in a football
        # box kept sport=Football and filed as Sports. file_batch.py rejects
        # a field it does not know, loudly, which is the better guard.
        rec = dict(defaults)
        for k, v in c.items():
            if k in ("n", "front", "back", "sku") or v in (None, ""):
                continue
            rec[k] = v
        rec["player"] = rec.pop("name")
        rows.append(rec)
        shots.append([os.path.join(work, c[w] + ".jpg")
                      for w in ("front", "back") if c.get(w)])

    # a card already in the workbook is worth saying so about -- a second copy
    # is normal out of a second box, but a repeat of the SAME batch is not
    inv = inventory(a.workbook)
    for r in rows:
        same = [i for i in inv
                if norm(i["name"]) == norm(r.get("player"))
                and norm(i["parallel"]) == norm(r.get("parallel"))
                and str(i["num"] or "").strip() == str(r.get("num") or "").strip()]
        if same:
            print("   note: %s %s #%s is already in the workbook as %s"
                  % (r.get("player"), r.get("parallel") or "", r.get("num") or "",
                     ", ".join(x["sku"] for x in same)))

    print("\n%d card(s) to add" % len(rows))
    if defaults:
        print("shared: " + ", ".join("%s=%s" % kv for kv in sorted(defaults.items())))
    for r in rows[:6]:
        print("   %-22s %-10s #%-8s %s" % (r.get("player"), r.get("parallel") or "",
                                           r.get("num") or "",
                                           r.get("insert") or ""))
    if len(rows) > 6:
        print("   ... and %d more" % (len(rows) - 6))

    if not a.go:
        print("\nNothing added. Add --go.")
        return 0

    added = file_batch.add_rows(a.workbook, rows)
    print("\nadded %d row(s): %s .. %s"
          % (len(added), added[0]["sku"], added[-1]["sku"]))

    filed = 0
    for entry, pics in zip(added, shots):
        pics = [p for p in pics if os.path.exists(p)]
        if not pics:
            continue
        r = helper(a.workbook, "add_photos.py", "--sku", entry["sku"], *pics)
        if r.returncode:
            print("   %s photo FAILED: %s" % (entry["sku"],
                                              (r.stdout + r.stderr)[:150]))
        else:
            filed += 1
    print("photos filed for %d card(s)" % filed)

    for script in ("autofill.py", "link_photos.py", "sport_tabs.py"):
        helper(a.workbook, script, *(["--go"] if script != "sport_tabs.py"
                                     else []))
    print("SKUs, categories, photo links and the game tabs are up to date.")

    review = [e for e in added if e.get("unsure")]
    if review:
        print("\n%d card(s) filed as Review -- settle the CHECK note in Notes, "
              "then set Status to Unlisted:" % len(review))
        for e in review:
            print("   %-9s %-22s %s" % (e["sku"], e["player"],
                                        ", ".join(e["unsure"])))
    return 0


def main():
    p = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    sub = p.add_subparsers(dest="cmd", required=True)

    pr = sub.add_parser("prep", help="convert a folder and sheet the fronts")
    pr.add_argument("src")
    pr.add_argument("--work", default=WORK)
    pr.add_argument("--out", default="batch-photos.json")
    pr.add_argument("--singles", action="store_true",
                    help="fronts only; do not pair them")
    pr.add_argument("--skip", help="comma-separated basenames to ignore")
    pr.add_argument("--backs", action="store_true",
                    help="also sheet the backs; the card number is only there")
    pr.add_argument("--defaults", metavar="k=v,k=v",
                    help="what the whole box shares, e.g. "
                         "lot=LOT-002,source=Big 5 Upland")
    pr.set_defaults(fn=do_prep)

    fi = sub.add_parser("file", help="file an identified batch onto its SKUs")
    fi.add_argument("batch")
    fi.add_argument("--workbook", default=WORKBOOK)
    fi.add_argument("--go", action="store_true")
    fi.set_defaults(fn=do_file)

    ad = sub.add_parser("add", help="a box not in the workbook yet: create the "
                                    "rows, then file the photos")
    ad.add_argument("batch")
    ad.add_argument("--workbook", default=WORKBOOK)
    ad.add_argument("--go", action="store_true")
    ad.set_defaults(fn=do_add)

    a = p.parse_args()
    return a.fn(a)


if __name__ == "__main__":
    sys.exit(main())
