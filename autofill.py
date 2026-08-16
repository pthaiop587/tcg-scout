"""Fill in what a hand-typed card can work out for itself.

    python autofill.py            # say what it would do
    python autofill.py --go       # do it

Two things, both of which are only missing because the row was typed rather
than added by a script.

file_batch.py and add_card.py assign a SKU when they add a card. Typing
straight into the Inventory tab does not, and a row without one is invisible
to everything downstream: make_ebay_csv.py cannot export it, add_photos.py has
nothing to file a picture against. Sixty cards typed in after a box rip is
sixty cards that quietly do not exist.

So this fills the gaps, in the order the rows already sit in, carrying on from
the highest SKU already used. It only ever writes into an EMPTY SKU cell -- a
card that already has one is never renumbered, because a SKU is what the
photographs on disk and any listing already out there are named after.

A row is a card if it has a name in "Player or card name". A row with a
quantity and nothing else is a half-typed line, not a card, and is left alone
rather than being given a number it will keep forever.

CATEGORY, from Sport or game. Football, Basketball and Baseball are Sports;
Pokemon, Palworld, One Piece and Disney are TCG. make_ebay_csv.py picks the
eBay category code off that column, and a card with it blank goes up under the
wrong one. It is derived rather than left as a formula because make_ebay_csv
reads the workbook with data_only=True -- a formula that Excel has not
recalculated reads as empty, and the listing would be miscategorised without
anything looking wrong in the sheet.
"""

import argparse
import os
import re
import sys

from openpyxl import load_workbook

WORKBOOK = "Card Run HQ - Master.xlsx"
NAME_COL = "Player or card name"

# Sport or game -> the eBay-facing Category. Anything not listed is left blank
# rather than guessed; "Other" genuinely could be either.
CATEGORY_OF = {
    "football": "Sports", "basketball": "Sports", "baseball": "Sports",
    "hockey": "Sports", "soccer": "Sports",
    "pokemon": "TCG", "palworld": "TCG", "one piece": "TCG",
    "disney": "TCG", "lorcana": "TCG", "magic": "TCG",
}


def scan(path):
    wb = load_workbook(path)
    if "Inventory" not in wb.sheetnames:
        sys.exit("%s has no Inventory tab" % path)
    ws = wb["Inventory"]
    hdr = [c.value for c in ws[1]]
    for needed in ("SKU", NAME_COL):
        if needed not in hdr:
            sys.exit("Inventory has no '%s' column" % needed)

    sku_col = hdr.index("SKU") + 1
    name_col = hdr.index(NAME_COL) + 1
    sport_col = hdr.index("Sport or game") + 1 if "Sport or game" in hdr else None
    cat_col = hdr.index("Category") + 1 if "Category" in hdr else None

    highest = 0
    blanks = []
    cats = []
    for r in range(2, ws.max_row + 1):
        sku = str(ws.cell(row=r, column=sku_col).value or "").strip()
        name = str(ws.cell(row=r, column=name_col).value or "").strip()
        m = re.match(r"CRH-(\d+)$", sku, re.I)
        if m:
            highest = max(highest, int(m.group(1)))
        elif name and not sku:
            blanks.append((r, name))

        if name and sport_col and cat_col:
            have = str(ws.cell(row=r, column=cat_col).value or "").strip()
            sport = str(ws.cell(row=r, column=sport_col).value or "").strip()
            want = CATEGORY_OF.get(sport.lower())
            if want and not have:
                cats.append((r, sport, want))
    return wb, ws, sku_col, cat_col, highest, blanks, cats


def main():
    p = argparse.ArgumentParser(
        description="Give a SKU to every hand-typed card that has none.")
    p.add_argument("--workbook", default=WORKBOOK)
    p.add_argument("--go", action="store_true",
                   help="do it; without this it only says what it would do")
    a = p.parse_args()

    if not os.path.exists(a.workbook):
        sys.exit("no workbook at %s" % a.workbook)

    wb, ws, sku_col, cat_col, highest, blanks, cats = scan(a.workbook)

    if not blanks and not cats:
        print("Nothing to fill in. Highest SKU is CRH-%04d." % highest)
        return 0

    if cats:
        print("%d card%s with no Category, which can be worked out from the "
              "sport:" % (len(cats), "" if len(cats) == 1 else "s"))
        seen = {}
        for _r, sport, want in cats:
            seen.setdefault((sport, want), 0)
            seen[(sport, want)] += 1
        for (sport, want), n in sorted(seen.items()):
            print("   %-14s -> %-10s %d card(s)" % (sport, want, n))

    if not blanks:
        if a.go:
            for row, _sport, want in cats:
                ws.cell(row=row, column=cat_col, value=want)
            wb.save(a.workbook)
            print("\nFilled %d Category cell(s)." % len(cats))
        else:
            print("\nThis was a dry run. Add --go to write them in.")
        return 0

    print("%d card%s with no SKU. The next free number is CRH-%04d."
          % (len(blanks), "" if len(blanks) == 1 else "s", highest + 1))
    for r, name in blanks[:8]:
        print("   row %-4d %s" % (r, name))
    if len(blanks) > 8:
        print("   ... and %d more" % (len(blanks) - 8))

    if not a.go:
        print("\nThis was a dry run. Add --go to write them in.")
        print("Nothing that already has a SKU is touched -- a SKU is what your "
              "photos and any live listing are named after.")
        return 0

    n = highest
    for row, _name in blanks:
        n += 1
        ws.cell(row=row, column=sku_col, value="CRH-%04d" % n)
    for row, _sport, want in cats:
        ws.cell(row=row, column=cat_col, value=want)
    wb.save(a.workbook)
    if cats:
        print("\nFilled %d Category cell(s) from the sport." % len(cats))

    print("\nWrote CRH-%04d to CRH-%04d." % (highest + 1, n))
    print("Those cards can now be exported to eBay and photographed. "
          "Nothing that already had a SKU was changed.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
