"""Put the workbook's Inventory onto the dashboard.

    python export_inventory.py            # for your own machine
    python export_inventory.py --publish  # also a safe copy for the public site
    python build_all.py . card-run-hq.html

Reads the Inventory tab and writes JSON that build_all.py renders as a
"My inventory" tab, so what you type into the spreadsheet is what you see on
the page. Run it whenever the workbook changes; it rewrites from scratch, so
it can never drift from the file.

TWO FILES, AND THE DIFFERENCE MATTERS.

`inventory.json` is everything -- cost, notes, the lot. It is gitignored, so
it stays on this machine and the dashboard shows it only in the copy of the
page built here. CI has never seen your workbook and cannot: it is gitignored
too.

`--publish` additionally writes `inventory-public.json`, which is committed
and served with the site. It carries what a card IS -- name, set, number,
parallel, condition, quantity, status, photo -- and NOT what it cost you, what
it sold for, or your notes. The repo is public, so anything in that file is
readable by anyone with the URL. That is a decision about your own privacy
rather than a technical one, which is why it takes a flag and is not the
default.

build_all.py prefers the local file when it is there and falls back to the
published one, so a build on this machine shows everything and a build in CI
shows only what you chose to publish.
"""

import argparse
import io
import json
import os
import re
import sys

from openpyxl import load_workbook

WORKBOOK = "Card Run HQ - Master.xlsx"
LOCAL = "inventory.json"
PUBLIC = "inventory-public.json"
PHOTOS = "photos"
PAGES = "https://pthaiop587.github.io/tcg-scout"

# Inventory column -> the short key the page uses. Anything not listed here is
# not carried, which is deliberate: the page should not quietly start showing
# a column somebody adds to the workbook.
FIELDS = {
    "SKU": "sku", "Status": "status", "Lot ID": "lot",
    "Category": "cat", "Sport or game": "sport", "League": "league",
    "Year": "year", "Brand / set": "brand", "Insert set": "insert",
    "Parallel": "parallel", "Player or card name": "name", "Card #": "num",
    "Serial /": "serial", "Team": "team", "RC": "rc", "Auto": "auto",
    "Relic": "relic", "Graded by": "grader", "Grade": "grade",
    "Cert #": "cert", "Card condition": "cond", "Qty": "qty",
    "Cost each": "cost", "Market value": "market", "Ask price": "ask",
    "eBay title": "title", "Notes": "notes",
}

# What NEVER goes in the published copy. Money you paid, money you made, and
# free text that could hold anything.
PRIVATE = {"cost", "notes", "lot"}


def photos_for(sku, folder=PHOTOS):
    """Paths relative to the page, not absolute URLs.

    The dashboard sits next to the photos folder in both places it is ever
    read -- built here, or published with the site -- so a relative path
    resolves in both. Absolute URLs pointed at the published site, which
    meant the thumbnails were broken on the machine the photos were actually
    on. eBay is the one consumer that genuinely needs a public https address,
    and make_ebay_csv.py builds those itself.
    """
    out = []
    for suffix in ("", "-back", "-2", "-3"):
        for ext in (".jpg", ".jpeg", ".png"):
            rel = "%s/%s%s%s" % (folder, sku, suffix, ext)
            if os.path.exists(rel):
                out.append(rel.replace("\\", "/"))
                break
    return out


def clean(value):
    """openpyxl hands back dates and Decimals; JSON wants neither."""
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        return value.isoformat()[:10]
    if isinstance(value, float) and value == int(value):
        return int(value)
    return value


def read_inventory(path, folder=PHOTOS):
    wb = load_workbook(path, data_only=True)
    if "Inventory" not in wb.sheetnames:
        sys.exit("%s has no Inventory tab" % path)
    ws = wb["Inventory"]

    hdr = [c.value for c in ws[1]]
    missing = [h for h in ("SKU", "Player or card name") if h not in hdr]
    if missing:
        sys.exit("Inventory is missing column(s): %s" % ", ".join(missing))

    cards = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = dict(zip(hdr, row))
        sku = str(rec.get("SKU") or "").strip()
        if not sku:
            continue
        card = {}
        for header, key in FIELDS.items():
            card[key] = clean(rec.get(header))
        # a title is a formula, so a workbook that has not been opened in Excel
        # since it was written has no cached value for it -- build one instead
        if not card["title"]:
            bits = [card["year"], card["brand"], card["insert"],
                    card["parallel"], card["name"]]
            card["title"] = " ".join(str(b) for b in bits if b).strip()
        card["photos"] = photos_for(sku, folder)
        cards.append(card)
    return cards


def totals(cards, money=True):
    """The same arithmetic as the workbook's Summary, so they agree.

    Reads with .get because the published copy has already had cost stripped
    out of every card -- and drops the cost total for the same reason. A
    summed cost is still a cost; leaving it in would publish what the stock
    was worth to buy while carefully removing it card by card.
    """
    held = [c for c in cards if c.get("status") in ("Unlisted", "Listed", "Review")]
    num = lambda v: float(v) if isinstance(v, (int, float)) else 0.0
    t = {
        "rows": len(cards),
        "cards": sum(int(num(c.get("qty")) or 1) for c in held),
        "cost": round(sum(num(c.get("qty") or 1) * num(c.get("cost")) for c in held), 2),
        "market": round(sum(num(c.get("qty") or 1) * num(c.get("market")) for c in held), 2),
        "review": sum(1 for c in cards if c.get("status") == "Review"),
        "unlisted": sum(1 for c in cards if c.get("status") == "Unlisted"),
        "listed": sum(1 for c in cards if c.get("status") == "Listed"),
        "sold": sum(1 for c in cards if c.get("status") == "Sold"),
        "nophoto": sum(1 for c in cards
                       if c.get("status") == "Unlisted" and not c.get("photos")),
    }
    if not money:
        t.pop("cost")
    return t


def strip_private(cards):
    return [{k: v for k, v in c.items() if k not in PRIVATE} for c in cards]


def write(path, cards, money):
    payload = {"money": money, "totals": totals(cards, money), "cards": cards}
    with io.open(path, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, ensure_ascii=False, separators=(",", ":"))
    return os.path.getsize(path)


def main():
    p = argparse.ArgumentParser(
        description="Put the workbook's Inventory onto the dashboard.")
    p.add_argument("--workbook", default=WORKBOOK)
    p.add_argument("--photos", default=PHOTOS)
    p.add_argument("--publish", action="store_true",
                   help="ALSO write the copy that goes on the public site, "
                        "without cost, notes or lot")
    a = p.parse_args()

    if not os.path.exists(a.workbook):
        sys.exit("no workbook at %s" % a.workbook)

    cards = read_inventory(a.workbook, a.photos)
    if not cards:
        print("no cards in the Inventory tab yet -- nothing to show.")
        return 0

    t = totals(cards)
    size = write(LOCAL, cards, money=True)
    print("%s: %d card%s, %d KB  (stays on this machine)"
          % (LOCAL, t["rows"], "" if t["rows"] == 1 else "s", size // 1024 or 1))
    print("   %d held, %s at cost, %s at market"
          % (t["cards"], "$%.2f" % t["cost"], "$%.2f" % t["market"]))
    if t["review"]:
        print("   %d held for review -- they cannot be exported to eBay" % t["review"])
    if t["nophoto"]:
        print("   %d ready to list with no photo" % t["nophoto"])

    if a.publish:
        size = write(PUBLIC, strip_private(cards), money=False)
        print("\n%s: the same cards WITHOUT cost, notes or lot, %d KB"
              % (PUBLIC, size // 1024 or 1))
        print("   this one is committed and served with the site, so treat it "
              "as public. Commit it to publish; delete it to stop.")

    print("\nNow rebuild the page:  python build_all.py . card-run-hq.html")
    return 0


if __name__ == "__main__":
    sys.exit(main())
