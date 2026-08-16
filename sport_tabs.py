"""A tab per sport, so each one can be looked at on its own.

    python sport_tabs.py                    # rebuild a tab for every sport in use
    python sport_tabs.py --add Basketball   # start one before there are any cards
    python sport_tabs.py --list             # what is in the workbook now

WHY THESE ARE VIEWS AND NOT SEPARATE INVENTORIES.

Everything downstream reads the **Inventory** tab and only that: file_batch.py
takes the next SKU from it, make_ebay_csv.py exports from it, the Summary and
Audit tabs count it. Type a
basketball card into a tab of its own and it gets a SKU somebody else already
has, never reaches an eBay upload, and is missing from every total -- silently,
because nothing is looking for a second inventory.

So there is one Inventory, with a Sport column, and these tabs are read-only
copies of the rows for one sport. Everything still works, and each sport can
still be looked at on its own.

They are regenerated from Inventory each run, which is why they carry a date
and a warning not to type in them: anything typed here is overwritten the next
time this runs, and `Update workbook.cmd` runs it.

FILTER() would have made them live rather than generated, and was tried --
it does not survive a real recalculation outside the newest Excel, so a
formula-driven tab would show "none" on a perfectly good workbook.
"""

import argparse
import os
import sys
from datetime import date

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import inuse

WORKBOOK = "Card Run HQ - Master.xlsx"

# Written into A1 of every tab this makes. Tabs are only ever replaced when
# this is found, so a sheet somebody made by hand is never overwritten because
# it happened to be called Basketball.
MARK = "VIEW — generated from Inventory. Do not type here."

SHOW = ["SKU", "Status", "Date in", "Year", "Brand / set", "Insert set",
        "Parallel", "Player or card name", "Card #", "Serial /", "Team",
        "League", "Card condition", "Qty", "Cost each", "Market value",
        "Ask price",
        "Raw price", "Raw last sold", "Raw last sale",
        "PSA 9 price", "PSA 9 last sold", "PSA 9 last sale",
        "PSA 10 price", "PSA 10 last sold", "PSA 10 last sale",
        "Listed on", "Sold on", "Notes"]

WIDTH = {"SKU": 12, "Status": 12, "Date in": 11, "Year": 7,
         "Brand / set": 24, "Insert set": 20, "Parallel": 18,
         "Player or card name": 22, "Card #": 9, "Serial /": 9, "Team": 18,
         "League": 9, "Card condition": 18, "Qty": 6, "Cost each": 11,
         "Market value": 12, "Ask price": 11, "Listed on": 11, "Sold on": 11,
         "Raw price": 11, "Raw last sold": 12, "Raw last sale": 12,
         "PSA 9 price": 11, "PSA 9 last sold": 13, "PSA 9 last sale": 13,
         "PSA 10 price": 11, "PSA 10 last sold": 14, "PSA 10 last sale": 14,
         "Notes": 34}

# Dates arrive from openpyxl as datetimes and would otherwise render as a
# five-digit serial number, which reads as a card number at a glance.
DATES = ("Date in", "Listed on", "Sold on", "Raw last sold",
         "PSA 9 last sold", "PSA 10 last sold")
CASH = ("Cost each", "Market value", "Ask price",
        "Raw price", "Raw last sale", "PSA 9 price", "PSA 9 last sale",
        "PSA 10 price", "PSA 10 last sale")
DATEFMT = "yyyy-mm-dd"

MONEY = '"$"#,##0.00'
HEADFILL = PatternFill("solid", fgColor="1B1F2A")
NOTEFILL = PatternFill("solid", fgColor="FFF6E5")
HEADFONT = Font(bold=True, color="FFFFFF", size=10)
SMALL = Font(size=9, color="55607A")


def read(path):
    wb = load_workbook(path)
    if "Inventory" not in wb.sheetnames:
        sys.exit("%s has no Inventory tab" % path)
    ws = wb["Inventory"]
    hdr = [c.value for c in ws[1]]
    if "Sport or game" not in hdr:
        sys.exit("Inventory has no 'Sport or game' column")
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = dict(zip(hdr, row))
        if not str(rec.get("SKU") or "").strip():
            continue
        rows.append(rec)
    return wb, hdr, rows


def generated(wb, name):
    """True only for a tab this script made -- never one typed by hand."""
    if name not in wb.sheetnames:
        return False
    return str(wb[name]["A1"].value or "").startswith("VIEW")


def sports_in(rows):
    out = {}
    for r in rows:
        s = str(r.get("Sport or game") or "").strip()
        if s:
            out.setdefault(s, 0)
            out[s] += 1
    return out


def stray(ws, mine, cols):
    """Cells on a generated tab that Inventory cannot account for.

    A1 saying "do not type here" is a sign, not a lock. Somebody typed on this
    tab once and the next refresh threw it away, which is the worst kind of
    data loss: silent, and caused by the tool doing exactly what it was told.

    So before this tab is destroyed, work out whether anything on it is NOT a
    copy of Inventory. Three ways that happens: a value typed past the columns
    this script writes, a row whose SKU is not in Inventory at all, and a cell
    edited to differ from the Inventory row it was copied from.
    """
    by_sku = {}
    for rec in mine:
        s = str(rec.get("SKU") or "").strip()
        if s:
            by_sku[s] = rec

    # Compare against the tab's OWN header, not against the column list we are
    # about to write. Those two drift every time a column is added, and
    # comparing by position across that drift makes every cell on the tab look
    # hand-edited -- which reports the whole sheet as precious and keeps a
    # useless copy of it.
    old_hdr = [ws.cell(row=2, column=c).value
               for c in range(1, ws.max_column + 1)]

    found = []
    for r in range(3, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value
                for c in range(1, ws.max_column + 1)]
        if not any(v not in (None, "") for v in vals):
            continue

        sku = str(vals[0] or "").strip()
        rec = by_sku.get(sku)
        if not rec:                                       # a row of its own
            found.append((r, "A", sku or "(no SKU)"))
            continue

        for i, name in enumerate(old_hdr):
            was = vals[i]
            if was in (None, ""):
                continue
            if not name or name not in rec:
                # a column Inventory has never heard of, so nothing here came
                # from it
                found.append((r, name or get_column_letter(i + 1), was))
                continue
            now = rec.get(name)
            if was == now:
                continue
            if isinstance(was, (int, float)) and isinstance(now, (int, float)) \
                    and abs(float(was) - float(now)) < 0.005:
                continue
            found.append((r, name, was))
    return found


def build_tab(wb, sport, rows, hdr):
    if sport in wb.sheetnames:
        if not generated(wb, sport):
            print("   %s: a tab of that name already exists and was not made "
                  "by this script, so it was left alone." % sport)
            return 0

        mine_now = [r for r in rows
                    if str(r.get("Sport or game") or "").strip().lower()
                    == sport.lower()]
        cols_now = [c for c in SHOW if c in hdr]
        typed = stray(wb[sport], mine_now, cols_now)
        if typed:
            keep = "%s (typed on)" % sport
            n = 2
            while keep in wb.sheetnames:
                keep = "%s (typed on %d)" % (sport, n)
                n += 1
            wb[sport].title = keep
            print("   %s: %d cell(s) on this tab were typed by hand, not "
                  "copied from Inventory." % (sport, len(typed)))
            for r, where, v in typed[:6]:
                print("        row %-4d %-18s %s" % (r, where, str(v)[:44]))
            if len(typed) > 6:
                print("        ... and %d more" % (len(typed) - 6))
            print("      Kept as %r rather than thrown away. Move what you "
                  "want into Inventory,\n      then delete that sheet." % keep)
        else:
            del wb[sport]

    ws = wb.create_sheet(sport)
    ws["A1"] = MARK + "  Rebuilt %s from the Inventory tab." % date.today().isoformat()
    ws["A1"].font = SMALL
    ws["A1"].fill = NOTEFILL
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(6, len(SHOW)))
    ws.row_dimensions[1].height = 18

    cols = [c for c in SHOW if c in hdr]
    for i, name in enumerate(cols, start=1):
        c = ws.cell(row=2, column=i, value=name)
        c.fill = HEADFILL
        c.font = HEADFONT
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = WIDTH.get(name, 14)
    ws.row_dimensions[2].height = 28
    ws.freeze_panes = ws.cell(row=3, column=1)

    mine = [r for r in rows
            if str(r.get("Sport or game") or "").strip().lower() == sport.lower()]
    for n, rec in enumerate(mine, start=3):
        for i, name in enumerate(cols, start=1):
            cell = ws.cell(row=n, column=i, value=rec.get(name))
            if name in CASH:
                cell.number_format = MONEY
            elif name in DATES:
                cell.number_format = DATEFMT
    if cols:
        ws.auto_filter.ref = "A2:%s%d" % (get_column_letter(len(cols)),
                                          max(3, len(mine) + 2))
    return len(mine)


def main():
    p = argparse.ArgumentParser(
        description="Rebuild a read-only tab per sport from the Inventory tab.")
    p.add_argument("--workbook", default=WORKBOOK)
    p.add_argument("--add", action="append", default=[], metavar="SPORT",
                   help="make a tab for this sport even if no cards use it yet "
                        "(repeatable)")
    p.add_argument("--list", action="store_true",
                   help="say what sports are in the workbook and stop")
    a = p.parse_args()

    inuse.refuse_if_open(a.workbook)

    if not os.path.exists(a.workbook):
        sys.exit("no workbook at %s -- run make_workbook.py first" % a.workbook)

    wb, hdr, rows = read(a.workbook)
    counts = sports_in(rows)

    if a.list:
        if not counts:
            print("No sport is set on any card yet.")
        for s in sorted(counts):
            print("   %-16s %d card(s)" % (s, counts[s]))
        made = [n for n in wb.sheetnames if generated(wb, n)]
        print("\nview tabs: %s" % (", ".join(made) if made else "none yet"))
        return 0

    # every sport in use, plus any asked for, plus any view tab already here
    wanted = set(counts) | set(a.add) | {n for n in wb.sheetnames if generated(wb, n)}
    if not wanted:
        print("No sport is set on any card, and none was asked for.")
        print("Try:  python sport_tabs.py --add Basketball")
        return 0

    print("rebuilding %d view tab(s) from Inventory:" % len(wanted))
    for sport in sorted(wanted):
        n = build_tab(wb, sport, rows, hdr)
        print("   %-16s %d card(s)" % (sport, n))

    wb.save(a.workbook)
    print("\nType into Inventory, not into these -- they are rebuilt from it "
          "every time this runs.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
