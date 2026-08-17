"""Tests for the price lookup and the daily audit.

Run: python -m pytest test_prices.py

Nothing here touches the network. What is worth guarding is the reasoning
around the numbers, not the fetching of them: which tier a completed sale
belongs to, whether the columns land in pairs the workbook can read, and
whether "what moved" is arithmetic anyone can check.

The grade test is the important one. Those sales tables carry no ids and no
headings -- their ORDER is the only thing separating ungraded from PSA 9 from
PSA 10 on the page, so the grade is read out of each listing's own title. Get
that wrong and PSA 10 money lands in the raw column, which reads as a very
good day.
"""

import subprocess
import sys

import pytest
from openpyxl import load_workbook

import prices


# --- which tier a sale belongs to -------------------------------------------

@pytest.mark.parametrize("title,tier", [
    ("2025 Prizm Draft Picks Arch Manning Gold Ice #166", "raw"),
    ("Panini 2025 Prizm Arch Manning Rated Prospect Texas #166", "raw"),
    ("2025 Prizm Draft Picks - Arch Manning #166 Gold Ice PSA 10 #166", "psa10"),
    ("2025 Prizm Arch Manning Gold Ice PSA 9 Texas #166", "psa9"),
    ("Arch Manning Gold Ice BGS 10 #166", "psa10"),
    ("Arch Manning Gold Ice SGC 10", "psa10"),
])
def test_the_grade_is_read_from_the_listing(title, tier):
    assert prices.grade_of(title) == tier


@pytest.mark.parametrize("title", [
    "2025 Prizm Draft Picks Arch Manning Gold Ice SGC 8.5 #166",
    "Panini 2025 Prizm Arch Manning Gold Ice #166 PSA 8 Texas",
    "Arch Manning Gold Ice BGS 9.5",
    "Arch Manning Gold Ice graded slab",
])
def test_a_grade_we_do_not_track_is_not_counted_as_raw(title):
    """An 8 or a 9.5 is not a raw sale. Counting it as one would drag the raw
    last-sale price up by a slab premium."""
    assert prices.grade_of(title) is None


def test_a_psa_95_is_not_a_psa_9():
    assert prices.grade_of("Arch Manning Gold Ice PSA 9.5 #166") is None


# --- money ------------------------------------------------------------------

def test_the_first_figure_is_the_sale_price():
    """A sale cell can hold the price and then the price with postage. The
    first is the one comparable with the guide."""
    assert prices.money("$54.00     $60.00") == 54.0
    assert prices.money("$1,234.50") == 1234.5
    assert prices.money("-") is None
    assert prices.money("") is None


# --- the columns the workbook has to be able to read -------------------------

def test_every_tier_has_a_price_a_date_and_a_sale():
    for tier in prices.TIERS:
        cols = prices.COL_OF[tier]
        assert len(cols) == 3, tier
        for name in cols:
            assert name in [n for n, _w in prices.PRICE_COLS], name


def test_the_columns_are_appended_never_inserted(tmp_path):
    """workbook_extra.py addresses Inventory by column LETTER. A column
    inserted rather than appended moves every formula silently."""
    wb = tmp_path / "Card Run HQ - Master.xlsx"
    subprocess.run([sys.executable, "make_workbook.py", "--out", str(wb)],
                   check=True, capture_output=True)
    before = [c.value for c in load_workbook(wb)["Inventory"][1]]

    book = load_workbook(wb)
    prices.ensure_columns(book["Inventory"])
    book.save(wb)

    after = [c.value for c in load_workbook(wb)["Inventory"][1]]
    assert after[:len(before)] == before, "existing columns moved"
    assert len(after) == len(before) + len(prices.PRICE_COLS)


def test_running_it_twice_does_not_add_the_columns_again(tmp_path):
    wb = tmp_path / "Card Run HQ - Master.xlsx"
    subprocess.run([sys.executable, "make_workbook.py", "--out", str(wb)],
                   check=True, capture_output=True)
    book = load_workbook(wb)
    prices.ensure_columns(book["Inventory"])
    prices.ensure_columns(book["Inventory"])
    book.save(wb)
    hdr = [c.value for c in load_workbook(wb)["Inventory"][1]]
    assert hdr.count("Raw price") == 1
    assert hdr.count("PSA 10 last sale") == 1


# --- names, which is what the number-based fallback is for -------------------

def test_a_name_matches_however_it_is_punctuated():
    assert prices.norm("T.J. Sanders") == prices.norm("TJ Sanders")
    assert prices.norm("Harold Fannin Jr.") == prices.norm("Harold Fannin")
    assert prices.norm("Chris Paul Jr.") == prices.norm("chris  paul  jr")


def test_the_card_url_is_built_not_guessed():
    u = prices.card_url("football-cards-2025-panini-prizm-draft-picks",
                        "Arch Manning", "Gold Ice", "166")
    assert u.endswith("/arch-manning-gold-ice-166")
    assert prices.card_url("s", "Arch Manning", "", "166").endswith(
        "/arch-manning-166")


# --- the audit --------------------------------------------------------------

def test_the_daily_flag_turns_on_what_a_daily_check_needs():
    """--daily has to imply writing AND overwriting: a check that only fills
    blanks would report nothing moved, every day, for ever."""
    src = open("prices.py", encoding="utf-8").read()
    i = src.index("if a.daily:")
    block = src[i:i + 220]
    assert "a.go" in block and "a.overwrite" in block
    assert "price-check.log" in block


def test_a_school_is_never_overwritten_by_a_daily_run():
    """Prices move; a player's college does not. Re-voting on sixty correct
    schools every morning buries the prices in noise."""
    src = open("prices.py", encoding="utf-8").read()
    # anchored on "if a.teams" rather than the whole line: the condition has
    # picked up a guard since (Pokemon cards have no college) and will again.
    i = src.index("if a.teams")
    block = src[i:i + 260]
    assert '"Team" in g' in block
    assert "blank_cell(cell.value)" in block
    assert "or a.overwrite" not in block


# --- picking one card -------------------------------------------------------

def test_skus_can_be_repeated_or_comma_separated_or_both():
    """All three are what somebody reaches for; none of them is wrong."""
    assert prices.wanted_skus(["CRH-0062"]) == {"CRH-0062"}
    assert prices.wanted_skus(["CRH-0062,CRH-0039"]) == {"CRH-0062", "CRH-0039"}
    assert prices.wanted_skus(["CRH-0062", "CRH-0039"]) == {"CRH-0062", "CRH-0039"}
    assert prices.wanted_skus(["crh-0062"]) == {"CRH-0062"}
    assert prices.wanted_skus(["CRH-0062, CRH-0039 ,"]) == {"CRH-0062", "CRH-0039"}


def test_no_skus_asked_for_is_empty_not_none():
    assert prices.wanted_skus(None) == set()
    assert prices.wanted_skus([]) == set()


def test_an_unknown_sku_stops_before_it_opens_a_browser(tmp_path):
    """Four minutes of scraping to be told the SKU was a typo would be a poor
    way to find out. It also means this test needs no network."""
    wb = tmp_path / "Card Run HQ - Master.xlsx"
    subprocess.run([sys.executable, "make_workbook.py", "--out", str(wb)],
                   check=True, capture_output=True)
    import file_batch as fb
    fb.add_rows(str(wb), [{"player": "Travis Hunter", "sport": "Football"}])

    r = subprocess.run([sys.executable, "prices.py", "--workbook", str(wb),
                        "--sku", "CRH-9999"], capture_output=True, text=True,
                       timeout=90)
    assert r.returncode != 0
    body = r.stdout + r.stderr
    assert "CRH-9999" in body
    assert "playwright" not in body.lower()


def test_naming_a_sku_ignores_the_sport_filter(tmp_path):
    """--sport defaults to Football. Asking for a Pokemon card by SKU should
    fetch it, not answer "no Football cards found", which is true and
    useless."""
    wb = tmp_path / "Card Run HQ - Master.xlsx"
    subprocess.run([sys.executable, "make_workbook.py", "--out", str(wb)],
                   check=True, capture_output=True)
    import file_batch as fb
    fb.add_rows(str(wb), [{"player": "Mega Darkrai ex", "sport": "Pokemon"}])

    # It must get PAST selection -- proved by the failure being about the
    # lookup rather than about the card not existing.
    r = subprocess.run([sys.executable, "prices.py", "--workbook", str(wb),
                        "--sku", "CRH-0001", "--report"],
                       capture_output=True, text=True, timeout=300)
    body = r.stdout + r.stderr
    assert "no card with SKU" not in body, body
    assert "no Football cards found" not in body, body


def test_the_set_page_is_only_fetched_when_it_is_needed():
    """It costs a couple of minutes and two thousand rows. A card whose own
    URL works should never trigger it."""
    src = open("prices.py", encoding="utf-8").read()
    assert "def console(slug):" in src
    i = src.index("urls = [card_url(")
    j = src.index("hit = find()", i)
    assert i < j, "the set page is consulted before the card's own URL"


def test_a_reported_move_says_so_when_it_was_not_written():
    """--go fills empty cells only. A card that already had a price keeps it
    while the audit reports the change anyway, which reads exactly like a
    successful update and is not one. That cost a wrong price sitting in the
    sheet for two runs."""
    src = open("prices.py", encoding="utf-8").read()
    i = src.index("price(s) moved since the last check")
    block = src[i:i + 700]
    assert "not a.overwrite" in block
    assert "NOT written" in block


def test_daily_always_overwrites():
    """Which is why the daily check does not hit that trap: prices move, and a
    check that could only ever fill blanks would report movement it never
    recorded, every morning."""
    src = open("prices.py", encoding="utf-8").read()
    i = src.index("if a.daily:")
    assert "a.overwrite" in src[i:i + 220]
# --- the second price guide -------------------------------------------------


def test_a_pokemon_number_loses_its_denominator():
    """The card is printed 013/084 and addressed 13. Sending the printed form
    asks for a page with a slash in its name."""
    assert prices.pokemon_num("013/084") == "13"
    assert prices.pokemon_num("111/084") == "111"
    assert prices.pokemon_num("#006/084") == "6"
    assert prices.pokemon_num("45") == "45"


def test_a_pokemon_card_routes_to_the_other_site():
    site, sets, num, par, poke = prices.route(
        {"brand": "Pitch Black", "num": "013/084", "parallel": "",
         "insert": ""})
    assert site == prices.SITES["pricecharting"]
    assert sets == ["pokemon-pitch-black"]
    assert (num, poke) == ("13", True)


def test_a_football_card_still_routes_where_it_did():
    site, sets, num, par, poke = prices.route(
        {"brand": "Panini Prizm Draft Picks", "num": "190",
         "parallel": "Gold Ice", "insert": ""})
    assert site == prices.SITES["sportscardspro"]
    assert sets[0] == prices.SET_SLUG
    assert (num, par, poke) == ("190", "Gold Ice", False)


def test_a_football_insert_still_gets_its_own_set_page():
    _site, sets, _n, _p, _k = prices.route(
        {"brand": "Panini Prizm Draft Picks", "num": "21",
         "parallel": "Gold Ice", "insert": "Student Orientation"})
    assert sets[0] == prices.INSERT_SLUG["student orientation"]
    assert prices.SET_SLUG in sets


def test_a_rarity_is_not_a_variant_and_stays_out_of_the_url():
    """Reverse Holo is a different card at ten times the price. Secret Rare is
    a description of one. Only the first belongs in the URL."""
    for rarity in ("Secret Rare", "Double Rare", "Uncommon", "Common"):
        _s, _sets, _n, par, _k = prices.route(
            {"brand": "Pitch Black", "num": "111/084", "parallel": rarity,
             "insert": ""})
        assert par == "", rarity


def test_a_reverse_holo_does_reach_the_url():
    _s, _sets, _n, par, _k = prices.route(
        {"brand": "Pitch Black", "num": "013/084",
         "parallel": "Reverse Holo", "insert": ""})
    assert par == "Reverse Holo"
    u = prices.card_url("pokemon-pitch-black", "Goldeen", par, "13",
                        base=prices.SITES["pricecharting"])
    assert u.endswith("/pokemon-pitch-black/goldeen-reverse-holo-13")


def test_the_two_sites_spell_an_apostrophe_differently():
    """sportscardspro drops it, pricecharting keeps it url-encoded. Guessing
    one costs a card; offering both costs one extra request."""
    poke = prices.name_forms("Misty's Vitality", True)
    assert poke[0] == "misty%27s-vitality"
    sport = prices.name_forms("Ja'Marr Chase", False)
    assert sport[0] == prices.slug("Ja'Marr Chase")
    assert "ja%27marr-chase" in sport
    assert prices.name_forms("Goldeen", True) == ["goldeen"]


def test_a_url_can_be_built_against_either_site():
    poke = prices.card_url("pokemon-pitch-black", "Rampardos ex", "", "45",
                           base=prices.SITES["pricecharting"])
    assert poke == ("https://www.pricecharting.com/game/pokemon-pitch-black"
                    "/rampardos-ex-45")
    # unchanged when nobody asks for a different site
    assert prices.card_url("s", "Arch Manning", "", "166").startswith(
        prices.BASE)


def test_one_run_can_cover_two_games(tmp_path):
    """The 8am job runs once. If the sport filter took a single name, half the
    workbook would silently never be priced."""
    wb = tmp_path / "b.xlsx"
    subprocess.run([sys.executable, "make_workbook.py", "--out", str(wb)],
                   check=True, capture_output=True)
    book = load_workbook(wb)
    ws = book["Inventory"]
    hdr = [c.value for c in ws[1]]
    g = {n: i + 1 for i, n in enumerate(hdr) if n}
    for r, (nm, sport) in enumerate([("Cam Ward", "Football"),
                                     ("Goldeen", "Pokemon"),
                                     ("Pikachu", "Palworld")], start=2):
        ws.cell(row=r, column=g["Player or card name"]).value = nm
        ws.cell(row=r, column=g["Sport or game"]).value = sport
        ws.cell(row=r, column=g["SKU"]).value = "CRH-000%d" % r
    book.save(wb)

    ws2 = load_workbook(wb)["Inventory"]
    hdr2 = [c.value for c in ws2[1]]
    got = [c["name"] for c in prices.read_cards(ws2, hdr2, "Football,Pokemon")]
    assert got == ["Cam Ward", "Goldeen"]
    assert len(prices.read_cards(ws2, hdr2, None)) == 3


def test_an_unknown_set_does_not_trigger_a_two_minute_harvest():
    """A card from a set the football page never covered still routes there,
    so nothing silently stops working -- but hunting two thousand rows for it
    every morning finds nothing and costs two minutes."""
    assert prices.on_default_set("Panini Prizm Draft Picks")
    assert prices.on_default_set("")
    assert not prices.on_default_set("MEE (set code -- confirm)")
    src = open("prices.py", encoding="utf-8").read()
    i = src.index("if not usable(prices)")
    assert "on_default_set" in src[i:i + 160]
