"""Tests for the listing pricer.

Run: python -m pytest test_price_listings.py

The rule is Mr. P's: ask the last price the card actually sold for, rounded up
to the next dollar; charge the buyer $1.50 postage; anything under 50c is bulk
rather than a listing. What is worth guarding is that "last sold" really means
the sale and not the guide average -- those disagree by multiples -- and that
the 50c floor is a floor rather than a suggestion.
"""

import subprocess
import sys

import pytest
from openpyxl import load_workbook

import price_listings as pl

FEE, FIX, SHIP, POST = 13.25, 0.30, 1.50, 0.70


def card(sku="CRH-0001", sale=None, raw=None, cost=1.26, market=None,
         status="Unlisted"):
    return {"row": 2, "sku": sku, "name": "A Card", "status": status,
            "cost": cost, "raw": raw, "sale": sale, "market": market}


# --- the arithmetic ---------------------------------------------------------

def test_the_ask_rounds_up_to_a_whole_dollar():
    assert pl.ask_for(1.19) == 2.0
    assert pl.ask_for(0.50) == 1.0
    assert pl.ask_for(22.72) == 23.0
    assert pl.ask_for(3.00) == 3.0, "a round number should not jump a dollar"


def test_the_buyer_paying_postage_is_what_makes_a_dollar_card_work():
    """Same card, same price. The only difference is who pays the postman."""
    seller_pays = pl.net_of(1.0, FEE, FIX, 0.0, POST)
    buyer_pays = pl.net_of(1.0, FEE, FIX, SHIP, POST)
    assert seller_pays < 0
    assert buyer_pays > 1.0


def test_ebay_takes_its_cut_of_the_postage_too():
    """Charging a dollar more postage does not put a dollar more in the bank,
    and a margin worked out as if it did would be wrong on every row."""
    a = pl.net_of(5.0, FEE, FIX, 1.50, POST)
    b = pl.net_of(5.0, FEE, FIX, 2.50, POST)
    assert b - a == pytest.approx(1.0 * (1 - FEE / 100.0))
    assert b - a < 1.0


# --- which number the price comes from --------------------------------------

def test_the_last_sale_beats_the_guide_price():
    """Dante Moore's guide said $1.19 while the card had just sold for $3.57.
    Asking $2 for a card the market pays $4 for is money left behind."""
    r = pl.plan([card(sale=3.57, raw=1.19)])[0]
    assert r["market"] == 3.57
    assert r["src"] == "sold"
    assert r["ask"] == 4.0


def test_the_guide_is_used_only_when_nothing_has_sold():
    r = pl.plan([card(sale=None, raw=1.19)])[0]
    assert (r["market"], r["src"], r["ask"]) == (1.19, "guide", 2.0)


def test_a_card_with_neither_is_never_guessed_at():
    r = pl.plan([card()])[0]
    assert r["bulk"] is True and r["ask"] is None
    assert "no price" in r["why"]


# --- the floor --------------------------------------------------------------

def test_under_fifty_cents_is_bulk_not_a_listing():
    r = pl.plan([card(sale=0.37)])[0]
    assert r["bulk"] is True
    assert r["ask"] is None
    assert "0.37" in r["why"]


def test_exactly_fifty_cents_is_listed_at_a_dollar():
    """The floor is 'under 50c', so 50c itself is in."""
    r = pl.plan([card(sale=0.50)])[0]
    assert r["bulk"] is False
    assert r["ask"] == 1.0


def test_the_floor_can_be_moved():
    assert pl.plan([card(sale=0.60)], min_list=1.00)[0]["bulk"] is True


def test_every_listed_card_actually_makes_money():
    """The floor is only worth having if nothing above it loses. Walk the
    range rather than trusting one example."""
    for cents in range(50, 500):
        r = pl.plan([card(sale=cents / 100.0)])[0]
        assert r["bulk"] is False
        assert pl.net_of(r["ask"], FEE, FIX, SHIP, POST) > 0


def test_market_value_is_rewritten_from_the_price_not_merely_filled():
    r = pl.plan([card(sale=0.37, market=10.56)])[0]
    assert r["market"] == 0.37
    assert r["was"] == 10.56


# --- the workbook -----------------------------------------------------------

def build(tmp_path, rows):
    wb = tmp_path / "Card Run HQ - Master.xlsx"
    subprocess.run([sys.executable, "make_workbook.py", "--out", str(wb)],
                   check=True, capture_output=True)
    book = load_workbook(wb)
    import prices
    prices.ensure_columns(book["Inventory"])
    ws = book["Inventory"]
    hdr = [c.value for c in ws[1]]
    g = {n: i + 1 for i, n in enumerate(hdr) if n}
    for i, d in enumerate(rows, start=2):
        for k, val in d.items():
            ws.cell(row=i, column=g[k]).value = val
    book.save(wb)
    return wb


def run(wb, *args):
    r = subprocess.run([sys.executable, "price_listings.py", "--workbook",
                        str(wb)] + list(args), capture_output=True, text=True)
    assert r.returncode == 0, r.stdout + r.stderr
    return r.stdout


def inv(wb):
    ws = load_workbook(wb, data_only=True)["Inventory"]
    hdr = [c.value for c in ws[1]]
    g = {n: i + 1 for i, n in enumerate(hdr) if n}
    out = {}
    for r in range(2, ws.max_row + 1):
        sku = ws.cell(row=r, column=g["SKU"]).value
        if sku:
            out[sku] = {n: ws.cell(row=r, column=i).value
                        for n, i in g.items()}
    return out


ROWS = [
    {"SKU": "CRH-0001", "Player or card name": "Arch Manning",
     "Status": "Unlisted", "Cost each": 1.26, "Raw price": 20.30,
     "Raw last sale": 12.50},
    {"SKU": "CRH-0002", "Player or card name": "James Cook",
     "Status": "Unlisted", "Cost each": 1.26, "Raw price": 0.37,
     "Raw last sale": 0.01, "Market value": 10.56},
    {"SKU": "CRH-0003", "Player or card name": "Basic Energy",
     "Status": "Review", "Cost each": 0.64},
]


def test_a_dry_run_changes_nothing(tmp_path):
    wb = build(tmp_path, ROWS)
    out = run(wb)
    assert "Nothing written" in out
    assert inv(wb)["CRH-0001"]["Ask price"] is None
    assert inv(wb)["CRH-0002"]["Market value"] == 10.56


def test_go_prices_from_the_sale_and_bulks_what_is_under_the_floor(tmp_path):
    wb = build(tmp_path, ROWS)
    run(wb, "--go")
    got = inv(wb)
    # priced off the $12.50 sale, not the $20.30 guide
    assert got["CRH-0001"]["Ask price"] == 13.0
    assert got["CRH-0002"]["Ask price"] is None
    assert got["CRH-0002"]["Status"] == "Bulk"


def test_the_stale_market_value_is_corrected(tmp_path):
    wb = build(tmp_path, ROWS)
    run(wb, "--go")
    assert inv(wb)["CRH-0002"]["Market value"] == 0.01


def test_a_review_row_is_left_for_the_person_it_is_waiting_on(tmp_path):
    wb = build(tmp_path, ROWS)
    run(wb, "--go")
    got = inv(wb)["CRH-0003"]
    assert got["Status"] == "Review"
    assert got["Ask price"] is None


def test_a_card_that_recovers_comes_back_off_the_bulk_pile(tmp_path):
    """Prices move. A card bulked in March that sells for $3 in August should
    not stay bulk for ever because of one bad week."""
    wb = build(tmp_path, ROWS)
    run(wb, "--go")
    assert inv(wb)["CRH-0002"]["Status"] == "Bulk"

    book = load_workbook(wb)
    ws = book["Inventory"]
    hdr = [c.value for c in ws[1]]
    g = {n: i + 1 for i, n in enumerate(hdr) if n}
    ws.cell(row=3, column=g["Raw last sale"]).value = 3.00
    book.save(wb)

    run(wb, "--go")
    got = inv(wb)["CRH-0002"]
    assert got["Status"] == "Unlisted"
    assert got["Ask price"] == 3.0


def test_no_bulk_leaves_status_alone(tmp_path):
    wb = build(tmp_path, ROWS)
    run(wb, "--go", "--no-bulk")
    assert inv(wb)["CRH-0002"]["Status"] == "Unlisted"


def test_the_export_skips_what_was_marked_bulk(tmp_path):
    wb = build(tmp_path, ROWS)
    run(wb, "--go")
    out = tmp_path / "e.csv"
    r = subprocess.run([sys.executable, "make_ebay_csv.py", "--workbook",
                        str(wb), "-o", str(out)], capture_output=True,
                       text=True)
    assert r.returncode == 0, r.stdout + r.stderr
    body = out.read_text(encoding="utf-8")
    assert "CRH-0001" in body
    assert "CRH-0002" not in body
