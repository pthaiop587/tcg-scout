"""One command: tidy the workbook after you have typed in it.

    python refresh.py

Run this after typing into Card Run HQ - Master.xlsx. It does, in order:

    embed_photos.py      thumbnails into the workbook's Photos tab
    autofill.py          SKU and Category for anything typed in by hand
    sport_tabs.py        a read-only tab per sport, rebuilt from Inventory

Each is still a script you can run on its own; this exists because running
three in the right order every time is exactly the kind of thing that stops
getting done. Double-click "Update workbook.cmd" and it runs this.

There used to be two more steps -- an export to JSON and a page built around
it. The dashboard was retired on 16 Aug 2026; the workbook is the whole system
now. They are in the git history if it ever comes back.

It still refuses to guess about the layout: if the workbook is on the old one
it stops and names the single command that moves it across without losing what
you typed, because rebuilding over the top would throw the data away.
"""

import argparse
import os
import subprocess
import sys

import inuse

WORKBOOK = "Card Run HQ - Master.xlsx"
# The old layout called the upload tab "eBay upload" and had no per-game tabs.
# That rename is the cleanest thing to test for -- checking for Summary or
# Photos would wrongly reject the short layout, which does not have them.
OLD_NAME = "eBay upload"
NEW_NAME = "eBay"


def run(args, why):
    print("\n> %s" % " ".join(args[1:]))
    r = subprocess.run([sys.executable] + args[1:], capture_output=True, text=True)
    out = (r.stdout or "").strip()
    if out:
        print("\n".join("   " + l for l in out.splitlines()))
    if r.returncode != 0:
        err = (r.stderr or "").strip()
        print("\n%s failed:" % why)
        print("\n".join("   " + l for l in (err or "no output").splitlines()))
        return False
    return True


def workbook_tabs(path):
    from openpyxl import load_workbook
    return set(load_workbook(path, read_only=True).sheetnames)


def main():
    p = argparse.ArgumentParser(
        description="Tidy the workbook: SKUs, categories and the game tabs.")
    p.add_argument("--workbook", default=WORKBOOK)
    a = p.parse_args()

    inuse.refuse_if_open(a.workbook)

    if not os.path.exists(a.workbook):
        print("There is no %s here." % a.workbook)
        print("Build one with:  python make_workbook.py")
        return 1

    tabs = workbook_tabs(a.workbook)
    if NEW_NAME not in tabs:
        print("%s is on the old layout -- its upload tab is still called "
              "\"%s\"." % (a.workbook, OLD_NAME))
        print("\nMove it across first. This keeps everything you have typed and "
              "writes a dated backup before it touches anything:")
        print("\n   python upgrade_workbook.py --go\n")
        print("Then run this again.")
        return 1

    # the Photos tab only exists in the --full layout, so this step is skipped
    # rather than failed when it is not there
    if ("Photos" in tabs and os.path.isdir("photos")
            and any(f.startswith("CRH-") for f in os.listdir("photos"))):
        if not run([sys.executable, "embed_photos.py", "--workbook", a.workbook],
                   "embed_photos.py"):
            return 1

    # A card typed straight into Inventory has no SKU, and without one it is
    # invisible to the eBay export, to photo filing and to the dashboard. This
    # only ever fills an EMPTY cell and never renumbers a card that has one, so
    # it is safe to do every run -- and doing it every run is the point, since
    # the failure it prevents is sixty cards silently not existing.
    if not run([sys.executable, "autofill.py", "--workbook", a.workbook, "--go"],
               "autofill.py"):
        return 1

    # a clickable link from each row to that card's pictures. Cheap, and it
    # runs every time so a photo filed since the last run is reachable from
    # the sheet rather than a folder away.
    if os.path.isdir("photos"):
        if not run([sys.executable, "link_photos.py", "--workbook",
                    a.workbook, "--go"], "link_photos.py"):
            return 1

    # a read-only tab per sport, rebuilt from Inventory each time
    if not run([sys.executable, "sport_tabs.py", "--workbook", a.workbook],
               "sport_tabs.py"):
        return 1

    print("\nDone. The workbook is tidied and the game tabs match Inventory.")
    print("Open Card Run HQ - Master.xlsx.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
