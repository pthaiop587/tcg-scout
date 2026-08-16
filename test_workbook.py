"""Tests for the master workbook.

Run: python -m pytest test_workbook.py

Two kinds. The structural ones check the tabs, headers and dropdowns exist,
and run anywhere. The arithmetic ones actually RECALCULATE the workbook with
LibreOffice and read the answers back, and skip if it is not installed.

That second kind earns its keep. openpyxl writes formulas without evaluating
them, so a formula can be perfectly well-formed and point at the wrong row
forever -- which is exactly what happened: the Summary's row references were
counted by hand, a section heading shifted them, and "unrealised gain" was
quietly reading the card count instead of the cost. Nothing but a real
recalculation catches that.
"""

import os
import shutil
import subprocess
import sys
from datetime import date

import pytest
from openpyxl import load_workbook

import embed_photos
import file_batch as fb

SOFFICE = [
    r"C:\Program Files\LibreOffice\program\soffice.exe",
    r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
    "/usr/bin/soffice", "/usr/local/bin/soffice", "soffice",
]

GAMES = ["Football", "Basketball", "Baseball", "Pokemon", "Palworld",
         "One Piece", "Disney"]

# The short layout is the default and the one Mr. P actually uses: the record,
# the upload, a tab per game, and the Read me. Lists is hidden -- it only feeds
# the Graded by dropdown, whose entries are too long for an inline list.
SHORT = ["Read me", "Inventory", "eBay"] + GAMES + ["Lists"]

# --full brings back the money and audit side.
TABS = (["Read me", "Summary", "Inventory", "eBay", "Purchases", "Box log",
         "Expenses", "Sales", "Photos", "Audit", "Reference"] + GAMES + ["Lists"])


def soffice():
    for p in SOFFICE:
        if os.path.isabs(p) and os.path.exists(p):
            return p
        if not os.path.isabs(p) and shutil.which(p):
            return shutil.which(p)
    return None


@pytest.fixture(scope="module")
def blank(tmp_path_factory):
    d = tmp_path_factory.mktemp("wb")
    out = d / "Card Run HQ - Master.xlsx"
    subprocess.run([sys.executable, "make_workbook.py", "--full", "--out", str(out)],
                   check=True, capture_output=True)
    return out


@pytest.fixture
def wb(blank, tmp_path):
    dest = tmp_path / "wb.xlsx"
    shutil.copy(blank, dest)
    return dest


# --- structure --------------------------------------------------------------

def test_every_tab_is_there(blank):
    assert load_workbook(blank).sheetnames == TABS


def test_the_short_layout_is_the_default(tmp_path):
    """Ten tabs and no more: the record, the upload, one per game, the guide."""
    out = tmp_path / "short.xlsx"
    subprocess.run([sys.executable, "make_workbook.py", "--out", str(out)],
                   check=True, capture_output=True)
    book = load_workbook(out)
    assert book.sheetnames == SHORT
    visible = [n for n in book.sheetnames if book[n].sheet_state == "visible"]
    assert len(visible) == 10, visible
    assert book["Lists"].sheet_state == "hidden",         "Lists feeds a dropdown; it should not be a tab anybody sees"


def test_a_tab_exists_for_every_game_before_any_cards(tmp_path):
    """You want the tab when you start on a game, not after the first card."""
    out = tmp_path / "short.xlsx"
    subprocess.run([sys.executable, "make_workbook.py", "--out", str(out)],
                   check=True, capture_output=True)
    book = load_workbook(out)
    for game in GAMES:
        assert game in book.sheetnames, game
        assert str(book[game]["A1"].value or "").startswith("VIEW")


def test_the_sport_dropdown_matches_the_game_tabs(tmp_path):
    """A value with no tab, or a tab with no value, is how a card gets lost."""
    out = tmp_path / "short.xlsx"
    subprocess.run([sys.executable, "make_workbook.py", "--out", str(out)],
                   check=True, capture_output=True)
    ws = load_workbook(out)["Inventory"]
    lists = " ".join(d.formula1 for d in ws.data_validations.dataValidation)
    for game in GAMES:
        assert game in lists, game


def test_purchases_records_a_receipt(blank):
    """A purchase nobody can produce a receipt for is a problem in April."""
    hdr = [c.value for c in load_workbook(blank)["Purchases"][1]]
    for col in ("Date", "Vendor / store", "What", "Total paid", "Tax",
                "Order / receipt #", "Receipt file", "Lot ID", "Paid with"):
        assert col in hdr, col


def test_expenses_covers_the_costs_that_are_not_cards(blank):
    hdr = [c.value for c in load_workbook(blank)["Expenses"][1]]
    for col in ("Date", "Category", "Amount", "Receipt file", "Deductible"):
        assert col in hdr, col


def test_photos_tab_has_room_for_the_picture_and_the_link(blank):
    """Both, and for different reasons -- see embed_photos.py."""
    hdr = [c.value for c in load_workbook(blank)["Photos"][1]]
    assert "Picture" in hdr, "nowhere to put the thumbnail"
    assert "Picture URL for eBay" in hdr, "nowhere to put the link eBay fetches"
    assert "Front" in hdr and "Back" in hdr


def test_the_upload_tab_is_called_ebay(blank):
    book = load_workbook(blank)
    assert "eBay" in book.sheetnames
    assert "eBay upload" not in book.sheetnames


def test_inventory_columns_are_untouched(blank):
    """make_ebay_csv.py finds its columns by header text; renaming one breaks
    the upload silently, so the names are pinned here."""
    hdr = [c.value for c in load_workbook(blank)["Inventory"][1]]
    for col in ("SKU", "Status", "Lot ID", "Qty", "Cost each", "Market value",
                "Ask price", "eBay title", "Card condition", "Serial /"):
        assert col in hdr, col


def test_status_can_be_review(blank):
    """Review is what holds an uncertain card back from the eBay export."""
    ws = load_workbook(blank)["Inventory"]
    lists = [dv.formula1 for dv in ws.data_validations.dataValidation]
    assert any("Review" in f and "Unlisted" in f for f in lists), lists


# --- the numbers, actually calculated ---------------------------------------

def populate(path):
    """A box, a bag of toploaders, two cards out of it, one of them sold."""
    book = load_workbook(path)
    p = book["Purchases"]
    p["A2"], p["B2"], p["C2"] = "PUR-001", date(2026, 8, 15), "Target"
    p["E2"], p["F2"], p["G2"] = "Prizm blaster", "Sealed box", "LOT-001"
    p["H2"], p["I2"], p["K2"] = 1, 29.99, 2.62
    p["P2"] = "receipts/target.jpg"

    e = book["Expenses"]
    e["A2"], e["B2"], e["E2"] = date(2026, 8, 15), "Supplies", 12.49

    s = book["Sales"]
    s["A2"], s["B2"], s["D2"] = date(2026, 8, 20), "CRH-0002", "eBay"
    s["E2"], s["F2"] = 7.99, 1.09
    s["H2"], s["I2"], s["J2"], s["L2"] = 1.19, 1.05, 0.35, 6.50
    book.save(path)

    fb.add_rows(str(path), [
        {"player": "Shedeur Sanders", "year": 2025, "brand": "Prizm",
         "num": "8", "cost": 6.50, "market": 12.50, "ask": 14.99, "lot": "LOT-001"},
        {"player": "Jonah Coleman", "year": 2025, "brand": "Prizm",
         "num": "169", "cost": 6.50, "market": 6.00, "ask": 7.99, "lot": "LOT-001"},
    ])


def recalc(path, tmp_path):
    exe = soffice()
    out = tmp_path / "out"
    out.mkdir(exist_ok=True)
    subprocess.run([exe, "--headless", "--calc", "--convert-to", "xlsx",
                    "--outdir", str(out), str(path)],
                   capture_output=True, timeout=240)
    done = out / (os.path.splitext(os.path.basename(path))[0] + ".xlsx")
    assert done.exists(), "LibreOffice wrote nothing"
    return load_workbook(done, data_only=True)


def summary(book):
    ws = book["Summary"]
    return {ws["A%d" % r].value: ws["B%d" % r].value
            for r in range(2, 30) if ws["A%d" % r].value}


needs_soffice = pytest.mark.skipif(soffice() is None,
                                   reason="LibreOffice not installed")


@needs_soffice
def test_no_formula_anywhere_returns_an_error(wb, tmp_path):
    populate(wb)
    book = recalc(wb, tmp_path)
    bad = []
    for name in book.sheetnames:
        for row in book[name].iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("#"):
                    bad.append("%s!%s = %s" % (name, c.coordinate, c.value))
    assert not bad, bad


@needs_soffice
def test_the_summary_adds_up(wb, tmp_path):
    populate(wb)
    s = summary(recalc(wb, tmp_path))
    assert round(s["Spent on stock"], 2) == 32.61, "29.99 plus 2.62 tax"
    assert round(s["Spent on everything else"], 2) == 12.49
    assert round(s["Total out of pocket"], 2) == 45.10
    assert s["Cards on hand"] == 2
    assert round(s["What they cost"], 2) == 13.00
    assert round(s["What they are worth"], 2) == 18.50
    assert round(s["Unrealised gain"], 2) == 5.50, \
        "worth minus cost -- this read the card count once"
    assert s["Ready to list"] == 2
    assert round(s["Gross"], 2) == 9.08, "sale price plus what they paid to ship"
    assert round(s["Net received"], 2) == 6.49
    assert round(s["Profit against everything spent"], 2) == -45.11, \
        "sold profit minus every pound out, not net minus spend"


@needs_soffice
def test_the_audit_notices_a_card_that_sold_but_is_still_in_stock(wb, tmp_path):
    """The way a card gets listed twice."""
    populate(wb)
    book = recalc(wb, tmp_path)
    ws = book["Audit"]
    found = {ws["A%d" % r].value: ws["B%d" % r].value for r in range(3, 25)
             if ws["A%d" % r].value}
    assert found["Sold but still Unlisted or Listed"] == 1


@needs_soffice
def test_the_audit_is_quiet_on_a_clean_workbook(wb, tmp_path):
    book = recalc(wb, tmp_path)
    ws = book["Audit"]
    for r in range(3, 25):
        label, count = ws["A%d" % r].value, ws["B%d" % r].value
        if label and isinstance(count, (int, float)):
            assert count == 0, "%s reported %s on an empty workbook" % (label, count)


# --- photos into the sheet --------------------------------------------------

def test_photos_are_embedded_not_linked(wb, tmp_path):
    """A thumbnail you can see, at a size that still opens."""
    photos = tmp_path / "photos"
    photos.mkdir()
    from PIL import Image
    for name in ("CRH-0001.jpg", "CRH-0001-back.jpg", "CRH-0002.jpg"):
        Image.new("RGB", (750, 1050), (180, 140, 60)).save(photos / name)

    fb.add_rows(str(wb), [{"player": "Shedeur Sanders", "brand": "Prizm"},
                          {"player": "Jonah Coleman", "brand": "Prizm"}])
    rc = subprocess.run([sys.executable, "embed_photos.py",
                         "--workbook", str(wb), "--photos", str(photos)],
                        capture_output=True, text=True)
    assert rc.returncode == 0, rc.stdout + rc.stderr

    ws = load_workbook(wb)["Photos"]
    assert len(ws._images) == 2, "one thumbnail per card, from its front"

    hdr = [c.value for c in ws[1]]
    rows = {}
    for row in ws.iter_rows(min_row=2, max_row=4, values_only=True):
        if row[0]:
            rows[row[0]] = dict(zip(hdr, row))
    assert rows["CRH-0001"]["How many"] == 2
    assert rows["CRH-0002"]["How many"] == 1
    assert rows["CRH-0001"]["Back"] == "CRH-0001-back.jpg"
    assert "Shedeur Sanders" in rows["CRH-0001"]["Card"]
    assert rows["CRH-0001"]["Picture URL for eBay"].startswith("https://")
    assert "CRH-0001.jpg" in rows["CRH-0001"]["Picture URL for eBay"]


def test_rerunning_photos_does_not_stack_up_images(wb, tmp_path):
    """It rewrites from disk, so it can never drift from the files."""
    photos = tmp_path / "photos"
    photos.mkdir()
    from PIL import Image
    Image.new("RGB", (750, 1050), (60, 90, 200)).save(photos / "CRH-0001.jpg")
    for _ in range(3):
        subprocess.run([sys.executable, "embed_photos.py",
                        "--workbook", str(wb), "--photos", str(photos)],
                       capture_output=True, check=True)
    ws = load_workbook(wb)["Photos"]
    assert len(ws._images) == 1, "three runs left %d images" % len(ws._images)


def test_a_photo_with_no_card_is_reported_not_dropped(wb, tmp_path, capsys):
    photos = tmp_path / "photos"
    photos.mkdir()
    from PIL import Image
    Image.new("RGB", (750, 1050), (200, 60, 60)).save(photos / "CRH-0099.jpg")
    rc = subprocess.run([sys.executable, "embed_photos.py",
                         "--workbook", str(wb), "--photos", str(photos)],
                        capture_output=True, text=True)
    assert "CRH-0099" in rc.stdout
    assert "not on Inventory" in rc.stdout


# --- upgrading an existing workbook -----------------------------------------

def test_upgrade_keeps_what_was_typed(wb, tmp_path):
    """make_workbook --force throws the data away; this must not."""
    fb.add_rows(str(wb), [{"player": "Shedeur Sanders", "brand": "Prizm",
                           "num": "8", "cost": 6.5, "lot": "LOT-001"}])
    book = load_workbook(wb)
    s = book["Sales"]
    s["A2"], s["B2"], s["E2"] = date(2026, 8, 20), "CRH-0001", 12.0
    book.save(wb)

    rc = subprocess.run([sys.executable, "upgrade_workbook.py",
                         "--workbook", str(wb), "--full", "--go"],
                        capture_output=True, text=True)
    assert rc.returncode == 0, rc.stdout + rc.stderr

    after = load_workbook(wb)
    assert after.sheetnames == TABS

    inv = after["Inventory"]
    hdr = [c.value for c in inv[1]]
    row = {h: v for h, v in zip(hdr, next(inv.iter_rows(min_row=2, values_only=True)))}
    assert row["SKU"] == "CRH-0001"
    assert row["Player or card name"] == "Shedeur Sanders"
    assert row["Cost each"] == 6.5, "a value landed in a different column"
    assert row["Lot ID"] == "LOT-001"

    sold = next(after["Sales"].iter_rows(min_row=2, values_only=True))
    assert sold[1] == "CRH-0001", "the sale did not come across"


def test_upgrade_does_not_duplicate_the_seeded_box(wb, tmp_path):
    """A fresh workbook already holds the one box bought, straight off its
    receipt. Carrying the old copy on top counted it twice."""
    def boxes(p):
        return len([r for r in load_workbook(p)["Box log"]
                    .iter_rows(min_row=2, max_row=6, values_only=True) if r[0]])
    before = boxes(wb)
    subprocess.run([sys.executable, "upgrade_workbook.py",
                    "--workbook", str(wb), "--full", "--go"],
                   capture_output=True, check=True)
    assert boxes(wb) == before, "the seeded box was duplicated"


def test_upgrade_backs_up_before_touching_anything(wb):
    subprocess.run([sys.executable, "upgrade_workbook.py",
                    "--workbook", str(wb), "--full", "--go"],
                   capture_output=True, check=True)
    backups = [f for f in os.listdir(os.path.dirname(str(wb))) if "backup" in f]
    assert backups, "no backup was written"


def test_dry_run_changes_nothing(wb):
    before = os.path.getsize(wb)
    rc = subprocess.run([sys.executable, "upgrade_workbook.py",
                         "--workbook", str(wb)], capture_output=True, text=True)
    assert "dry run" in rc.stdout
    assert os.path.getsize(wb) == before


# --- SKUs for hand-typed cards ----------------------------------------------

def typed_cards(path, names, start_row=2):
    """Rows typed straight into Inventory: a name, no SKU."""
    book = load_workbook(path)
    ws = book["Inventory"]
    hdr = [c.value for c in ws[1]]
    name_col = hdr.index("Player or card name") + 1
    for i, n in enumerate(names):
        ws.cell(row=start_row + i, column=name_col, value=n)
    book.save(path)


def skus(path):
    ws = load_workbook(path)["Inventory"]
    hdr = [c.value for c in ws[1]]
    s, n = hdr.index("SKU"), hdr.index("Player or card name")
    return [(r[s], r[n]) for r in ws.iter_rows(min_row=2, values_only=True) if r[n]]


def fill(path, *args):
    r = subprocess.run([sys.executable, os.path.abspath("fill_skus.py"),
                        "--workbook", str(path)] + list(args),
                       capture_output=True, text=True)
    assert r.returncode == 0, r.stdout + r.stderr
    return r.stdout


def test_a_hand_typed_card_gets_a_sku(wb):
    """Without one it is invisible to the export, to photos and to the page."""
    typed_cards(wb, ["Jalen Milroe", "Cam Ward"])
    fill(wb, "--go")
    assert skus(wb) == [("CRH-0001", "Jalen Milroe"), ("CRH-0002", "Cam Ward")]


def test_it_carries_on_from_the_highest_already_used(wb):
    fb.add_rows(str(wb), [{"player": "Shedeur Sanders"}])
    typed_cards(wb, ["Jalen Milroe"], start_row=3)
    fill(wb, "--go")
    assert skus(wb) == [("CRH-0001", "Shedeur Sanders"), ("CRH-0002", "Jalen Milroe")]


def test_a_card_that_has_a_sku_is_never_renumbered(wb):
    """Photos on disk and any live listing are named after it."""
    fb.add_rows(str(wb), [{"player": "Shedeur Sanders"}])
    typed_cards(wb, ["Jalen Milroe"], start_row=3)
    fill(wb, "--go")
    fill(wb, "--go")
    fill(wb, "--go")
    assert skus(wb) == [("CRH-0001", "Shedeur Sanders"), ("CRH-0002", "Jalen Milroe")]


def test_a_row_with_no_name_is_not_a_card(wb):
    """A half-typed line should not be handed a permanent number."""
    book = load_workbook(wb)
    ws = book["Inventory"]
    hdr = [c.value for c in ws[1]]
    ws.cell(row=2, column=hdr.index("Qty") + 1, value=1)
    book.save(wb)
    fill(wb, "--go")
    assert skus(wb) == []


def test_dry_run_writes_nothing(wb):
    typed_cards(wb, ["Jalen Milroe"])
    out = fill(wb)
    assert "dry run" in out
    assert skus(wb) == [(None, "Jalen Milroe")]


def test_it_says_so_when_there_is_nothing_to_do(wb):
    fb.add_rows(str(wb), [{"player": "Shedeur Sanders"}])
    assert "already has a SKU" in fill(wb)
