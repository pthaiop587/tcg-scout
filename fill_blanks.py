"""Fill in what a card shares with the rest of its box, from one row you typed.

    python fill_blanks.py --from CRH-0001            # say what it would do
    python fill_blanks.py --from CRH-0001 --go       # do it

Sixty cards out of one box share almost everything: what they cost, where they
came from, what condition they are in, how many of each you have. Typing that
sixty times is how the sixty-first row ends up saying something different from
the other sixty, and a total quietly stops adding up.

So: type ONE row properly, then point this at it.

Two rules keep it safe.

It only ever writes into an EMPTY cell. A row you have already corrected is
never overwritten, so this can be run again after adding cards without undoing
anything.

And it only copies columns on an explicit list of things a box genuinely shares.
That list is an allow-list, not a block-list, because the failure here is
silent and expensive: copy "Card #" across sixty rows and every card claims to
be card 8, copy "Player or card name" and they are all Shedeur Sanders. A new
column added to the workbook later is not shared until somebody says it is.

Market value is handled separately. It is not shared -- each card has its own
-- but it is not typed either: it is what the card is worth raw, which
prices.py has already looked up. --market-from-raw copies that across, so
Profit and the Summary have something to work from.
"""

import argparse
import sys

from openpyxl import load_workbook

WORKBOOK = "Card Run HQ - Master.xlsx"
NAME = "Player or card name"

# What a box genuinely shares. Deliberately an allow-list: anything not named
# here is per-card and stays blank until someone types it.
SHARED = [
    "Status", "Date in", "Source", "Lot ID", "Category", "Sport or game",
    "League", "Year", "Brand / set", "Card condition", "Qty", "Cost each",
]

# Never copied, whatever else changes. Belt and braces on top of the allow-list
# above -- these are the ones that would be worst to get wrong.
NEVER = {
    "SKU", NAME, "Card #", "Serial /", "Parallel", "Insert set", "Team",
    "Graded by", "Grade", "Cert #", "Market value", "Ask price", "eBay title",
    "Notes", "Listed on", "eBay item #", "Sold on", "Sold for", "Fees paid",
    "Raw price", "Raw last sold", "PSA 9 price", "PSA 9 last sold",
    "PSA 10 price", "PSA 10 last sold",
}


def blank(v):
    return v is None or (isinstance(v, str) and not v.strip())


def main():
    p = argparse.ArgumentParser(
        description="Copy the shared details of a box across its cards.")
    p.add_argument("--workbook", default=WORKBOOK)
    p.add_argument("--from", dest="src", required=True,
                   help="SKU of the row you filled in properly")
    p.add_argument("--sport", help="only cards of this sport (default: the "
                                   "template's own sport)")
    p.add_argument("--lot", help="only cards with this Lot ID")
    p.add_argument("--market-from-raw", action="store_true",
                   help="also set Market value to the looked-up raw price")
    p.add_argument("--go", action="store_true", help="write the values in")
    a = p.parse_args()

    wb = load_workbook(a.workbook)
    ws = wb["Inventory"]
    hdr = [c.value for c in ws[1]]
    g = {n: i + 1 for i, n in enumerate(hdr) if n}
    if NAME not in g:
        sys.exit("Inventory has no %r column" % NAME)

    rows = [r for r in range(2, ws.max_row + 1)
            if not blank(ws.cell(row=r, column=g[NAME]).value)]

    src = None
    for r in rows:
        if str(ws.cell(row=r, column=g["SKU"]).value or "").strip().upper() \
                == a.src.strip().upper():
            src = r
            break
    if not src:
        sys.exit("no card with SKU %r" % a.src)

    sport = a.sport or str(ws.cell(row=src, column=g["Sport or game"]).value
                           or "").strip()
    print("template: row %d, %s (%s)"
          % (src, ws.cell(row=src, column=g["SKU"]).value,
             ws.cell(row=src, column=g[NAME]).value))

    copyable = [c for c in SHARED
                if c in g and c not in NEVER
                and not blank(ws.cell(row=src, column=g[c]).value)]
    print("shares  : " + ", ".join(
        "%s=%s" % (c, ws.cell(row=src, column=g[c]).value) for c in copyable))

    targets = []
    for r in rows:
        if r == src:
            continue
        if sport and str(ws.cell(row=r, column=g["Sport or game"]).value
                         or "").strip() != sport:
            continue
        if a.lot and str(ws.cell(row=r, column=g["Lot ID"]).value
                         or "").strip() != a.lot:
            continue
        targets.append(r)
    print("cards   : %d %s card(s) to look at\n" % (len(targets), sport))

    counts = {c: 0 for c in copyable}
    mkt = 0
    for r in targets:
        for c in copyable:
            cell = ws.cell(row=r, column=g[c])
            if blank(cell.value):
                if a.go:
                    cell.value = ws.cell(row=src, column=g[c]).value
                    cell.number_format = ws.cell(row=src,
                                                 column=g[c]).number_format
                counts[c] += 1
    # Market value is not copied from anywhere -- each card has its own, taken
    # from what prices.py looked up. So it applies to the template row as well,
    # which the loop above deliberately skips.
    if a.market_from_raw and "Market value" in g and "Raw price" in g:
        for r in targets + [src]:
            m = ws.cell(row=r, column=g["Market value"])
            raw = ws.cell(row=r, column=g["Raw price"]).value
            if blank(m.value) and not blank(raw):
                if a.go:
                    m.value = raw
                    m.number_format = '"$"#,##0.00'
                mkt += 1

    for c in copyable:
        if counts[c]:
            print("   %-18s %3d blank cell(s)" % (c, counts[c]))
    if a.market_from_raw:
        print("   %-18s %3d from the raw price" % ("Market value", mkt))
    if not any(counts.values()) and not mkt:
        print("   nothing blank -- everything is already filled in.")
        return 0

    if a.go:
        wb.save(a.workbook)
        print("\nSaved %s. Run sport_tabs.py to push it onto the game tabs."
              % a.workbook)
    else:
        print("\nNothing written. Add --go to fill them in.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
