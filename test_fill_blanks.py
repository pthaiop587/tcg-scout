"""Tests for filling a box's shared details, and for reading a school.

Run: python -m pytest test_fill_blanks.py

The dangerous direction here is copying too much. Sixty rows that all share a
cost is right; sixty rows that all share a card number is a ruined inventory
that still looks tidy, and nothing downstream would complain -- make_ebay_csv
would happily list sixty cards as card #8. So the tests that matter are about
what must NOT travel.
"""

import subprocess
import sys

import pytest
from openpyxl import load_workbook

import colleges
import file_batch as fb

CARDS = [
    {"player": "Shedeur Sanders", "year": 2025, "brand": "Prizm Draft Picks",
     "num": "8", "sport": "Football", "league": "NCAA"},
    {"player": "Travis Hunter", "year": 2025, "brand": "Prizm Draft Picks",
     "num": "7", "sport": "Football", "league": "NCAA"},
    {"player": "Arch Manning", "year": 2025, "brand": "Prizm Draft Picks",
     "num": "166", "sport": "Football", "league": "NCAA"},
    {"player": "Cooper Flagg", "year": 2025, "brand": "Prizm",
     "num": "1", "sport": "Basketball", "league": "NCAA"},
]


@pytest.fixture
def wb(tmp_path):
    out = tmp_path / "Card Run HQ - Master.xlsx"
    subprocess.run([sys.executable, "make_workbook.py", "--out", str(out)],
                   check=True, capture_output=True)
    fb.add_rows(str(out), [dict(c) for c in CARDS])

    book = load_workbook(out)
    ws = book["Inventory"]
    hdr = [c.value for c in ws[1]]
    g = {n: i + 1 for i, n in enumerate(hdr) if n}
    # the template row, filled in the way a person would
    ws.cell(row=2, column=g["Cost each"], value=1.26)
    ws.cell(row=2, column=g["Qty"], value=1)
    ws.cell(row=2, column=g["Card condition"], value="Near Mint or Better")
    ws.cell(row=2, column=g["Source"], value="Big 5 Upland")
    ws.cell(row=2, column=g["Team"], value="Colorado Buffaloes")
    # a raw price on two of them, as prices.py would leave it
    start = ws.max_column
    ws.cell(row=1, column=start + 1, value="Raw price")
    ws.cell(row=2, column=start + 1, value=2.92)
    ws.cell(row=3, column=start + 1, value=2.28)
    book.save(out)
    return out


def run(wb, *args):
    r = subprocess.run([sys.executable, "fill_blanks.py", "--workbook",
                        str(wb), "--from", "CRH-0001"] + list(args),
                       capture_output=True, text=True)
    assert r.returncode == 0, r.stdout + r.stderr
    return r.stdout


def sheet(wb):
    ws = load_workbook(wb)["Inventory"]
    hdr = [c.value for c in ws[1]]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[hdr.index("Player or card name")]:
            out.append(dict(zip(hdr, row)))
    return out


# --- what should travel -----------------------------------------------------

def test_the_shared_details_reach_the_other_cards(wb):
    run(wb, "--go")
    rows = sheet(wb)
    football = [r for r in rows if r["Sport or game"] == "Football"]
    assert len(football) == 3
    for r in football:
        assert r["Cost each"] == 1.26
        assert r["Qty"] == 1
        assert r["Card condition"] == "Near Mint or Better"
        assert r["Source"] == "Big 5 Upland"


def test_a_dry_run_writes_nothing(wb):
    out = run(wb)
    assert "Nothing written" in out
    assert all(r["Cost each"] is None for r in sheet(wb)[1:])


# --- what must NOT travel ---------------------------------------------------

def test_the_card_itself_is_never_copied(wb):
    """The whole point. Copy these and every row is the same card."""
    run(wb, "--go")
    rows = sheet(wb)
    assert [r["Player or card name"] for r in rows] == \
        ["Shedeur Sanders", "Travis Hunter", "Arch Manning", "Cooper Flagg"]
    assert [str(r["Card #"]) for r in rows] == ["8", "7", "166", "1"]


def test_the_team_is_not_copied(wb):
    """Shared by a box, but not by its cards -- they went to different schools."""
    run(wb, "--go")
    rows = sheet(wb)
    assert rows[0]["Team"] == "Colorado Buffaloes"
    assert all(r["Team"] is None for r in rows[1:])


def test_a_value_you_already_typed_is_left_alone(wb):
    book = load_workbook(wb)
    ws = book["Inventory"]
    hdr = [c.value for c in ws[1]]
    ws.cell(row=3, column=hdr.index("Card condition") + 1, value="Played")
    book.save(wb)

    run(wb, "--go")
    assert sheet(wb)[1]["Card condition"] == "Played"


def test_another_sport_is_untouched(wb):
    """A basketball card must not inherit the price of a football box.

    Checked on Cost each and Source, which are genuinely empty here --
    file_batch.py already defaults Card condition and Qty on every row it
    adds, so those prove nothing either way."""
    run(wb, "--go")
    bball = [r for r in sheet(wb) if r["Sport or game"] == "Basketball"][0]
    assert bball["Cost each"] is None
    assert bball["Source"] is None


def test_running_it_twice_changes_nothing_more(wb):
    run(wb, "--go")
    first = sheet(wb)
    out = run(wb, "--go")
    assert sheet(wb) == first
    assert "nothing blank" in out


# --- market value from the looked-up price ----------------------------------

def test_market_value_comes_from_the_raw_price(wb):
    run(wb, "--go", "--market-from-raw")
    rows = sheet(wb)
    assert rows[0]["Market value"] == 2.92
    assert rows[1]["Market value"] == 2.28
    assert rows[2]["Market value"] is None      # no raw price looked up


def test_market_value_is_not_touched_without_the_flag(wb):
    run(wb, "--go")
    assert all(r["Market value"] is None for r in sheet(wb))


# --- reading a school off a listing -----------------------------------------

def test_the_longer_school_name_wins():
    """'Texas' sits inside 'Texas A&M' and 'Texas Tech'. Shortest-first would
    send every Aggie and Red Raider to Austin."""
    assert colleges.team_in("Shemar Turner Texas A&M 2025") == "Texas A&M Aggies"
    assert colleges.team_in("Tahj Brooks Texas Tech") == "Texas Tech Red Raiders"
    assert colleges.team_in("Arch Manning Texas") == "Texas Longhorns"
    assert colleges.team_in("Washington State Cougars") == "Washington State Cougars"
    assert colleges.team_in("Washington Huskies") == "Washington Huskies"


def test_a_title_with_no_school_says_so():
    assert colleges.team_in("2025 Prizm Draft Picks Silver #15") is None
    assert colleges.team_in("") is None


def test_the_commonest_school_wins_the_vote():
    """One seller naming the opponent should not decide it."""
    school, n = colleges.vote([
        "Shedeur Sanders Colorado Buffaloes",
        "Shedeur Sanders Colorado",
        "Shedeur Sanders vs Nebraska",
    ])
    assert school == "Colorado Buffaloes"
    assert n == 2


def test_no_school_anywhere_is_reported_not_guessed():
    assert colleges.vote(["nothing here", "or here"]) == (None, 0)
