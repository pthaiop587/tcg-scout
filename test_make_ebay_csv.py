"""Tests for the eBay export, mostly the title.

Run: python -m pytest test_make_ebay_csv.py

The title is the listing. It is what a buyer searches, and eBay gives you 80
characters and silently wastes anything past that, so the two things worth
guarding are that the words a buyer would type are present and that the line
fits. Everything else on a listing can be edited after the fact; a title that
nobody searches for just sits there.
"""

import csv
import subprocess
import sys

import pytest
from openpyxl import load_workbook

import make_ebay_csv as m


def card(**kw):
    d = {"year": "2025", "brand": "Panini Prizm Draft Picks", "insert": "",
         "parallel": "Gold Ice", "player": "Arch Manning", "num": "166",
         "serial": "", "rc": True, "auto": False, "relic": False,
         "cat": "Sports", "sport": "Football", "team": "Texas Longhorns"}
    d.update(kw)
    return d


# --- what a buyer types -----------------------------------------------------

def test_a_pokemon_title_says_pokemon():
    """"2026 Pitch Black Secret Rare Misty's Vitality" is invisible to anyone
    searching for a Pokemon card, which is everyone who would buy it."""
    t = m.build_title(card(cat="TCG", sport="Pokemon", brand="Pitch Black",
                           year="2026", parallel="Secret Rare",
                           player="Misty's Vitality", num="111/084",
                           rc=False, team=""))
    assert t.startswith("Pokemon ")
    assert "Pitch Black" in t and "Misty's Vitality" in t
    assert "111/084" in t


def test_the_game_is_not_repeated_when_the_set_already_names_it():
    t = m.build_title(card(cat="TCG", sport="Pokemon",
                           brand="Pokemon Pitch Black", player="Pikachu",
                           num="1", rc=False, parallel="", team=""))
    assert t.lower().count("pokemon") == 1


def test_a_sports_title_is_unchanged_by_that():
    """Sports titles already carry the maker, which does the same job."""
    t = m.build_title(card())
    assert t == "2025 Panini Prizm Draft Picks Gold Ice Arch Manning #166 RC"


# --- the 80-character ceiling -----------------------------------------------

def test_a_long_title_is_cut_down_rather_than_run_over():
    t = m.build_title(card(insert="Instant Impact", player="James Pearce Jr.",
                           num="II-JPJ"))
    assert len(t) <= 80
    assert "James Pearce Jr." in t, "the player is never what gets dropped"


def test_the_player_and_the_set_always_survive():
    t = m.build_title(card(insert="Student Orientation",
                           player="Shedeur Sanders", num="II-SSS",
                           serial="49", auto=True, relic=True))
    assert len(t) <= 80
    assert "Shedeur Sanders" in t
    assert "Prizm" in t


def test_the_game_survives_a_squeeze_too():
    """It is prefixed with priority None, so a long Pokemon card keeps it."""
    t = m.build_title(card(cat="TCG", sport="Pokemon",
                           brand="Pitch Black Ultra Premium Collection",
                           parallel="Special Illustration Rare",
                           player="Misty's Vitality And Friends",
                           num="111/084", serial="150", rc=False, team=""))
    assert len(t) <= 80
    assert t.startswith("Pokemon ")


# --- the file ---------------------------------------------------------------

def test_the_export_writes_the_ask_price_as_the_start_price(tmp_path):
    wb = tmp_path / "Card Run HQ - Master.xlsx"
    subprocess.run([sys.executable, "make_workbook.py", "--out", str(wb)],
                   check=True, capture_output=True)
    book = load_workbook(wb)
    ws = book["Inventory"]
    hdr = [c.value for c in ws[1]]
    g = {n: i + 1 for i, n in enumerate(hdr) if n}
    for col, val in (("SKU", "CRH-0001"), ("Status", "Unlisted"),
                     ("Player or card name", "Arch Manning"),
                     ("Year", 2025), ("Brand / set", "Panini Prizm"),
                     ("Card #", "166"), ("Ask price", 25.99),
                     ("Sport or game", "Football"), ("Category", "Sports")):
        ws.cell(row=2, column=g[col]).value = val
    book.save(wb)

    out = tmp_path / "e.csv"
    r = subprocess.run([sys.executable, "make_ebay_csv.py", "--workbook",
                        str(wb), "-o", str(out)], capture_output=True,
                       text=True)
    assert r.returncode == 0, r.stdout + r.stderr
    with open(out, newline="", encoding="utf-8-sig") as fh:
        rows = list(csv.DictReader(fh))
    assert len(rows) == 1
    assert rows[0]["StartPrice"] == "25.99"
    assert "Arch Manning" in rows[0]["Title"]
    assert rows[0]["CustomLabel"] == "CRH-0001"


def test_the_postage_charge_reaches_the_file(tmp_path):
    """$1.50 to the buyer is what makes a dollar card worth listing. If it
    does not reach the CSV the listing goes up with free postage and every
    cheap card sells at a loss."""
    wb = tmp_path / "Card Run HQ - Master.xlsx"
    subprocess.run([sys.executable, "make_workbook.py", "--out", str(wb)],
                   check=True, capture_output=True)
    book = load_workbook(wb)
    ws = book["Inventory"]
    hdr = [c.value for c in ws[1]]
    g = {n: i + 1 for i, n in enumerate(hdr) if n}
    for col, val in (("SKU", "CRH-0001"), ("Status", "Unlisted"),
                     ("Player or card name", "Arch Manning"),
                     ("Ask price", 13.0), ("Sport or game", "Football")):
        ws.cell(row=2, column=g[col]).value = val
    book.save(wb)

    out = tmp_path / "e.csv"
    r = subprocess.run([sys.executable, "make_ebay_csv.py", "--workbook",
                        str(wb), "-o", str(out)], capture_output=True,
                       text=True)
    assert r.returncode == 0, r.stdout + r.stderr
    with open(out, newline="", encoding="utf-8-sig") as fh:
        row = next(iter(csv.DictReader(fh)))
    assert row["ShippingService-1:Cost"] == "1.50"
    assert row["ShippingType"] == "Flat"
    assert row["ShippingCostPaidByOption"] == "Buyer"


def test_the_pricer_and_the_exporter_agree_on_the_postage():
    """Two files each holding their own copy of 1.50 is two files that can
    disagree. If they ever do, the margin in the sheet is fiction."""
    import price_listings
    assert m.SHIP_COST == price_listings.SHIP_CHARGE


# --- a TCG card is not a sports card ----------------------------------------

def test_a_pokemon_card_is_not_typed_as_a_sports_card():
    """C:Type was a global default, so every Pokemon listing claimed to be a
    Sports Trading Card. eBay believes item specifics."""
    v = m.values_for(dict(card(cat="TCG", sport="Pokemon", player="Sinistcha",
                               brand="Pitch Black", parallel="Uncommon"),
                          sku="CRH-0136", title="t", graded=False, grader="",
                          grade="", cert="", condition="Near Mint or Better",
                          grade_txt="", qty="1", ask="2", league="",
                          num="006/084",
                          serial="", auto=False, relic=False, rc=False))
    assert v.get("spec:type", "") == ""
    assert v.get("C:Type", "") == "", "the default must not leak back in"


def test_a_sports_card_still_is_one():
    v = m.values_for(dict(card(), sku="CRH-0060", title="t", graded=False,
                          grader="", grade="", cert="",
                          grade_txt="",
                          condition="Near Mint or Better", qty="1", ask="13",
                          league="NCAA", serial="", relic=False))
    assert v["spec:type"] == "Sports Trading Card"


def test_a_tcg_card_answers_the_questions_that_category_asks():
    """CCG Singles asks Game and Rarity; it has no Sport or Parallel field."""
    v = m.values_for(dict(card(cat="TCG", sport="Pokemon",
                               player="Misty's Vitality", brand="Pitch Black",
                               parallel="Secret Rare"),
                          sku="CRH-0122", title="t", graded=False, grader="",
                          grade="", cert="", condition="Near Mint or Better",
                          grade_txt="", qty="1", ask="19", league="",
                          num="111/084",
                          serial="", auto=False, relic=False, rc=False))
    assert v["spec:game"] == "Pokemon"
    assert v["spec:rarity"] == "Secret Rare"
    assert v["spec:cardname"] == "Misty's Vitality"


def test_a_sports_card_gains_no_game_or_rarity():
    v = m.values_for(dict(card(), sku="CRH-0060", title="t", graded=False,
                          grader="", grade="", cert="",
                          grade_txt="",
                          condition="Near Mint or Better", qty="1", ask="13",
                          league="NCAA", serial="", relic=False))
    assert v.get("spec:game", "") == ""
    assert v["spec:parallelvariety"] == "Gold Ice"
