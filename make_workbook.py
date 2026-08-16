"""Build the Card Run HQ master workbook.

One file, several tabs, meant to be edited by hand in Excel or Google Sheets.
Inventory is the master record; make_ebay_csv.py reads it and writes the
upload file. Everything eBay-specific lives in Reference so the codes can be
checked against eBay's own documentation rather than trusted blindly.

    python make_workbook.py            # writes Card Run HQ - Master.xlsx
    python make_workbook.py --force    # overwrite even if the file exists

Refuses to clobber an existing workbook without --force, because the whole
point is that Mr. P types into it.
"""

import argparse
import os
import sys
from datetime import date

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = "Card Run HQ - Master.xlsx"

# ---------------------------------------------------------------- eBay codes
# Verified Aug 2026. Sources are written into the Reference tab so they can be
# re-checked -- eBay moved bulk upload off File Exchange into Seller Hub
# Reports and the codes have changed before.
CAT_SPORTS = 261328   # Sports Trading Card Singles
CAT_NONSPT = 183050   # Non-Sport Trading Card Singles
CAT_CCG    = 183454   # CCG Individual Cards (Pokemon, Magic, Lorcana, ...)

COND_GRADED   = 2750
COND_UNGRADED = 4000

# CD:Card Condition - (ID: 40001). eBay wants the text label in the cell; the
# numeric id is kept alongside only so the mapping is auditable.
CARD_CONDITIONS = [
    ("Near Mint or Better", 400010),
    ("Excellent",           400011),
    ("Very Good",           400012),
    ("Poor",                400013),
]

GRADERS = [
    "Professional Sports Authenticator (PSA)",
    "Beckett Grading Services (BGS)",
    "Sportscard Guaranty Corporation (SGC)",
    "Certified Guaranty Company (CGC)",
    "Hybrid Grading Approach (HGA)",
    "Technical Authentication & Grading (TAG)",
    "Other",
]

# ------------------------------------------------------------------- styling
INK      = "1B1F2A"
ACCENT   = "3B4A6B"
HEADFILL = PatternFill("solid", fgColor=INK)
SUBFILL  = PatternFill("solid", fgColor=ACCENT)
CALCFILL = PatternFill("solid", fgColor="EEF1F6")
NOTEFILL = PatternFill("solid", fgColor="FFF6E5")
HEADFONT = Font(bold=True, color="FFFFFF", size=10)
TITLEFONT = Font(bold=True, size=14, color=INK)
SMALL    = Font(size=9, color="55607A")
HAIR     = Side(style="thin", color="C9D0DE")
BOX      = Border(left=HAIR, right=HAIR, top=HAIR, bottom=HAIR)

MONEY = '"$"#,##0.00'
PCT   = "0.0%"


def head(ws, cols, row=1):
    """Write a header row from (title, width) pairs and freeze beneath it."""
    for i, (name, width) in enumerate(cols, start=1):
        c = ws.cell(row=row, column=i, value=name)
        c.fill = HEADFILL
        c.font = HEADFONT
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = BOX
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[row].height = 30
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def note(ws, cell, text):
    ws[cell] = text
    ws[cell].font = SMALL
    ws[cell].alignment = Alignment(wrap_text=True, vertical="top")


def dv(ws, values, cells, allow_blank=True):
    """Attach a dropdown built from a literal list."""
    joined = ",".join(values)
    if len(joined) > 250:                     # Excel's inline-list ceiling
        raise ValueError("list too long for inline validation: " + joined[:60])
    d = DataValidation(type="list", formula1='"%s"' % joined,
                       allow_blank=allow_blank, showDropDown=False)
    ws.add_data_validation(d)
    d.add(cells)
    return d


def dv_range(ws, ref, cells, allow_blank=True):
    """Dropdown sourced from a range on the Lists sheet."""
    d = DataValidation(type="list", formula1=ref, allow_blank=allow_blank,
                       showDropDown=False)
    ws.add_data_validation(d)
    d.add(cells)
    return d


# --------------------------------------------------------------- Inventory
# Column order is load-bearing: make_ebay_csv.py looks columns up by header
# text, but the formulas below are written with these letters.
INVENTORY_COLS = [
    ("SKU", 12), ("Status", 13), ("Date in", 11), ("Source", 18),
    ("Lot ID", 11), ("Category", 10), ("Sport or game", 14), ("League", 10),
    ("Year", 7), ("Brand / set", 24), ("Insert set", 20), ("Parallel", 18),
    ("Player or card name", 22), ("Card #", 9), ("Serial /", 9),
    ("Team", 18), ("RC", 6), ("Auto", 6), ("Relic", 7),
    ("Graded by", 20), ("Grade", 7), ("Cert #", 13),
    ("Card condition", 18), ("Qty", 6), ("Cost each", 11),
    ("Market value", 12), ("Ask price", 11), ("Est fees", 10),
    ("Est net", 10), ("Margin", 9), ("eBay title", 52), ("Len", 6),
    ("Sold comps", 12), ("Notes", 34),
]

# Column letters are derived from the headers above, never typed. Inserting a
# column used to mean hand-editing every formula and silently getting one
# wrong.
COL = {name: get_column_letter(i)
       for i, (name, _w) in enumerate(INVENTORY_COLS, start=1)}
LAST_COL = get_column_letter(len(INVENTORY_COLS))

INV_ROWS = 400          # pre-formulated rows; add more by dragging down


def build_inventory(wb):
    ws = wb.create_sheet("Inventory")
    head(ws, INVENTORY_COLS)

    c = COL
    for r in range(2, INV_ROWS + 2):
        f = {k: "%s%d" % (v, r) for k, v in c.items()}

        # Fees: eBay final value fee plus the per-order fee, which steps at $10.
        ws[f["Est fees"]] = (
            '=IF(N({ask})=0,"",ROUND({ask}*0.1325,2)+IF({ask}<10,0.3,0.4))'
        ).format(ask=f["Ask price"])
        ws[f["Est net"]] = '=IF(N({ask})=0,"",{ask}-{fee})'.format(
            ask=f["Ask price"], fee=f["Est fees"])
        ws[f["Margin"]] = (
            '=IF(N({net})=0,"",IF(N({cost})=0,"",({net}-{cost})/{net}))'
        ).format(net=f["Est net"], cost=f["Cost each"])

        # Title: year, set, insert, parallel, name, number, then the badges.
        title = (
            'TRIM({yr}&" "&{brand}&" "&{ins}&" "&{par}&" "&{who}'
            '&IF({num}=""," "," #"&{num})'
            '&IF({ser}=""," "," /"&{ser})'
            '&IF({rc}="Yes"," RC","")'
            '&IF({au}="Yes"," AUTO","")'
            '&IF({rel}="Yes"," PATCH",""))'
        ).format(yr=f["Year"], brand=f["Brand / set"], ins=f["Insert set"],
                 par=f["Parallel"], who=f["Player or card name"],
                 num=f["Card #"], ser=f["Serial /"], rc=f["RC"],
                 au=f["Auto"], rel=f["Relic"])
        ws[f["eBay title"]] = '=IF({who}="","",{t})'.format(
            who=f["Player or card name"], t=title)
        ws[f["Len"]] = '=IF({t}="","",LEN({t}))'.format(t=f["eBay title"])

        # One tap to the sold comps for this exact card.
        ws[f["Sold comps"]] = (
            '=IF({who}="","",HYPERLINK('
            '"https://www.ebay.com/sch/i.html?_nkw="'
            '&SUBSTITUTE(TRIM({yr}&" "&{brand}&" "&{ins}&" "&{par}&" "&{who}'
            '&" "&{num})," ","+")'
            '&"&_sacat=0&LH_Sold=1&LH_Complete=1&_ipg=100","sold"))'
        ).format(who=f["Player or card name"], yr=f["Year"],
                 brand=f["Brand / set"], ins=f["Insert set"],
                 par=f["Parallel"], num=f["Card #"])

        for name in ("Cost each", "Market value", "Ask price",
                     "Est fees", "Est net"):
            ws[f[name]].number_format = MONEY
        ws[f["Margin"]].number_format = PCT
        ws[f["Date in"]].number_format = "yyyy-mm-dd"
        for name in ("Est fees", "Est net", "Margin", "eBay title", "Len",
                     "Sold comps"):
            ws[f[name]].fill = CALCFILL

    last = INV_ROWS + 1

    def span(name):
        return "{c}2:{c}{last}".format(c=c[name], last=last)

    dv(ws, ["Unlisted", "Listed", "Sold", "Kept", "At COMC", "Bulk"],
       span("Status"))
    dv(ws, ["Sports", "TCG", "Non-sport"], span("Category"))
    dv(ws, ["NCAA", "NFL", "NBA", "MLB", "NHL", "MLS", "WWE", "UFC"],
       span("League"))
    for name in ("RC", "Auto", "Relic"):
        dv(ws, ["Yes", "No"], span(name))
    dv(ws, [x[0] for x in CARD_CONDITIONS], span("Card condition"))
    dv_range(ws, "=Lists!$A$2:$A$%d" % (len(GRADERS) + 1), span("Graded by"))

    # A title over 80 characters is rejected on upload, so make it loud.
    ws.conditional_formatting.add(
        span("Len"),
        CellIsRule(operator="greaterThan", formula=["80"],
                   fill=PatternFill("solid", fgColor="FFC7CE"),
                   font=Font(bold=True, color="9C0006")))
    ws.conditional_formatting.add(
        span("Len"),
        CellIsRule(operator="between", formula=["1", "80"],
                   fill=PatternFill("solid", fgColor="D8F0DC")))
    # Negative margin, i.e. selling it loses money.
    ws.conditional_formatting.add(
        span("Margin"),
        CellIsRule(operator="lessThan", formula=["0"],
                   font=Font(bold=True, color="9C0006")))
    ws.auto_filter.ref = "A1:%s%d" % (LAST_COL, last)
    return ws


# ------------------------------------------------------------------ Box log
BOXLOG_COLS = [
    ("Lot ID", 11), ("Date", 11), ("Time", 9), ("Store", 20),
    ("Location", 20), ("Product", 40), ("Qty", 6), ("Unit price", 11),
    ("Subtotal", 11), ("Tax", 9), ("Total paid", 11), ("List price", 11),
    ("Over / under list", 15), ("Cards pulled", 12), ("Cost per card", 12),
    ("Value pulled", 12), ("Net vs cost", 12), ("Best hit", 30), ("Notes", 34),
]


def build_boxlog(wb):
    ws = wb.create_sheet("Box log")
    head(ws, BOXLOG_COLS)

    for r in range(2, 202):
        ws["I%d" % r] = '=IF(N(H{r})=0,"",G{r}*H{r})'.format(r=r)
        ws["K%d" % r] = '=IF(N(I{r})=0,"",I{r}+N(J{r}))'.format(r=r)
        ws["M%d" % r] = '=IF(N(L{r})=0,"",K{r}-L{r}*N(G{r}))'.format(r=r)
        ws["O%d" % r] = '=IF(N(N{r})=0,"",K{r}/N{r})'.format(r=r)
        ws["Q%d" % r] = '=IF(N(P{r})=0,"",P{r}-K{r})'.format(r=r)
        for col in ("H", "I", "J", "K", "L", "M", "O", "P", "Q"):
            ws["%s%d" % (col, r)].number_format = MONEY
        ws["B%d" % r].number_format = "yyyy-mm-dd"
        for col in ("I", "K", "M", "O", "Q"):
            ws["%s%d" % (col, r)].fill = CALCFILL

    ws.conditional_formatting.add(
        "Q2:Q201",
        CellIsRule(operator="lessThan", formula=["0"],
                   font=Font(bold=True, color="9C0006")))
    ws.conditional_formatting.add(
        "Q2:Q201",
        CellIsRule(operator="greaterThan", formula=["0"],
                   font=Font(bold=True, color="1E6B33")))

    # The one box actually bought so far, straight off the receipt.
    ws.append([])  # keeps append from landing on row 2 formulas
    seed = {
        "A2": "LOT-001", "B2": date(2026, 8, 15), "D2": "Big 5",
        "E2": "Upland, CA",
        "F2": "2025 Panini Prizm Draft Picks College Football - Mega Box",
        "G2": 1, "H2": 69.99, "J2": 5.42, "L2": 55.99, "N2": 60,
        "S2": "6 packs x 10 cards. 18 mega-box exclusives per box.",
    }
    for ref, val in seed.items():
        ws[ref] = val
    ws["B2"].number_format = "yyyy-mm-dd"
    for col in ("H", "J", "L"):
        ws["%s2" % col].number_format = MONEY
    ws.auto_filter.ref = "A1:S201"
    return ws


# -------------------------------------------------------------------- Sales
SALES_COLS = [
    ("Date sold", 11), ("SKU", 12), ("What it was", 40), ("Channel", 14),
    ("Sale price", 11), ("Shipping charged", 15), ("Gross", 11),
    ("Fees", 10), ("Shipping cost", 13), ("Supplies", 10), ("Net", 11),
    ("Cost basis", 11), ("Profit", 11), ("Margin", 9), ("Notes", 34),
]


def build_sales(wb):
    ws = wb.create_sheet("Sales")
    head(ws, SALES_COLS)
    for r in range(2, 302):
        ws["G%d" % r] = '=IF(N(E{r})=0,"",E{r}+N(F{r}))'.format(r=r)
        ws["K%d" % r] = ('=IF(N(G{r})=0,"",G{r}-N(H{r})-N(I{r})-N(J{r}))'
                         ).format(r=r)
        ws["M%d" % r] = '=IF(N(K{r})=0,"",K{r}-N(L{r}))'.format(r=r)
        ws["N%d" % r] = '=IF(N(K{r})=0,"",M{r}/K{r})'.format(r=r)
        for col in ("E", "F", "G", "H", "I", "J", "K", "L", "M"):
            ws["%s%d" % (col, r)].number_format = MONEY
        ws["N%d" % r].number_format = PCT
        ws["A%d" % r].number_format = "yyyy-mm-dd"
        for col in ("G", "K", "M", "N"):
            ws["%s%d" % (col, r)].fill = CALCFILL
    dv(ws, ["eBay", "TCGplayer", "COMC", "Whatnot", "Local", "Other"],
       "D2:D301")
    ws.conditional_formatting.add(
        "M2:M301",
        CellIsRule(operator="lessThan", formula=["0"],
                   font=Font(bold=True, color="9C0006")))
    ws.auto_filter.ref = "A1:O301"
    return ws


# ---------------------------------------------------------------- Reference
def build_reference(wb):
    ws = wb.create_sheet("Reference")
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 62
    ws["A1"] = "eBay codes, checked August 2026"
    ws["A1"].font = TITLEFONT
    r = 3

    def block(title, rows, cols=("What", "Code", "Notes")):
        nonlocal r
        ws.cell(row=r, column=1, value=title).font = Font(bold=True, size=11)
        r += 1
        for i, name in enumerate(cols, start=1):
            c = ws.cell(row=r, column=i, value=name)
            c.fill = SUBFILL
            c.font = HEADFONT
        r += 1
        for row in rows:
            for i, val in enumerate(row, start=1):
                cell = ws.cell(row=r, column=i, value=val)
                cell.border = BOX
                cell.alignment = Alignment(wrap_text=(i == 3), vertical="top")
            r += 1
        r += 1

    block("Category", [
        ("Sports card singles", CAT_SPORTS, "Football, basketball, baseball, hockey, soccer. Sport goes in its own item specific."),
        ("Non-sport card singles", CAT_NONSPT, "Garbage Pail Kids, movie and TV cards."),
        ("CCG individual cards", CAT_CCG, "Pokemon, Magic, Lorcana, One Piece, Yu-Gi-Oh."),
    ])

    block("Condition ID", [
        ("Graded", COND_GRADED, "Slabbed by a grading company. Grader, grade and cert number then become required."),
        ("Ungraded", COND_UNGRADED, "Raw card. Needs the card condition descriptor below. 'Used' is no longer accepted for cards."),
    ])

    block("Card condition descriptor -- CD:Card Condition (ID: 40001)", [
        (name, code, "Required on every ungraded card since 2023.")
        for name, code in CARD_CONDITIONS
    ])

    block("Professional grader -- CD:Professional Grader (ID: 27501)", [
        (g, "", "Graded cards only. Grade goes in ID 27502, cert number in ID 27503.")
        for g in GRADERS
    ])

    block("Fees used by the formulas", [
        ("eBay final value fee", "13.25%", "Trading cards, charged on the item plus the shipping the buyer pays."),
        ("eBay per-order fee", "$0.30 / $0.40", "$0.30 when the order is under $10, otherwise $0.40."),
        ("eBay insertion", "$0.35", "Only past the 250 free listings a month."),
        ("eBay Standard Envelope", "$0.74 - $1.32", "Raw cards under $20. Tracked, and the cheapest way to move a single."),
        ("TCGplayer", "10.75% + 2.5% + $0.30", "The $0.30 is per order, not per card, so a cart of singles amortises it. TCG only."),
        ("COMC", "5% + 10% cash out", "Plus roughly $0.50-1.00 a card to submit. Worth it on volumes of cheap sports singles."),
    ])

    block("Where these came from", [
        ("Condition IDs and descriptors", "", "eBay Developers, Condition Descriptor IDs; eBay community thread on Seller Hub reports for trading card ungraded values."),
        ("Category IDs", "", "eBay's own browse URLs, e.g. ebay.com/b/Trading-Card-Singles/261328."),
        ("Bulk upload route", "", "File Exchange is retired. Uploads now go through Seller Hub, Reports, Uploads."),
        ("Header row", "", "The header row is tied to your account and template version. Download the template from your own Seller Hub and let make_ebay_csv.py match it."),
    ])
    return ws


def build_lists(wb):
    ws = wb.create_sheet("Lists")
    ws["A1"] = "Graders"
    for i, g in enumerate(GRADERS, start=2):
        ws["A%d" % i] = g
    ws["C1"] = "Card conditions"
    for i, (name, _code) in enumerate(CARD_CONDITIONS, start=2):
        ws["C%d" % i] = name
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["C"].width = 22
    ws.sheet_state = "hidden"
    return ws


# ------------------------------------------------------------------ Read me
README = [
    ("Inventory",
     "Every card you own, one row each. This is the master -- everything else "
     "reads from it. Type into the white columns; the shaded ones work "
     "themselves out. The eBay title builds itself and Len turns red past 80 "
     "characters, which is eBay's hard limit."),
    ("eBay upload",
     "Written for you by make_ebay_csv.py. Do not type here -- it gets "
     "rebuilt. It holds only the Inventory rows marked Unlisted."),
    ("Box log",
     "Every box and pack bought: where, when, what it cost with tax, how many "
     "cards came out and what they were worth. Net vs cost is the number that "
     "says whether ripping beat selling it sealed."),
    ("Sales",
     "What actually sold and what was left after fees. Profit is measured "
     "against cost basis, so allocate the box cost across the cards it "
     "produced -- Cost per card on the Box log gives you that figure."),
    ("Reference",
     "The eBay codes the upload depends on, with where each one came from. "
     "Check this tab first if an upload gets rejected."),
]

STEPS = [
    "Photograph the card, front and back, and send it to Claude.",
    "Claude reads the set, number and parallel off the card and gives you a row for Inventory.",
    "Tap the 'sold' link in that row to see what the card actually sells for, and put that in Market value.",
    "Set Ask price. Est fees, Est net and Margin fill in on their own.",
    "Run: python make_ebay_csv.py",
    "In eBay, go to Seller Hub, Reports, Uploads, and download the template for Trading Card Singles.",
    "Put that file in this folder as ebay-template.csv and run the script again -- it will match your template's header row exactly.",
    "Upload the CSV it writes, then check the results report on the same page.",
]


def build_readme(wb):
    ws = wb.create_sheet("Read me", 0)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 96
    ws["A1"] = "Card Run HQ -- master workbook"
    ws["A1"].font = TITLEFONT
    ws["A2"] = "Built %s. Inventory is the master record; the eBay upload is generated from it." % date.today().isoformat()
    ws["A2"].font = SMALL

    r = 4
    ws.cell(row=r, column=1, value="Photo to listing").font = Font(bold=True, size=11)
    r += 1
    for i, step in enumerate(STEPS, start=1):
        ws.cell(row=r, column=1, value=str(i)).font = Font(bold=True, color=ACCENT)
        note(ws, "B%d" % r, step)
        ws.row_dimensions[r].height = 15
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="The tabs").font = Font(bold=True, size=11)
    r += 1
    for name, what in README:
        c = ws.cell(row=r, column=1, value=name)
        c.font = Font(bold=True)
        c.alignment = Alignment(vertical="top")
        note(ws, "B%d" % r, what)
        ws.row_dimensions[r].height = 44
        r += 1

    r += 1
    warn = ws.cell(row=r, column=1, value="Read this before the first upload")
    warn.font = Font(bold=True, size=11, color="9C0006")
    r += 1
    for line in [
        "eBay retired File Exchange. Bulk uploads now go through Seller Hub, "
        "Reports, Uploads.",
        "The header row of an upload file is tied to your account and to the "
        "template version, so a header typed from documentation can be "
        "rejected even when every value in it is right. Download your own "
        "template once, save it in this folder as ebay-template.csv, and the "
        "script will conform to it.",
        "Upload one card the first time. Read the results report. Then do the "
        "rest in bulk.",
    ]:
        note(ws, "B%d" % r, line)
        ws["B%d" % r].fill = NOTEFILL
        ws.row_dimensions[r].height = 44
        r += 1
    return ws


def build_upload_placeholder(wb):
    ws = wb.create_sheet("eBay upload")
    ws.column_dimensions["A"].width = 100
    ws["A1"] = "Generated -- do not type here"
    ws["A1"].font = TITLEFONT
    note(ws, "A2", "Run  python make_ebay_csv.py  and this tab is rebuilt from "
                   "the Inventory rows marked Unlisted, alongside the CSV that "
                   "gets uploaded.")
    return ws


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--force", action="store_true",
                    help="overwrite an existing workbook")
    ap.add_argument("-o", "--out", default=OUT)
    args = ap.parse_args()

    if os.path.exists(args.out) and not args.force:
        sys.exit("%s already exists. Use --force if you really mean to "
                 "replace it -- you will lose anything typed in." % args.out)

    wb = Workbook()
    wb.remove(wb.active)
    build_readme(wb)
    build_inventory(wb)
    build_upload_placeholder(wb)
    build_boxlog(wb)
    build_sales(wb)
    build_reference(wb)
    build_lists(wb)
    wb.save(args.out)
    print("wrote %s (%d tabs)" % (args.out, len(wb.sheetnames)))
    print("tabs: " + ", ".join(wb.sheetnames))


if __name__ == "__main__":
    main()
