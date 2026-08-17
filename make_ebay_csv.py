"""Turn the Inventory tab into an eBay bulk-upload CSV.

    python make_ebay_csv.py                  # every Unlisted row
    python make_ebay_csv.py --sku CRH-0001   # just one
    python make_ebay_csv.py --all            # ignore Status

eBay retired File Exchange; uploads now go through Seller Hub, Reports,
Uploads. The header row there is tied to your account and template version,
so a header typed out of documentation can be rejected even when every value
under it is correct. Drop your own downloaded template in this folder as
ebay-template.csv and this script conforms to it column for column, filling
what it recognises and leaving the rest blank for you to check.

Without a template it falls back to the documented File Exchange header,
which Seller Hub still accepts but which is a best effort, not a guarantee.
"""

import argparse
import csv
import json
import os
import re
import sys
from datetime import date

from openpyxl import load_workbook

import inuse

WORKBOOK = "Card Run HQ - Master.xlsx"
TEMPLATE = "ebay-template.csv"
# Written by ebay_pics.py: SKU -> the URLs eBay hosts for it.
EPS_URLS = "photos/eps-urls.json"
PAGES    = "https://pthaiop587.github.io/tcg-scout"

CAT = {"Sports": 261328, "Non-sport": 183050, "TCG": 183454}
COND_GRADED, COND_UNGRADED = 2750, 4000

# Postage charged to the buyer for a single card. It is not a detail: eBay
# takes its percentage of the postage as well, and charging $1.50 against a
# real cost of about $0.70 is what covers the fixed fee and makes a dollar
# card worth listing at all. price_listings.py assumes the same figure.
SHIP_COST = 1.50

# Defaults for the columns eBay needs but that are not a property of the card.
DEFAULTS = {
    "Format": "FixedPrice",
    "Duration": "GTC",
    "Location": "Upland, CA 91786",
    "DispatchTimeMax": "1",
    "ReturnsAcceptedOption": "ReturnsAccepted",
    "ReturnsWithinOption": "Days_30",
    "RefundOption": "MoneyBack",
    "ShippingCostPaidByOption": "Buyer",
    # Flat $1.50 for one card, which is what Mr. P charges. These loose
    # fields are for accounts NOT on Business Policies; an account that is on
    # them takes postage from ShippingProfileName and rejects these. Fill
    # whichever pair your account uses -- same either/or as returns below.
    "ShippingType": "Flat",
    "ShippingService-1:Option": "USPSFirstClass",
    "ShippingService-1:Cost": "%.2f" % SHIP_COST,
    "Country": "US",
    "Currency": "USD",
    "PayPalAccepted": "0",
    "C:Language": "English",
    "C:Original/Licensed Reprint": "Original",
    "C:Card Size": "Standard",
    # C:Type is deliberately NOT here. As a default it is forced onto every
    # row, and "Sports Trading Card" on a Pokemon card is a wrong answer to a
    # question eBay's CCG category does not even ask. values_for sets it per
    # card instead.
    "C:Country/Region of Manufacture": "United States",
}

ACTION_HEADER = ("Action(SiteID=US|Country=US|Currency=USD"
                 "|Version=1193|CC=UTF-8)")

FALLBACK_HEADER = [
    ACTION_HEADER, "CustomLabel", "Category", "Title", "ConditionID",
    "CD:Card Condition - (ID: 40001)",
    "CD:Professional Grader - (ID: 27501)",
    "CD:Grade - (ID: 27502)",
    "CDA:Certification Number - (ID: 27503)",
    "C:Sport", "C:Player/Athlete", "C:Season", "C:Manufacturer", "C:Set",
    "C:Card Number", "C:Parallel/Variety", "C:Features", "C:League",
    "C:Team", "C:Autographed", "C:Graded", "C:Card Size", "C:Language",
    "C:Original/Licensed Reprint", "C:Type", "C:Vintage",
    "PicURL", "Description", "Format", "Duration", "StartPrice",
    "Quantity", "Location", "DispatchTimeMax",
    # Accounts opted into Business Policies -- most of them now -- take
    # shipping, returns and payment from named profiles, and reject the
    # loose fields below. Accounts that are not take the loose fields and
    # ignore the profiles. Fill whichever pair your account uses; the other
    # stays blank.
    "ShippingProfileName", "ReturnProfileName", "PaymentProfileName",
    "ReturnsAcceptedOption", "ReturnsWithinOption", "RefundOption",
    "ShippingCostPaidByOption",
    "ShippingType", "ShippingService-1:Option", "ShippingService-1:Cost",
]


def norm(s):
    """Collapse a header to something comparable across template versions."""
    return re.sub(r"[^a-z0-9]", "", str(s or "").lower())


# Aliases let a template's own spelling find our value. First match wins, so
# the more specific patterns are listed before the looser ones.
ALIASES = {
    "action": "Action",
    "customlabel": "CustomLabel", "sku": "CustomLabel",
    "category": "Category", "categoryid": "Category",
    "primarycategory": "Category",
    "title": "Title",
    "subtitle": "Subtitle",
    "conditionid": "ConditionID", "condition": "ConditionID",
    "picurl": "PicURL", "pictureurl": "PicURL", "imageurl": "PicURL",
    "description": "Description",
    "format": "Format", "listingtype": "Format",
    "duration": "Duration", "listingduration": "Duration",
    "startprice": "StartPrice", "price": "StartPrice",
    "quantity": "Quantity",
    "location": "Location",
    "dispatchtimemax": "DispatchTimeMax", "handlingtime": "DispatchTimeMax",
}


def keys_for(header):
    """Candidate lookup keys for one template column, best guess first.

    Templates spell the same thing several ways -- "C:Player/Athlete" or bare
    "Player/Athlete", "CD:Card Condition - (ID: 40001)" or "Card Condition" --
    so return every form we might have stored it under and let the caller take
    the first that holds a value.
    """
    h = str(header).strip()
    n = norm(h)
    if n.startswith("action"):
        return ("Action",)

    cands = []
    if n in ALIASES:
        cands.append(ALIASES[n])

    # cda before cd before c: regex alternation takes the first branch that
    # matches, so shortest-first would swallow "CD:" as a bare "C".
    # The separator is required, otherwise plain words beginning with c
    # ("Category", "Currency") get their first letter eaten.
    m = re.match(r"^(?:cda|cd|c)[:\-]\s*(.+)$", h, flags=re.I)
    tail = m.group(1) if m else h
    tail = re.sub(r"\s*\(id:\s*\d+\s*\)\s*$", "", tail, flags=re.I)
    tail = tail.strip().rstrip("-").strip()

    cands.append("spec:" + norm(tail))
    cands.append(h)        # plain columns: ReturnsAcceptedOption, Location
    cands.append(tail)
    return tuple(cands)


def build_title(c, limit=80):
    """Assemble the listing title in the order a card seller writes it.

    Parts stay in reading order -- year, set, insert, parallel, player,
    number, badges -- which is what the workbook's own formula produces, so
    the sheet and the uploaded file agree. Only when the line runs past
    eBay's 80-character ceiling does anything get dropped, cheapest first.
    """
    def layout(brand):
        # A sports title carries its maker -- "Panini Prizm Draft Picks" --
        # and that is what people search. A TCG set name does not: "2026
        # Pitch Black Secret Rare Misty's Vitality #111/084" never says
        # Pokemon, so nobody looking for a Pokemon card finds it. Lead with
        # the game, and never drop it.
        game = c.get("sport", "") if c.get("cat") == "TCG" else ""
        if game and game.lower() in ("%s %s" % (brand, c["player"])).lower():
            game = ""

        # (text, how readily it can go; None = never). A parallel name, a
        # serial and an AUTO badge are what a buyer searches on, so they
        # outrank the year and the card number.
        p = [
            (game, None),
            (c["year"], 50),
            (brand, None),
            (c["insert"], 20),
            (c["parallel"], 80),
            (c["player"], None),
            (("#" + c["num"]) if c["num"] else "", 40),
            (("/" + c["serial"]) if c["serial"] else "", 70),
            ("RC" if c["rc"] else "", 65),
            ("AUTO" if c["auto"] else "", 60),
            ("PATCH" if c["relic"] else "", 30),
        ]
        # The team is deliberately absent: it wins few searches, it has its
        # own item specific, and leaving it out keeps this identical to the
        # workbook's own title formula.
        return [(t, pr) for t, pr in p if t]

    def joined(ps):
        return re.sub(r"\s+", " ", " ".join(t for t, _ in ps)).strip()

    # Squeeze the set name before giving up any actual content -- "Update
    # Series" and "Edition" cost characters and win no searches, whereas
    # "Superfractor" does.
    brands = [c["brand"]]
    squeezed = re.sub(r"\s+(Update Series|Edition|Series)\b", "", c["brand"])
    if squeezed != c["brand"]:
        brands.append(squeezed)
    bare = re.sub(r"^(Panini|Topps|Upper Deck)\s+", "", squeezed)
    if bare != squeezed:
        brands.append(bare)

    # Try each set-name variant and see how much content it costs to fit.
    # Fewest drops wins; on a tie the longest, most searchable set name does.
    best = None
    for rank, brand in enumerate(brands):
        parts = layout(brand)
        dropped = 0
        out = joined(parts)
        while len(out) > limit:
            droppable = [(pr, i) for i, (_t, pr) in enumerate(parts)
                         if pr is not None]
            if not droppable:
                break
            parts.pop(min(droppable)[1])
            dropped += 1
            out = joined(parts)
        if len(out) <= limit and (best is None or dropped < best[0]):
            best = (dropped, rank, out)
            if dropped == 0:
                break
    if best:
        return best[2]
    return joined(layout(brands[-1]))[:limit].strip()


def build_features(c):
    f = []
    if c["rc"]:
        f.append("Rookie")
    if c["insert"]:
        f.append("Insert")
    if c["parallel"]:
        f.append("Parallel/Variety")
    if c["serial"]:
        f.append("Serial Numbered")
    if c["auto"]:
        f.append("Autographed")
    if c["relic"]:
        f.append("Memorabilia")
    return ", ".join(f)


def build_description(c):
    bits = [
        ("Set", " ".join(x for x in [c["year"], c["brand"]] if x)),
        ("Insert", c["insert"]),
        ("Parallel", c["parallel"]),
        ("Card number", c["num"]),
        ("Serial", ("/" + c["serial"]) if c["serial"] else ""),
        ("Player", c["player"]),
        ("Team", c["team"]),
        ("Condition", c["grade_txt"] or c["condition"]),
    ]
    rows = "".join(
        "<tr><td style='padding:3px 12px 3px 0'><b>%s</b></td>"
        "<td style='padding:3px 0'>%s</td></tr>" % (k, v)
        for k, v in bits if v)
    return (
        "<div style='font-family:Arial,Helvetica,sans-serif;font-size:14px'>"
        "<p>%s</p><table>%s</table>"
        "<p>Shipped in a penny sleeve and top loader, inside a rigid mailer. "
        "Photos are of the actual card you receive.</p></div>"
    ) % (c["title"], rows)


def yes(v):
    return str(v or "").strip().lower() in ("yes", "y", "true", "1", "x")


def txt(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def read_inventory(path):
    wb = load_workbook(path, data_only=True)
    if "Inventory" not in wb.sheetnames:
        sys.exit("no Inventory tab in %s" % path)
    ws = wb["Inventory"]
    head = [txt(c.value) for c in ws[1]]
    idx = {norm(h): i for i, h in enumerate(head) if h}

    def get(row, name):
        i = idx.get(norm(name))
        return txt(row[i]) if i is not None and i < len(row) else ""

    cards = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        player = get(row, "Player or card name")
        if not player:
            continue
        grader = get(row, "Graded by")
        graded = bool(grader) and grader.lower() != "ungraded"
        c = {
            "sku": get(row, "SKU"), "status": get(row, "Status"),
            "cat": get(row, "Category") or "Sports",
            "sport": get(row, "Sport or game"), "year": get(row, "Year"),
            "brand": get(row, "Brand / set"), "insert": get(row, "Insert set"),
            "parallel": get(row, "Parallel"), "player": player,
            "num": get(row, "Card #"), "serial": get(row, "Serial /"),
            "team": get(row, "Team"), "league": get(row, "League"),
            "rc": yes(get(row, "RC")), "auto": yes(get(row, "Auto")),
            "relic": yes(get(row, "Relic")),
            "graded": graded, "grader": grader if graded else "",
            "grade": get(row, "Grade"), "cert": get(row, "Cert #"),
            "condition": get(row, "Card condition") or "Near Mint or Better",
            "qty": get(row, "Qty") or "1",
            "ask": get(row, "Ask price"), "market": get(row, "Market value"),
        }
        c["grade_txt"] = ("%s %s" % (c["grader"], c["grade"])).strip() if graded else ""
        c["title"] = build_title(c)
        cards.append(c)
    return cards


def eps_urls(path=None):
    """SKU -> the URLs eBay gave back, if ebay_pics.py has run.

    Read once per path and cached, because this is consulted for every card.
    The default is looked up at call time, not bound at definition time: as a
    default argument it would freeze the module constant and quietly ignore
    anyone who changed it.
    """
    if path is None:
        path = EPS_URLS
    cache = getattr(eps_urls, "_cache", None)
    if cache is None:
        cache = eps_urls._cache = {}
    if path not in cache:
        got = {}
        if os.path.exists(path):
            try:
                with open(path, encoding="utf-8") as fh:
                    got = json.load(fh)
            except (ValueError, OSError):
                got = {}
        cache[path] = got
    return cache[path]


def photo_urls(sku):
    """PicURL must be a public https URL.

    Prefer a picture eBay is already hosting: it needs no public site at all,
    it cannot 404 mid-upload, and it does not put 500KB scans into a git
    history that keeps them for ever. The Pages site is the fallback for
    cards that have not been through ebay_pics.py.
    """
    hosted = eps_urls().get(str(sku or ""))
    if hosted:
        return "|".join(hosted)

    found = []
    for suffix in ("", "-back", "-2", "-3"):
        for ext in (".jpg", ".jpeg", ".png"):
            rel = "photos/%s%s%s" % (sku, suffix, ext)
            if sku and os.path.exists(rel):
                found.append("%s/%s" % (PAGES, rel.replace("\\", "/")))
                break
    return "|".join(found)


def values_for(c):
    """Everything we know about one card, keyed the way field_for names it."""
    v = dict(DEFAULTS)
    v.update({
        "Action": "Add",
        "CustomLabel": c["sku"],
        "Category": str(CAT.get(c["cat"], CAT["Sports"])),
        "Title": c["title"],
        "ConditionID": str(COND_GRADED if c["graded"] else COND_UNGRADED),
        "StartPrice": c["ask"],
        "Quantity": c["qty"],
        "PicURL": photo_urls(c["sku"]),
        "Description": build_description(c),
    })
    if c["graded"]:
        v["spec:professionalgrader"] = c["grader"]
        v["spec:grade"] = c["grade"]
        v["spec:certificationnumber"] = c["cert"]
    else:
        v["spec:cardcondition"] = c["condition"]

    specs = {
        "sport": c["sport"], "playerathlete": c["player"],
        "player": c["player"], "athlete": c["player"],
        "season": c["year"], "year": c["year"],
        "manufacturer": (c["brand"].split()[0] if c["brand"] else ""),
        # eBay's Set specific carries the year; Season carries it too, and
        # both are expected to.
        "set": " ".join(x for x in [c["year"], c["brand"]] if x),
        "insertset": c["insert"], "insert": c["insert"],
        "cardnumber": c["num"], "parallelvariety": c["parallel"],
        "parallel": c["parallel"], "features": build_features(c),
        # League is not the sport. A college card is NCAA, not "Football",
        # and a wrong specific is worse than an absent one.
        "league": c["league"], "team": c["team"],
        "autographed": "Yes" if c["auto"] else "No",
        "graded": "Yes" if c["graded"] else "No",
        "cardname": c["player"],
        "printrun": c["serial"],
        "vintage": "No",
        "language": "English",
        "cardsize": "Standard",
        "originallicensedreprint": "Original",
    }

    # A Pokemon card is not a sports card and eBay's CCG Singles category
    # asks different questions: Game rather than Sport, Rarity rather than
    # Parallel/Variety, Character rather than Player. Store both vocabularies
    # and let the template take whichever it asks for. Type is the one that
    # must NOT carry over -- a wrong specific is worse than an absent one,
    # because eBay believes it.
    if str(c.get("cat", "")).upper() == "TCG":
        specs.update({
            "game": c["sport"], "cardgame": c["sport"],
            "rarity": c["parallel"], "character": c["player"],
            "type": "",
        })
    else:
        specs["type"] = "Sports Trading Card"
    for k, val in specs.items():
        v.setdefault("spec:" + k, val)
    # Also expose the C:-prefixed fallback names.
    for key in list(DEFAULTS):
        if key.startswith("C:"):
            v["spec:" + norm(key[2:])] = DEFAULTS[key]
    return v


def load_header(path):
    """Read the header row out of an eBay-downloaded template."""
    with open(path, newline="", encoding="utf-8-sig") as fh:
        for row in csv.reader(fh):
            if any(str(c).strip() for c in row):
                return [str(c).strip() for c in row]
    return []


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--workbook", default=WORKBOOK)
    ap.add_argument("--template", default=TEMPLATE)
    ap.add_argument("--sku", action="append", help="only these SKUs")
    ap.add_argument("--all", action="store_true",
                    help="include rows whose Status is not Unlisted")
    ap.add_argument("-o", "--out")
    args = ap.parse_args()

    inuse.refuse_if_open(args.workbook)

    if not os.path.exists(args.workbook):
        sys.exit("%s not found -- run make_workbook.py first" % args.workbook)

    cards = read_inventory(args.workbook)
    if args.sku:
        want = {s.strip().lower() for s in args.sku}
        cards = [c for c in cards if c["sku"].lower() in want]
    elif not args.all:
        cards = [c for c in cards
                 if (c["status"] or "Unlisted").lower() == "unlisted"]
    if not cards:
        sys.exit("no rows to write. Check the Status column, or pass --all.")

    if os.path.exists(args.template):
        header = load_header(args.template)
        source = "your template (%s)" % args.template
    else:
        header = list(FALLBACK_HEADER)
        source = "the documented fallback header -- NOT your account's template"

    fields = [keys_for(h) for h in header]
    out = args.out or "ebay-upload-%s.csv" % date.today().isoformat()

    unfilled, rows = set(), []
    for c in cards:
        v = values_for(c)
        row = []
        for h, cands in zip(header, fields):
            val = next((v[k] for k in cands if v.get(k, "") != ""), "")
            if val == "":
                unfilled.add(h)
            row.append(val)
        rows.append(row)

    # newline="" plus \r\n is what eBay expects; utf-8-sig keeps Excel happy.
    with open(out, "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.writer(fh, quoting=csv.QUOTE_MINIMAL, lineterminator="\r\n")
        w.writerow(header)
        w.writerows(rows)

    # Mirror it into the workbook so the sheet and the file never disagree.
    wb = load_workbook(args.workbook)
    # "eBay upload" was the old name; take either so an older workbook still works
    for old in ("eBay", "eBay upload"):
        if old in wb.sheetnames:
            del wb[old]
    ws = wb.create_sheet("eBay", 2)
    ws.append(header)
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    wb.save(args.workbook)

    print("wrote %s -- %d listing%s, %d columns"
          % (out, len(rows), "" if len(rows) == 1 else "s", len(header)))
    print("header came from %s" % source)
    for c in cards:
        flag = "" if len(c["title"]) <= 80 else "  <-- TITLE TOO LONG"
        print("  %-10s %2d chars  %s%s"
              % (c["sku"] or "-", len(c["title"]), c["title"], flag))
        if not photo_urls(c["sku"]):
            print("             no photo found at photos/%s.jpg" % c["sku"])
    if unfilled:
        print("\nleft blank (%d): %s" % (len(unfilled),
                                         ", ".join(sorted(unfilled)[:12])))
        print("fill anything eBay marks required before uploading.")

    # A picture eBay cannot fetch fails the listing, and the exporter has no
    # way to know what is actually published -- os.path.exists says the file
    # is on THIS disk, which is a different question. So count the two hosts
    # and say so, rather than handing over a file that looks complete.
    hosted = eps_urls()
    on_ebay = sum(1 for c in cards if str(c["sku"]) in hosted)
    on_pages = len(cards) - on_ebay
    if on_pages:
        print("\n%d of %d listing(s) still point at the Pages site."
              % (on_pages, len(cards)))
        print("Only what is committed to the repo is actually served there;"
              " the rest will 404 and eBay will reject those listings.")
        print("Run ebay_pics.py --go to hand the pictures to eBay instead.")
    if on_ebay:
        print("%d listing(s) use pictures eBay already hosts." % on_ebay)


if __name__ == "__main__":
    main()
