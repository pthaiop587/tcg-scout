"""Give a SKU to every card that was typed in by hand.

    python fill_skus.py            # say what it would do
    python fill_skus.py --go       # do it

file_batch.py and add_card.py assign a SKU when they add a card. Typing
straight into the Inventory tab does not, and a row without one is invisible
to everything downstream: make_ebay_csv.py cannot export it, add_photos.py has
nothing to file a picture against, and export_inventory.py skips it, so it
never reaches the dashboard either. Sixty cards typed in after a box rip is
sixty cards that quietly do not exist.

So this fills the gaps, in the order the rows already sit in, carrying on from
the highest SKU already used. It only ever writes into an EMPTY SKU cell -- a
card that already has one is never renumbered, because a SKU is what the
photographs on disk and any listing already out there are named after.

A row is a card if it has a name in "Player or card name". A row with a
quantity and nothing else is a half-typed line, not a card, and is left alone
rather than being given a number it will keep forever.
"""

import argparse
import os
import re
import sys

from openpyxl import load_workbook

WORKBOOK = "Card Run HQ - Master.xlsx"
NAME_COL = "Player or card name"


def scan(path):
    wb = load_workbook(path)
    if "Inventory" not in wb.sheetnames:
        sys.exit("%s has no Inventory tab" % path)
    ws = wb["Inventory"]
    hdr = [c.value for c in ws[1]]
    for needed in ("SKU", NAME_COL):
        if needed not in hdr:
            sys.exit("Inventory has no '%s' column" % needed)

    sku_col = hdr.index("SKU") + 1
    name_col = hdr.index(NAME_COL) + 1

    highest = 0
    blanks = []
    for r in range(2, ws.max_row + 1):
        sku = str(ws.cell(row=r, column=sku_col).value or "").strip()
        name = str(ws.cell(row=r, column=name_col).value or "").strip()
        m = re.match(r"CRH-(\d+)$", sku, re.I)
        if m:
            highest = max(highest, int(m.group(1)))
        elif name and not sku:
            blanks.append((r, name))
    return wb, ws, sku_col, highest, blanks


def main():
    p = argparse.ArgumentParser(
        description="Give a SKU to every hand-typed card that has none.")
    p.add_argument("--workbook", default=WORKBOOK)
    p.add_argument("--go", action="store_true",
                   help="do it; without this it only says what it would do")
    a = p.parse_args()

    if not os.path.exists(a.workbook):
        sys.exit("no workbook at %s" % a.workbook)

    wb, ws, sku_col, highest, blanks = scan(a.workbook)

    if not blanks:
        print("Every card already has a SKU. Highest is CRH-%04d." % highest)
        return 0

    print("%d card%s with no SKU. The next free number is CRH-%04d."
          % (len(blanks), "" if len(blanks) == 1 else "s", highest + 1))
    for r, name in blanks[:8]:
        print("   row %-4d %s" % (r, name))
    if len(blanks) > 8:
        print("   ... and %d more" % (len(blanks) - 8))

    if not a.go:
        print("\nThis was a dry run. Add --go to write them in.")
        print("Nothing that already has a SKU is touched -- a SKU is what your "
              "photos and any live listing are named after.")
        return 0

    n = highest
    for row, _name in blanks:
        n += 1
        ws.cell(row=row, column=sku_col, value="CRH-%04d" % n)
    wb.save(a.workbook)

    print("\nWrote CRH-%04d to CRH-%04d." % (highest + 1, n))
    print("Those cards can now be exported, photographed and shown on the "
          "dashboard. Nothing that already had a SKU was changed.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
