"""Append a card to the Inventory tab.

    python add_card.py --player "Shedeur Sanders" --year 2025 \
        --brand "Panini Prizm Draft Picks" --insert "Student Orientation" \
        --parallel "Gold Ice" --num 8 --team "Colorado Buffaloes" \
        --sport Football --rc --lot LOT-001

SKU is assigned automatically. Values land in the input columns only; the
title, fees and comps link are formulas already sitting in the row.
"""

import argparse
import re
import sys
from datetime import date

from openpyxl import load_workbook

import inuse

WORKBOOK = "Card Run HQ - Master.xlsx"


def norm(s):
    return re.sub(r"[^a-z0-9]", "", str(s or "").lower())


def next_sku(ws, col):
    n = 0
    for row in ws.iter_rows(min_row=2, min_col=col, max_col=col,
                            values_only=True):
        m = re.match(r"CRH-(\d+)$", str(row[0] or "").strip())
        if m:
            n = max(n, int(m.group(1)))
    return "CRH-%04d" % (n + 1)


def first_empty(ws, col):
    r = 2
    while ws.cell(row=r, column=col).value not in (None, ""):
        r += 1
    return r


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--workbook", default=WORKBOOK)
    p.add_argument("--player", required=True)
    p.add_argument("--year"); p.add_argument("--brand")
    p.add_argument("--insert", default=""); p.add_argument("--parallel", default="")
    p.add_argument("--num", default=""); p.add_argument("--serial", default="")
    p.add_argument("--team", default=""); p.add_argument("--sport", default="Football")
    p.add_argument("--league", default="",
                   help="NCAA, NFL, NBA... leave blank rather than guess")
    p.add_argument("--category", default="Sports",
                   choices=["Sports", "TCG", "Non-sport"])
    p.add_argument("--condition", default="Near Mint or Better")
    p.add_argument("--grader", default=""); p.add_argument("--grade", default="")
    p.add_argument("--cert", default="")
    p.add_argument("--rc", action="store_true"); p.add_argument("--auto", action="store_true")
    p.add_argument("--relic", action="store_true")
    p.add_argument("--qty", type=int, default=1)
    p.add_argument("--cost", type=float); p.add_argument("--market", type=float)
    p.add_argument("--ask", type=float)
    p.add_argument("--source", default=""); p.add_argument("--lot", default="")
    p.add_argument("--notes", default="")
    a = p.parse_args()

    inuse.refuse_if_open(a.workbook)

    wb = load_workbook(a.workbook)
    ws = wb["Inventory"]
    idx = {norm(c.value): c.column for c in ws[1] if c.value}

    def put(header, value):
        if value in (None, "", False):
            return
        col = idx.get(norm(header))
        if col is None:
            sys.exit("no '%s' column in Inventory" % header)
        ws.cell(row=r, column=col, value=value)

    r = first_empty(ws, idx[norm("Player or card name")])
    sku = next_sku(ws, idx[norm("SKU")])

    put("SKU", sku)
    put("Status", "Unlisted")
    put("Date in", date.today())
    put("Source", a.source)
    put("Lot ID", a.lot)
    put("Category", a.category)
    put("Sport or game", a.sport)
    put("League", a.league)
    put("Year", a.year)
    put("Brand / set", a.brand)
    put("Insert set", a.insert)
    put("Parallel", a.parallel)
    put("Player or card name", a.player)
    put("Card #", a.num)
    put("Serial /", a.serial)
    put("Team", a.team)
    put("RC", "Yes" if a.rc else "No")
    put("Auto", "Yes" if a.auto else "No")
    put("Relic", "Yes" if a.relic else "No")
    put("Graded by", a.grader)
    put("Grade", a.grade)
    put("Cert #", a.cert)
    put("Card condition", a.condition)
    put("Qty", a.qty)
    put("Cost each", a.cost)
    put("Market value", a.market)
    put("Ask price", a.ask)
    put("Notes", a.notes)

    ws.cell(row=r, column=idx[norm("Date in")]).number_format = "yyyy-mm-dd"
    wb.save(a.workbook)
    print("%s added on row %d: %s" % (sku, r, a.player))


if __name__ == "__main__":
    main()
