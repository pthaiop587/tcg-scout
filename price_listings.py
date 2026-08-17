"""Set Market value and Ask price from what the cards actually sold for.

    python price_listings.py              # say what it would do
    python price_listings.py --go         # write it in

THE RULE, as Mr. P set it:

    ask       the last price the card actually SOLD for on eBay, rounded up
              to the next whole dollar
    postage   $1.50, charged to the buyer, per single card
    floor     under 50c it is not a listing, it goes in the bulk pile

WHY THE LAST SALE AND NOT THE GUIDE PRICE.

The guide is a smoothed average of past sales; the last sale is what somebody
paid. They disagree often and by a lot -- Dante Moore's guide said $1.19 while
the most recent sale was $3.57 -- and when they disagree the guide is the one
describing a market that has moved on. prices.py records both, so this uses
the sale and falls back to the guide only when a card has no recorded sale.

WHY THE POSTAGE CHARGE CHANGES EVERYTHING.

eBay takes about 13.25% plus $0.30, and that percentage applies to the postage
the buyer pays as well as the item. Charging $1.50 against a real cost of
about $0.70 leaves roughly 80c, which is what covers the fixed fee. That is
the difference between a dollar card being worth listing and not:

    seller-paid postage   $1.00 card returns  -$0.13   (you pay to sell it)
    buyer-paid $1.50      $1.00 card returns  +$1.17

So the same collection that was 5 listings and 155 bulk under the old
assumption is mostly listable under this one. The arithmetic did not change;
who pays the postman did.

Market value is rewritten, not just filled -- it was set once from the guide
price and went stale when 34 cards were re-priced after the Silver/base
correction. A stale number is worse than a blank one; a blank does not look
like an answer.

Nothing is written without --go.
"""

import argparse
import math
import sys

from openpyxl import load_workbook

import inuse

WORKBOOK = "Card Run HQ - Master.xlsx"

FEE_PCT = 13.25          # eBay's cut, charged on the postage too
FEE_FIXED = 0.30         # per order
SHIP_CHARGE = 1.50       # what the buyer pays for one card
POSTAGE = 0.70           # what it really costs to send one card, tracked
MIN_LIST = 0.50          # under this it is bulk, not a listing
BULK = "Bulk"


def net_of(ask, fee_pct, fee_fixed, ship_charge, postage):
    """What lands in the bank once eBay and the postman are paid.

    The fee percentage applies to the postage as well, which is why charging
    more for postage does not simply add to the margin.
    """
    return (ask + ship_charge) * (1 - fee_pct / 100.0) - fee_fixed - postage


def ask_for(price):
    """Up to the next whole dollar. $1.19 asks $2; $3.00 stays $3."""
    return float(math.ceil(round(price, 6) - 1e-9))


def num(x):
    return x if isinstance(x, (int, float)) else None


def plan(cards, min_list=MIN_LIST):
    """Decide each card's market value, ask price, and whether it is bulk.

    Kept apart from the workbook so the arithmetic can be tested without one.
    """
    out = []
    for c in cards:
        row = dict(c, was=num(c.get("market")), market=None, ask=None,
                   bulk=False, src="", why="")
        sale, guide = num(c.get("sale")), num(c.get("raw"))
        price, src = (sale, "sold") if sale is not None else (guide, "guide")
        if price is None:
            row["bulk"] = True
            row["why"] = "no price found for it"
            out.append(row)
            continue
        row["market"], row["src"] = price, src
        if price < min_list:
            row["bulk"] = True
            row["why"] = "last %s %.2f, under the %.2f floor" % (src, price,
                                                                 min_list)
            out.append(row)
            continue
        row["ask"] = ask_for(price)
        out.append(row)
    return out


def read(ws, g):
    cards = []
    for r in range(2, ws.max_row + 1):
        sku = ws.cell(row=r, column=g["SKU"]).value
        if not sku:
            continue
        cards.append({
            "row": r, "sku": sku,
            "name": ws.cell(row=r, column=g["Player or card name"]).value,
            "status": str(ws.cell(row=r, column=g["Status"]).value
                          or "").strip(),
            "cost": ws.cell(row=r, column=g["Cost each"]).value,
            "raw": ws.cell(row=r, column=g["Raw price"]).value,
            "sale": ws.cell(row=r, column=g["Raw last sale"]).value,
            "market": ws.cell(row=r, column=g["Market value"]).value,
        })
    return cards


def main():
    p = argparse.ArgumentParser(
        description="Price the listings from what each card last sold for.")
    p.add_argument("--workbook", default=WORKBOOK)
    p.add_argument("--fee-pct", type=float, default=FEE_PCT)
    p.add_argument("--fee-fixed", type=float, default=FEE_FIXED)
    p.add_argument("--ship-charge", type=float, default=SHIP_CHARGE,
                   help="postage charged to the buyer (default %.2f)"
                        % SHIP_CHARGE)
    p.add_argument("--postage", type=float, default=POSTAGE,
                   help="what sending one card really costs (default %.2f)"
                        % POSTAGE)
    p.add_argument("--min-list", type=float, default=MIN_LIST,
                   help="below this a card is bulk, not a listing "
                        "(default %.2f)" % MIN_LIST)
    p.add_argument("--no-bulk", action="store_true",
                   help="leave Status alone on cards under the floor")
    p.add_argument("--go", action="store_true", help="write the values in")
    a = p.parse_args()

    inuse.refuse_if_open(a.workbook)
    wb = load_workbook(a.workbook)
    ws = wb["Inventory"]
    hdr = [c.value for c in ws[1]]
    g = {n: i + 1 for i, n in enumerate(hdr) if n}
    for need in ("SKU", "Status", "Cost each", "Raw price", "Raw last sale",
                 "Market value", "Ask price"):
        if need not in g:
            sys.exit("no %s column in %s" % (need, a.workbook))

    cards = read(ws, g)
    # A Review row is waiting on a person. Pricing it would bury the question.
    held = [c for c in cards if c["status"].lower() == "review"]
    live = [c for c in cards if c["status"].lower() != "review"]

    rows = plan(live, a.min_list)
    priced = [r for r in rows if r["ask"] is not None]
    bulk = [r for r in rows if r["bulk"]]
    guessed = [r for r in priced if r["src"] == "guide"]
    moved = [r for r in rows if r["market"] is not None
             and r["was"] is not None and abs(r["was"] - r["market"]) > 0.005]

    print("%d card(s): %d to list, %d for the bulk pile, %d held on Review"
          % (len(cards), len(priced), len(bulk), len(held)))
    print("ask = last sold price rounded up to the dollar; under $%.2f is bulk"
          % a.min_list)
    print("buyer pays $%.2f postage; eBay takes %.2f%% of that too, and $%.2f"
          % (a.ship_charge, a.fee_pct, a.fee_fixed))
    if guessed:
        print("%d card(s) have no recorded sale -- priced off the guide "
              "instead" % len(guessed))
    print()

    if moved:
        print("Market value changed on %d row(s):" % len(moved))
        for r in sorted(moved, key=lambda x: -abs(x["was"] - x["market"]))[:6]:
            print("   %-9s %-22s %6.2f -> %6.2f"
                  % (r["sku"], str(r["name"])[:22], r["was"], r["market"]))
        if len(moved) > 6:
            print("   ... and %d more" % (len(moved) - 6))
        print()

    take = sum(net_of(r["ask"], a.fee_pct, a.fee_fixed, a.ship_charge,
                      a.postage) for r in priced)
    cost = sum(num(r["cost"]) or 0 for r in priced)
    print("top of the list:")
    for r in sorted(priced, key=lambda x: -x["ask"])[:12]:
        print("   %-9s %-24s %-5s %6.2f  ask %6.2f  net %6.2f"
              % (r["sku"], str(r["name"])[:24], r["src"], r["market"],
                 r["ask"], net_of(r["ask"], a.fee_pct, a.fee_fixed,
                                  a.ship_charge, a.postage)))
    if len(priced) > 12:
        print("   ... and %d more" % (len(priced) - 12))

    print("\nif every listing sold: $%.2f net against $%.2f of card cost"
          % (take, cost))
    print("bulk pile: %d card(s)" % len(bulk))

    if not a.go:
        print("\nNothing written. Add --go.")
        return

    for r in rows:
        if r["market"] is not None:
            ws.cell(row=r["row"], column=g["Market value"]).value = r["market"]
        if r["ask"] is not None:
            ws.cell(row=r["row"], column=g["Ask price"]).value = r["ask"]
            if ws.cell(row=r["row"], column=g["Status"]).value == BULK:
                ws.cell(row=r["row"], column=g["Status"]).value = "Unlisted"
        elif r["bulk"] and not a.no_bulk:
            ws.cell(row=r["row"], column=g["Ask price"]).value = None
            ws.cell(row=r["row"], column=g["Status"]).value = BULK
    wb.save(a.workbook)
    print("\nSaved %s." % a.workbook)


if __name__ == "__main__":
    main()
