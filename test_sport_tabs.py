"""Tests for the per-sport view tabs.

Run: python -m pytest test_sport_tabs.py

The point of these tabs is that they are VIEWS. If they ever became a second
place cards live, the damage is silent: file_batch.py takes the next SKU from
Inventory only, make_ebay_csv.py exports from Inventory only, and the Summary
and Audit tabs count Inventory only. A basketball card typed into a Basketball
tab would get a SKU another card already has, never reach an upload, and be
missing from every total, with nothing anywhere saying so.

So the tests that matter are: Inventory still holds every card, the export
still sees every card, and a tab somebody made by hand is never overwritten.
"""

import os
import subprocess
import sys

import pytest
from openpyxl import load_workbook

import file_batch as fb
import sport_tabs as st

CARDS = [
    {"player": "Shedeur Sanders", "year": 2025, "brand": "Prizm Draft Picks",
     "num": "8", "sport": "Football", "league": "NCAA", "cost": 1.26, "market": 12.5},
    {"player": "Victor Wembanyama", "year": 2024, "brand": "Prizm",
     "num": "136", "sport": "Basketball", "league": "NBA", "cost": 4.0, "market": 38.0},
    {"player": "Cooper Flagg", "year": 2025, "brand": "Prizm Draft Picks",
     "num": "1", "sport": "Basketball", "league": "NCAA", "cost": 3.5, "market": 22.0},
]


@pytest.fixture
def wb(tmp_path):
    out = tmp_path / "Card Run HQ - Master.xlsx"
    subprocess.run([sys.executable, "make_workbook.py", "--out", str(out)],
                   check=True, capture_output=True)
    fb.add_rows(str(out), [dict(c) for c in CARDS])
    return out


def run(wb, *args):
    r = subprocess.run([sys.executable, os.path.abspath("sport_tabs.py"),
                        "--workbook", str(wb)] + list(args),
                       capture_output=True, text=True)
    assert r.returncode == 0, r.stdout + r.stderr
    return r.stdout


def rows_on(wb, tab):
    ws = load_workbook(wb)[tab]
    hdr = [c.value for c in ws[2]]
    out = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0]:
            out.append(dict(zip(hdr, row)))
    return out


# --- the split --------------------------------------------------------------

def test_a_tab_per_sport_in_use(wb):
    run(wb)
    names = load_workbook(wb).sheetnames
    assert "Basketball" in names
    assert "Football" in names


def test_each_tab_holds_only_its_own_sport(wb):
    run(wb)
    bball = rows_on(wb, "Basketball")
    assert {r["Player or card name"] for r in bball} == \
        {"Victor Wembanyama", "Cooper Flagg"}
    assert [r["Player or card name"] for r in rows_on(wb, "Football")] == \
        ["Shedeur Sanders"]


def test_add_makes_a_tab_before_there_are_any_cards(wb):
    """You want the tab when you start on a sport, not after."""
    out = run(wb, "--add", "Hockey")
    assert "Hockey" in load_workbook(wb).sheetnames
    assert rows_on(wb, "Hockey") == []
    assert "Hockey" in out


# --- the thing that must not happen -----------------------------------------

def test_inventory_still_holds_every_card(wb):
    """The views are copies. Inventory stays the one record."""
    run(wb)
    ws = load_workbook(wb)["Inventory"]
    skus = [r[0] for r in ws.iter_rows(min_row=2, values_only=True) if r[0]]
    assert skus == ["CRH-0001", "CRH-0002", "CRH-0003"]


def test_the_ebay_export_still_sees_every_card(wb, tmp_path):
    """If a sport tab ever became a real inventory, these would go missing."""
    run(wb)
    r = subprocess.run([sys.executable, "make_ebay_csv.py",
                        "--workbook", str(wb), "--out", str(tmp_path / "o.csv")],
                       capture_output=True, text=True)
    body = r.stdout + r.stderr
    assert "3 listings" in body, body


def test_the_next_sku_still_comes_from_inventory(wb):
    """A card added after the tabs exist must not reuse a number."""
    run(wb)
    added = fb.add_rows(str(wb), [{"player": "Travis Hunter", "sport": "Football"}])
    assert [e["sku"] for e in added] == ["CRH-0004"]


# --- rebuilding -------------------------------------------------------------

def test_running_it_twice_changes_nothing(wb):
    run(wb)
    before = load_workbook(wb).sheetnames
    run(wb)
    assert load_workbook(wb).sheetnames == before
    assert len(rows_on(wb, "Basketball")) == 2


def test_a_new_card_shows_up_on_the_next_run(wb):
    run(wb)
    fb.add_rows(str(wb), [{"player": "Ace Bailey", "sport": "Basketball"}])
    run(wb)
    assert len(rows_on(wb, "Basketball")) == 3


def test_a_tab_you_made_yourself_is_never_overwritten(wb):
    """Only sheets carrying the generated marker are replaced."""
    book = load_workbook(wb)
    # a fresh workbook already ships a generated Basketball tab, so stand in
    # for someone who deleted it and made their own
    if "Basketball" in book.sheetnames:
        del book["Basketball"]
    mine = book.create_sheet("Basketball")
    mine["A1"] = "my own notes"
    mine["A2"] = "do not delete this"
    book.save(wb)

    out = run(wb)
    ws = load_workbook(wb)["Basketball"]
    assert ws["A1"].value == "my own notes"
    assert ws["A2"].value == "do not delete this"
    assert "left alone" in out


def test_the_tab_says_it_is_generated(wb):
    """Otherwise somebody types in it and loses the lot on the next run."""
    run(wb)
    ws = load_workbook(wb)["Basketball"]
    a1 = str(ws["A1"].value)
    assert a1.startswith("VIEW")
    assert "Do not type here" in a1


def test_the_tab_shows_when_each_card_was_logged(wb):
    """The point of the tab is looking at one game on its own; "when did this
    come in" is part of that and used to be missing."""
    run(wb)
    rows = rows_on(wb, "Basketball")
    assert rows, "no rows to check"
    for r in rows:
        assert "Date in" in r, r.keys()
        assert r["Date in"] is not None


def test_the_date_is_formatted_as_a_date(wb):
    """openpyxl hands back a datetime; without a number format Excel shows it
    as a five-digit serial, which reads as a card number at a glance."""
    run(wb)
    ws = load_workbook(wb)["Basketball"]
    hdr = [c.value for c in ws[2]]
    col = hdr.index("Date in") + 1
    cell = ws.cell(row=3, column=col)
    assert cell.number_format == st.DATEFMT, cell.number_format


def test_every_date_column_carries_a_date_format(wb):
    """Listed on and Sold on travel to the tab too, and have the same trap.

    Only the ones the workbook actually has: the price-lookup columns are added
    by prices.py and a workbook that has never been priced will not carry
    them."""
    run(wb)
    ws = load_workbook(wb)["Basketball"]
    hdr = [c.value for c in ws[2]]
    checked = 0
    for name in st.DATES:
        if name not in hdr:
            continue
        cell = ws.cell(row=3, column=hdr.index(name) + 1)
        assert cell.number_format == st.DATEFMT, (name, cell.number_format)
        checked += 1
    assert checked >= 3, "expected at least Date in, Listed on, Sold on"


def test_a_priced_workbook_shows_its_prices_and_their_dates(wb):
    """prices.py appends six columns; the tab has to carry them, formatted, or
    a price sits next to a five-digit serial where its date should be."""
    import datetime
    book = load_workbook(wb)
    inv = book["Inventory"]
    start = inv.max_column
    for i, name in enumerate(["Raw price", "Raw last sold", "PSA 9 price",
                              "PSA 9 last sold", "PSA 10 price",
                              "PSA 10 last sold"], start=1):
        inv.cell(row=1, column=start + i, value=name)
        for r in (2, 3, 4):
            inv.cell(row=r, column=start + i,
                     value=datetime.date(2026, 8, 1) if "sold" in name
                     else 12.34)
    book.save(wb)

    run(wb)
    ws = load_workbook(wb)["Basketball"]
    hdr = [c.value for c in ws[2]]
    for name in ("Raw price", "PSA 9 price", "PSA 10 price"):
        assert name in hdr, name
        assert ws.cell(row=3, column=hdr.index(name) + 1).number_format == st.MONEY
    for name in ("Raw last sold", "PSA 9 last sold", "PSA 10 last sold"):
        assert name in hdr, name
        assert ws.cell(row=3, column=hdr.index(name) + 1).number_format == st.DATEFMT


def test_the_date_survives_a_rebuild(wb):
    """Tabs are thrown away and rebuilt every refresh; the date has to come
    back each time, not just the first."""
    run(wb)
    run(wb)
    rows = rows_on(wb, "Basketball")
    assert all(r["Date in"] is not None for r in rows)


def test_something_typed_onto_the_tab_is_never_thrown_away(wb):
    """The tab says "do not type here". Somebody typed there anyway, a refresh
    destroyed it, and nothing said so. A1 is a sign, not a lock -- the rebuild
    has to check before it deletes."""
    run(wb)
    book = load_workbook(wb)
    ws = book["Basketball"]
    hdr = [c.value for c in ws[2]]
    ws.cell(row=3, column=hdr.index("Notes") + 1, value="listed on ebay 8/16")
    book.save(wb)

    out = run(wb)
    after = load_workbook(wb)
    assert "Basketball (typed on)" in after.sheetnames, after.sheetnames
    kept = after["Basketball (typed on)"]
    khdr = [c.value for c in kept[2]]
    assert kept.cell(row=3, column=khdr.index("Notes") + 1).value \
        == "listed on ebay 8/16"
    assert "typed by hand" in out
    # and the view still gets rebuilt alongside it
    assert "Basketball" in after.sheetnames
    assert len(rows_on(wb, "Basketball")) == 2


def test_a_row_that_is_not_in_inventory_is_kept(wb):
    """A whole card typed straight onto the tab -- the exact thing that was
    lost."""
    run(wb)
    book = load_workbook(wb)
    ws = book["Basketball"]
    ws.cell(row=9, column=1, value="CRH-9999")
    ws.cell(row=9, column=7, value="Typed straight onto the tab")
    book.save(wb)

    run(wb)
    after = load_workbook(wb)
    assert "Basketball (typed on)" in after.sheetnames
    assert after["Basketball (typed on)"].cell(row=9, column=1).value \
        == "CRH-9999"


def test_an_untouched_tab_is_still_replaced_cleanly(wb):
    """The guard must not make every refresh leave junk sheets behind."""
    run(wb)
    run(wb)
    run(wb)
    names = load_workbook(wb).sheetnames
    assert not [n for n in names if "typed on" in n], names


def test_list_reports_what_is_there(wb):
    run(wb)
    out = run(wb, "--list")
    assert "Basketball" in out and "Football" in out
    assert "2 card(s)" in out


def test_a_fresh_workbook_already_has_a_tab_per_game(tmp_path):
    """make_workbook builds them empty, so a game can be started before its
    first card. sport_tabs then just keeps them up to date."""
    out = tmp_path / "wb.xlsx"
    subprocess.run([sys.executable, "make_workbook.py", "--out", str(out)],
                   check=True, capture_output=True)
    r = subprocess.run([sys.executable, os.path.abspath("sport_tabs.py"),
                        "--workbook", str(out)], capture_output=True, text=True)
    assert r.returncode == 0, r.stdout + r.stderr
    for game in ("Football", "Basketball", "Baseball", "Pokemon",
                 "Palworld", "One Piece", "Disney"):
        assert game in r.stdout
    assert "0 card(s)" in r.stdout
