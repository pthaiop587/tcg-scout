"""Move an existing workbook onto the current tab layout, keeping what is in it.

    python upgrade_workbook.py                 # say what it would do
    python upgrade_workbook.py --go            # actually do it

make_workbook.py --force builds a fresh workbook and throws away everything
typed into the old one, which is fine on day one and unusable afterwards. This
does the same rebuild and then carries the data across.

Rows are matched BY HEADER NAME, never by position. That is the whole point:
the layout gained five tabs and the columns moved, so copying cell A2 to cell
A2 would put a purchase date into a SKU column and nobody would notice until
an upload failed. A column that no longer exists is reported rather than
dropped in silence.

The old file is copied to a dated backup first, always, without being asked.
"""

import argparse
import os
import shutil
import subprocess
import sys
from datetime import datetime

from openpyxl import load_workbook

WORKBOOK = "Card Run HQ - Master.xlsx"

# Only tabs a person types into. Summary, Audit, Read me, Reference, Lists and
# the eBay upload are generated or fixed, so they come from the new build.
CARRY = ["Inventory", "Box log", "Sales", "Purchases", "Expenses", "Photos"]

# Purchases and Expenses became one Costs tab, and a Box log row is a box that
# was bought. Without this the short layout has "no home" for all three and the
# spending is dropped into the backup -- which is exactly the money you need to
# know a box paid for itself. Old column -> new column, by name.
MERGE_INTO = {
    "Purchases": ("Costs", {
        "Date": "Date", "What": "What", "Type": "Type",
        "Vendor / store": "Vendor / store", "Lot ID": "Lot ID", "Qty": "Qty",
        "Unit price": "Unit price", "Tax": "Tax", "Shipping": "Shipping",
        "Paid with": "Paid with", "Order / receipt #": "Order / receipt #",
        "Receipt file": "Receipt file", "Notes": "Notes"}),
    "Expenses": ("Costs", {
        "Date": "Date", "What": "What", "Category": "Type",
        "Vendor": "Vendor / store", "Amount": "Unit price",
        "Paid with": "Paid with", "Receipt file": "Receipt file",
        "Notes": "Notes"}),
    "Box log": ("Costs", {
        "Date": "Date", "Product": "What", "Store": "Vendor / store",
        "Lot ID": "Lot ID", "Qty": "Qty", "Unit price": "Unit price",
        "Tax": "Tax", "Notes": "Notes"}),
}

# An Expenses row had no Qty and a Box log row had no Type, and Costs needs
# both for its subtotal and its stock/running split.
MERGE_DEFAULTS = {"Expenses": {"Qty": 1}, "Box log": {"Type": "Sealed box"}}


def headers(ws):
    return [c.value for c in ws[1]]


def data_rows(ws):
    """The typed rows, stopping where the data stops.

    Formulas are left behind -- the new sheet brings its own, and they will be
    the current ones.

    It stops at the first completely empty row, and that is the whole point.
    Every tab carries a block of explanatory notes below its data, separated by
    a blank row. Reading "every row with something in it" swept those notes up
    as records, and they were duly carried across: the Costs tab came out with
    "The costs that are not a card." sitting in its Date column as though it
    were a purchase. Anything typed after the gap is reported rather than
    silently dropped.
    """
    hdr = headers(ws)
    out = []
    stopped_at = None
    for n, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        rec = {}
        for name, value in zip(hdr, row):
            if not name or value in (None, ""):
                continue
            if isinstance(value, str) and value.startswith("="):
                continue
            rec[name] = value
        if not rec:
            stopped_at = n
            break
        out.append(rec)

    if stopped_at:
        for row in ws.iter_rows(min_row=stopped_at + 1, values_only=True):
            for name, value in zip(hdr, row):
                if name and value not in (None, "") and not (
                        isinstance(value, str) and value.startswith("=")):
                    print("   note: %s has something below the blank row at %d "
                          "-- left where it is" % (ws.title, stopped_at))
                    return out
    return out


def first_free(ws, hdr):
    """The first row with nothing typed in it, ignoring formula columns."""
    r = 2
    while r <= ws.max_row:
        row = [ws.cell(row=r, column=i + 1).value for i in range(len(hdr))]
        typed = [v for v in row
                 if v not in (None, "") and not (isinstance(v, str) and v.startswith("="))]
        if not typed:
            return r
        r += 1
    return r


def main():
    p = argparse.ArgumentParser(
        description="Rebuild the workbook on the current layout, keeping the data.")
    p.add_argument("--workbook", default=WORKBOOK)
    p.add_argument("--full", action="store_true",
                   help="rebuild on the FULL layout (Purchases, Box log, "
                        "Expenses, Sales, Photos, Summary, Audit, Reference) "
                        "instead of the short one")
    p.add_argument("--go", action="store_true",
                   help="do it; without this it only says what it would do")
    a = p.parse_args()

    if not os.path.exists(a.workbook):
        sys.exit("no workbook at %s -- there is nothing to upgrade. "
                 "make_workbook.py builds a new one." % a.workbook)

    old = load_workbook(a.workbook)
    print("current file: %s" % a.workbook)
    print("  tabs: %s" % ", ".join(old.sheetnames))

    carried = {}
    for name in CARRY:
        if name in old.sheetnames:
            rows = data_rows(old[name])
            if rows:
                carried[name] = rows
    if carried:
        for name, rows in carried.items():
            print("  %-12s %d row(s) to carry over" % (name, len(rows)))
    else:
        print("  nothing typed in yet -- an upgrade would just be a rebuild.")

    if not a.go:
        print("\nThis was a dry run. Add --go to do it.")
        print("The old file will be copied to a dated backup first.")
        return 0

    stamp = datetime.now().strftime("%Y-%m-%d-%H%M%S")
    backup = "%s (backup %s).xlsx" % (os.path.splitext(a.workbook)[0], stamp)
    shutil.copy(a.workbook, backup)
    print("\nbacked up to: %s" % backup)

    fresh = "%s.new.xlsx" % os.path.splitext(a.workbook)[0]
    build = [sys.executable, "make_workbook.py", "--out", fresh, "--force"]
    if a.full:
        build.append("--full")
    subprocess.run(build, check=True, capture_output=True)
    new = load_workbook(fresh)

    lost = {}
    for name, rows in carried.items():
        target, mapping = name, None
        if name not in new.sheetnames and name in MERGE_INTO:
            target, mapping = MERGE_INTO[name]
        if target not in new.sheetnames:
            # nowhere for it at all -- say so loudly, because the rows are only
            # in the backup from here on
            lost[name] = ("%d row(s) -- this layout has no %s tab. "
                          "Re-run with --full to keep it." % (len(rows), name))
            continue
        if mapping:
            defaults = MERGE_DEFAULTS.get(name, {})
            moved = []
            for rec in rows:
                out = dict(defaults)
                out.update({mapping[k]: v for k, v in rec.items() if k in mapping})
                if out:
                    moved.append(out)
            rows = moved
            print("   %s -> %s (%d row(s))" % (name, target, len(rows)))
        ws = new[target]
        hdr = headers(ws)
        col = {h: i + 1 for i, h in enumerate(hdr) if h}

        # A fresh workbook is not empty: make_workbook.py seeds the Box log
        # with the one box actually bought, straight off its receipt. Carrying
        # the old copy in on top of it puts the same box in twice, and the
        # cost-per-card figures then count it twice. So anything the new sheet
        # already has, identically, is left alone.
        already = [frozenset(rec.items()) for rec in data_rows(ws)]
        r = first_free(ws, hdr)
        skipped = 0

        for rec in rows:
            if frozenset(rec.items()) in already:
                skipped += 1
                continue
            for key, value in rec.items():
                if key not in col:
                    lost.setdefault(name, set())
                    if isinstance(lost[name], set):
                        lost[name].add(key)
                    continue
                ws.cell(row=r, column=col[key], value=value)
            r += 1
        print("carried %d row(s) into %s%s"
              % (len(rows) - skipped, target,
                 "" if not skipped else
                 " (%d already there, left alone)" % skipped))

    new.save(fresh)
    os.replace(fresh, a.workbook)
    print("\n%s is now on the current layout (%d tabs)."
          % (a.workbook, len(new.sheetnames)))

    if lost:
        print("\nCOLUMNS THAT NO LONGER EXIST -- the values are in the backup, "
              "not in the new file:")
        for tab, what in lost.items():
            print("   %s: %s" % (tab, what if isinstance(what, str)
                                 else ", ".join(sorted(what))))
    print("\nNext: python embed_photos.py   (fills the Photos tab from photos/)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
