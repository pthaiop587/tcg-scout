"""Put the card photographs into the workbook's Photos tab.

    python embed_photos.py
    python embed_photos.py --workbook "Card Run HQ - Master.xlsx"

Reads the photos folder, writes one row per SKU, and drops an actual
thumbnail into the Picture column, so the Photos tab is a visual stock check
rather than a list of filenames. Re-run it whenever photos are added; it
rewrites the tab from what is on disk, so it can never drift from the files.

WHY A THUMBNAIL AND A URL, AND NOT THE PHOTO ITSELF. The workbook carries a
small copy for looking at and a link for eBay. eBay will not take an embedded
picture -- PicURL has to be a public https address it can fetch -- so the real
files live in photos/ and are published with the site. That also keeps the
workbook small: a hundred cards at full resolution would be a 200 MB
spreadsheet nobody can open.

So the thumbnails are inventory control, and the URLs are the listing. If a
URL does not resolve, the photos have not been pushed.
"""

import argparse
import io
import os
import re
import sys

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

import inuse
from PIL import Image

WORKBOOK = "Card Run HQ - Master.xlsx"
PHOTOS = "photos"
PAGES = "https://pthaiop587.github.io/tcg-scout"

# Wide enough to tell two gold parallels apart, small enough that a hundred of
# them still opens. Excel sizes a picture in points; a column width of 22 is
# about 150 px, so match it and set the row height to suit.
THUMB_W = 132
SKU_RE = re.compile(r"^(CRH-\d{4})(-back|-2|-3)?\.(jpg|jpeg|png)$", re.I)
SLOT_ORDER = ["", "-back", "-2", "-3"]


def find_photos(folder):
    """SKU -> {slot: filename}, from whatever is actually on disk."""
    out = {}
    if not os.path.isdir(folder):
        return out
    for name in sorted(os.listdir(folder)):
        m = SKU_RE.match(name)
        if not m:
            continue
        sku, slot = m.group(1).upper(), (m.group(2) or "").lower()
        out.setdefault(sku, {})[slot] = name
    return out


def card_names(ws):
    """SKU -> the eBay title, so a row says what it is a picture of."""
    hdr = [c.value for c in ws[1]]
    try:
        i_sku = hdr.index("SKU")
        i_name = hdr.index("Player or card name")
        i_year = hdr.index("Year")
        i_brand = hdr.index("Brand / set")
    except ValueError:
        return {}
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        sku = row[i_sku]
        if not sku:
            continue
        bits = [row[i_year], row[i_brand], row[i_name]]
        out[str(sku).strip().upper()] = " ".join(
            str(b).strip() for b in bits if b not in (None, ""))
    return out


def thumbnail(path, width=THUMB_W):
    """A small in-memory copy. The original is never touched."""
    im = Image.open(path)
    im = im.convert("RGB")
    h = max(1, int(round(im.height * width / float(im.width))))
    im = im.resize((width, h), Image.LANCZOS)
    buf = io.BytesIO()
    im.save(buf, "JPEG", quality=80)
    buf.seek(0)
    return buf, width, h


def clear_rows(ws, first=2):
    """Wipe the data rows without touching the header or the notes below.

    The images go FIRST. openpyxl reads the existing pictures back into
    ws._images when the workbook is opened, so re-adding without clearing
    stacks another copy on every run -- three runs, three thumbnails piled on
    one cell and three JPEGs inside the file. Doing it after the loop was not
    enough either: the loop stops at the notes block and used to return from
    the function there, so the line never ran at all.
    """
    ws._images = []
    last = ws.max_row
    for row in ws.iter_rows(min_row=first, max_row=last, max_col=len(ws[1])):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("What each"):
                return
            cell.value = None


def fill(ws, photos, names, folder):
    hdr = [c.value for c in ws[1]]
    col = {h: get_column_letter(i + 1) for i, h in enumerate(hdr) if h}
    written = 0

    for r, sku in enumerate(sorted(photos), start=2):
        slots = photos[sku]
        ws["%s%d" % (col["SKU"], r)] = sku
        ws["%s%d" % (col["Card"], r)] = names.get(sku, "")
        ws["%s%d" % (col["Front"], r)] = slots.get("", "")
        ws["%s%d" % (col["Back"], r)] = slots.get("-back", "")
        ws["%s%d" % (col["Extras"], r)] = ", ".join(
            slots[s] for s in ("-2", "-3") if s in slots)
        ws["%s%d" % (col["How many"], r)] = len(slots)
        ws["%s%d" % (col["Picture URL for eBay"], r)] = "|".join(
            "%s/%s/%s" % (PAGES, PHOTOS, slots[s])
            for s in SLOT_ORDER if s in slots)

        front = slots.get("") or slots.get("-back")
        if front:
            buf, w, h = thumbnail(os.path.join(folder, front))
            img = XLImage(buf)
            img.width, img.height = w, h
            ws.add_image(img, "%s%d" % (col["Picture"], r))
            ws.row_dimensions[r].height = h * 0.78   # points, not pixels
        written += 1
    return written


def main():
    p = argparse.ArgumentParser(
        description="Put the card photos into the workbook's Photos tab.")
    p.add_argument("--workbook", default=WORKBOOK)
    p.add_argument("--photos", default=PHOTOS)
    a = p.parse_args()

    inuse.refuse_if_open(a.workbook)

    if not os.path.exists(a.workbook):
        sys.exit("no workbook at %s -- run make_workbook.py first" % a.workbook)

    photos = find_photos(a.photos)
    if not photos:
        print("no CRH-#### pictures in %s, so there is nothing to show."
              % a.photos)
        print("file some with add_photos.py or file_batch.py first.")
        return 0

    wb = load_workbook(a.workbook)
    if "Photos" not in wb.sheetnames:
        sys.exit("this workbook has no Photos tab -- rebuild it with "
                 "make_workbook.py --force (it will not keep what you typed).")

    names = card_names(wb["Inventory"]) if "Inventory" in wb.sheetnames else {}
    ws = wb["Photos"]
    clear_rows(ws)
    n = fill(ws, photos, names, a.photos)
    wb.save(a.workbook)

    pics = sum(len(v) for v in photos.values())
    print("Photos tab rewritten: %d card%s, %d picture%s."
          % (n, "" if n == 1 else "s", pics, "" if pics == 1 else "s"))
    missing = [s for s in sorted(photos) if "" not in photos[s]]
    if missing:
        print("no front picture (only a back or an extra): %s"
              % ", ".join(missing))
    unnamed = [s for s in sorted(photos) if not names.get(s)]
    if unnamed:
        print("not on Inventory yet: %s" % ", ".join(unnamed))
    return 0


if __name__ == "__main__":
    sys.exit(main())
