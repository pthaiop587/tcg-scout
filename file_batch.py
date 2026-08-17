"""Put a whole scanned batch into the workbook, photos and all.

    python file_batch.py batch.json
    python file_batch.py batch.json --crops photos/crops --move
    python file_batch.py batch.json --dry-run

This is the path where you do no data entry. You scan, Claude reads the crops
and writes the batch file, and this files the lot: a row per card with a SKU,
the photos renamed onto those SKUs, and the eBay export left ready.

    python crop_scans.py --src "G:/Scans" --rotate 180
    (Claude looks at photos/crops and writes batch.json)
    python file_batch.py batch.json
    python make_ebay_csv.py

The batch file is JSON because the workbook has thirty columns and a pipe
format would be unreadable long before it was complete:

    {
      "pairs": true,
      "cards": [
        {"player": "Shedeur Sanders", "year": 2025,
         "brand": "Panini Prizm Draft Picks", "insert": "Student Orientation",
         "parallel": "Gold Cracked Ice", "num": "8", "rc": true,
         "team": "Colorado Buffaloes", "league": "NCAA",
         "unsure": ["parallel", "market"]}
      ]
    }

"pairs" says the crops are front, back, front, back -- which is how a batch
comes off a flatbed -- so two pictures are filed onto each card.

WHAT "unsure" IS FOR, and why it is not decoration. A parallel's name is not
printed anywhere on a card: a Silver Prizm does not say "Silver", so it is
read off the colour and pattern, and that is a judgement rather than a fact.
Sports singles have no free price feed either, so a market value cannot be
looked up at all. Any card carrying an unsure field is filed as **Review**
instead of Unlisted, and make_ebay_csv.py only ever exports Unlisted -- so an
uncertain card physically cannot reach a listing until somebody opens the
workbook, checks the field and changes the status by hand. The guard is at
the export, which is the boundary that matters.
"""

import argparse
import io
import json
import os
import re
import sys
from datetime import date

from openpyxl import load_workbook

import autofill

import add_photos

import inuse

WORKBOOK = "Card Run HQ - Master.xlsx"
CROPS = os.path.join("photos", "crops")

# JSON key -> Inventory column header. Anything not here is rejected rather
# than dropped, so a typo in a batch file is loud instead of a missing value.
FIELDS = {
    "player": "Player or card name",
    "year": "Year",
    "brand": "Brand / set",
    "insert": "Insert set",
    "parallel": "Parallel",
    "num": "Card #",
    "serial": "Serial /",
    "team": "Team",
    "sport": "Sport or game",
    "league": "League",
    "category": "Category",
    "condition": "Card condition",
    "grader": "Graded by",
    "grade": "Grade",
    "cert": "Cert #",
    "qty": "Qty",
    "cost": "Cost each",
    "market": "Market value",
    "ask": "Ask price",
    "source": "Source",
    "lot": "Lot ID",
    "notes": "Notes",
}
FLAGS = {"rc": "RC", "auto": "Auto", "relic": "Relic"}
DEFAULTS = {"sport": "Football",
            "condition": "Near Mint or Better", "qty": 1}


def category_for(card):
    """Sports or TCG, worked out from the game rather than assumed.

    This used to default to "Sports" whatever the card was, so a Pokemon card
    filed from a batch arrived as Sports -- and Category is the column
    make_ebay_csv.py picks the eBay category code from. The listing goes up
    under the wrong category with nothing looking wrong in the sheet.

    The map lives in autofill.py and is imported rather than repeated: two
    copies of it would agree until the day somebody added a game to one.
    """
    if card.get("category"):
        return card["category"]
    sport = str(card.get("sport") or DEFAULTS["sport"]).strip().lower()
    return autofill.CATEGORY_OF.get(sport)      # None rather than a guess

# "photos" is how many pictures in the crops folder belong to this card: 1 for
# a front only, 2 for front and back. Card desk writes it per card, because a
# batch is rarely uniform -- you photograph the back of the ones worth it. It
# is why the pictures are filed here rather than handed to add_photos --pairs,
# which can only do the same count for every card in the batch.
ALLOWED = set(FIELDS) | set(FLAGS) | {"unsure", "photos"}


def norm(s):
    return re.sub(r"[^a-z0-9]", "", str(s or "").lower())


def load_batch(path):
    with io.open(path, encoding="utf-8") as fh:
        data = json.load(fh)
    if isinstance(data, list):                 # a bare list of cards is fine
        data = {"cards": data}
    cards = data.get("cards")
    if not isinstance(cards, list) or not cards:
        sys.exit("%s has no cards in it" % path)

    for i, c in enumerate(cards, 1):
        if not isinstance(c, dict):
            sys.exit("card %d is not an object" % i)
        if not str(c.get("player", "")).strip():
            sys.exit("card %d has no player -- nothing can be filed without a name" % i)
        stray = set(c) - ALLOWED
        if stray:
            sys.exit("card %d has unknown field(s): %s\nknown: %s"
                     % (i, ", ".join(sorted(stray)), ", ".join(sorted(ALLOWED))))
        bad = [u for u in c.get("unsure", []) if u not in FIELDS]
        if bad:
            sys.exit("card %d marks unknown field(s) unsure: %s"
                     % (i, ", ".join(bad)))
        n = c.get("photos")
        if n is not None and (not isinstance(n, int) or n < 0 or n > 4):
            sys.exit("card %d wants %r photos; it has to be 0 to 4" % (i, n))

    pairs = bool(data.get("pairs"))
    if pairs:                       # shorthand for "every card has two"
        for c in cards:
            c.setdefault("photos", 2)
    return pairs, cards


def add_rows(wb_path, cards, dry_run=False):
    """Write every card in one open/save, and hand back the SKUs in order."""
    wb = load_workbook(wb_path)
    ws = wb["Inventory"]
    idx = {norm(c.value): c.column for c in ws[1] if c.value}

    missing = [h for h in list(FIELDS.values()) + list(FLAGS.values())
               + ["SKU", "Status", "Date in"] if norm(h) not in idx]
    if missing:
        sys.exit("Inventory is missing column(s): %s" % ", ".join(missing))

    row = 2
    name_col = idx[norm(FIELDS["player"])]
    while ws.cell(row=row, column=name_col).value not in (None, ""):
        row += 1

    highest = 0
    for vals in ws.iter_rows(min_row=2, min_col=idx[norm("SKU")],
                             max_col=idx[norm("SKU")], values_only=True):
        m = re.match(r"CRH-(\d+)$", str(vals[0] or "").strip())
        if m:
            highest = max(highest, int(m.group(1)))

    out = []
    for card in cards:
        highest += 1
        sku = "CRH-%04d" % highest
        unsure = list(card.get("unsure", []))

        if not dry_run:
            def put(header, value):
                if value in (None, "", False):
                    return
                ws.cell(row=row, column=idx[norm(header)], value=value)

            put("SKU", sku)
            # the whole guard: Review is not Unlisted, and only Unlisted exports
            put("Status", "Review" if unsure else "Unlisted")
            put("Date in", date.today())
            for key, header in FIELDS.items():
                if key == "category":
                    put(header, category_for(card))
                    continue
                put(header, card.get(key, DEFAULTS.get(key)))
            for key, header in FLAGS.items():
                put(header, "Yes" if card.get(key) else "No")
            if unsure:
                note = "CHECK: " + ", ".join(unsure)
                had = card.get("notes")
                ws.cell(row=row, column=idx[norm("Notes")],
                        value=(had + " | " + note) if had else note)
            ws.cell(row=row, column=idx[norm("Date in")]).number_format = "yyyy-mm-dd"

        out.append({"sku": sku, "row": row, "player": card["player"],
                    "unsure": unsure, "photos": int(card.get("photos", 0))})
        row += 1

    if not dry_run:
        wb.save(wb_path)
    return out


def main():
    p = argparse.ArgumentParser(
        description="File a scanned batch into the workbook, photos and all.")
    p.add_argument("batch", help="the JSON Claude wrote")
    p.add_argument("--workbook", default=WORKBOOK)
    p.add_argument("--crops", default=CROPS,
                   help="folder of cropped card pictures (default photos/crops)")
    p.add_argument("--no-photos", action="store_true",
                   help="just add the rows; leave the pictures alone")
    p.add_argument("--move", action="store_true",
                   help="clear the crops folder once the photos are filed")
    p.add_argument("--dry-run", action="store_true",
                   help="say what would happen and change nothing")
    a = p.parse_args()

    inuse.refuse_if_open(a.workbook)

    if not os.path.exists(a.workbook):
        sys.exit("no workbook at %s -- run make_workbook.py first" % a.workbook)

    pairs, cards = load_batch(a.batch)
    added = add_rows(a.workbook, cards, a.dry_run)

    print("%s%d card%s" % ("would add " if a.dry_run else "added ",
                           len(added), "" if len(added) == 1 else "s"))
    for e in added:
        print("   %s  %-28s %s" % (e["sku"], e["player"][:28],
                                   ("REVIEW - check " + ", ".join(e["unsure"]))
                                   if e["unsure"] else "Unlisted"))

    flagged = [e for e in added if e["unsure"]]
    if flagged:
        print("\n%d filed as Review, so make_ebay_csv.py will not export %s."
              % (len(flagged), "it" if len(flagged) == 1 else "them"))
        print("Open the workbook, settle the CHECK fields, set Status to "
              "Unlisted, and they join the next export.")

    if a.no_photos or a.dry_run:
        if not a.dry_run:
            print("\nphotos left alone.")
        return 0

    if not os.path.isdir(a.crops):
        print("\nno crops folder at %s -- rows added, photos not filed." % a.crops)
        return 0

    return file_photos(added, a.crops, a.move)


def file_photos(added, crops, move=False):
    """Put each card's pictures on its SKU, in order.

    Pictures are taken in filename order and handed out a card at a time --
    the first card's `photos` of them, then the next card's. Card desk names
    its downloads so that order is the order shown on the queue, front before
    back.

    The count has to match exactly. One picture out and every card after it
    gets somebody else's photo, which is a wrong picture on a live listing --
    so a mismatch files nothing at all and says what it found instead.
    """
    files = sorted(f for f in os.listdir(crops)
                   if os.path.splitext(f)[1].lower() in add_photos.EXTS)
    want = sum(e["photos"] for e in added)

    if not want:
        print("\nno photo counts in the batch, so no pictures were filed.")
        if files:
            print("%d picture(s) are sitting in %s -- add \"photos\": 1 or 2 to "
                  "each card, or file them by hand with add_photos.py --assign."
                  % (len(files), crops))
        return 0

    if want != len(files):
        print("\nNOTHING FILED. The batch accounts for %d picture%s and %s holds %d."
              % (want, "" if want == 1 else "s", crops, len(files)))
        print("Filing them anyway would put one card's photo on another card's "
              "listing, so fix the count and run it again.")
        return 1

    print("\nfiling %d picture%s from %s:" % (len(files), "" if len(files) == 1 else "s", crops))
    i = 0
    for e in added:
        for k in range(e["photos"]):
            src = os.path.join(crops, files[i])
            add_photos.place(src, e["sku"], add_photos.SLOTS[k][0])
            if move:
                os.remove(src)
            i += 1
    print("\nfiled %d onto %d card%s." % (len(files), len(added),
                                          "" if len(added) == 1 else "s"))
    return 0


if __name__ == "__main__":
    sys.exit(main())
