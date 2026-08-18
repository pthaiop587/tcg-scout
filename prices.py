"""Look up what each card is worth raw, in a 9, and in a 10 -- and when one last sold.

    python prices.py --report            # match only; touch nothing
    python prices.py --limit 5           # try five, print what it found
    python prices.py --go                # fill the columns in

A price with no date beside it is the thing that loses money here. "$200" on a
PSA 10 reads as fact, but if the last one actually sold in April it is a
guess dressed as a number, and you will price a real card against it. So every
price column has a date column next to it, and the date is the date a card of
THAT grade last changed hands -- not the date this script ran.

Where the numbers come from: sportscardspro.com, which is the PriceCharting
family. Two kinds of page.

The set page lists every card and parallel with Ungraded, Grade 9 and PSA 10
side by side, so all three prices for all your cards come from one page load.
It lazy-loads as you scroll, so this scrolls until the row count stops moving.

The individual card page carries the sales themselves, dated, in tables under
#price_comparison. Those tables have no ids and no headings, and their ORDER is
the only thing separating ungraded from PSA 9 from PSA 10 -- which would break
silently the first time the site added a grade tier. So the grade is read out
of each listing's own title instead ("... PSA 10 #166"), and a row that names
no grade is a raw sale. Order is then only used as a cross-check.

The site blocks plain fetching, so this drives a real browser, one page at a
time with a pause between, which is what you would be doing by hand anyway. It
is roughly one page per card: slow, not heavy.

Nothing is written without --go, and --go never overwrites a price you typed
yourself unless you pass --overwrite.
"""

import argparse
import csv
import datetime as dt
import os
import re
import sys
import time
from copy import copy

from openpyxl import load_workbook

import colleges
import inuse

WORKBOOK = "Card Run HQ - Master.xlsx"
BASE = "https://www.sportscardspro.com"
UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36")

# The six columns this fills, in pairs. Appended to Inventory, never inserted:
# workbook_extra.py addresses Inventory by column letter.
PRICE_COLS = [
    ("Raw price", 11), ("Raw last sold", 12), ("Raw last sale", 12),
    ("PSA 9 price", 11), ("PSA 9 last sold", 13), ("PSA 9 last sale", 13),
    ("PSA 10 price", 11), ("PSA 10 last sold", 14), ("PSA 10 last sale", 14),
]
TIERS = ["raw", "psa9", "psa10"]
# guide price, date of the most recent sale, and what THAT sale actually made.
# The guide is a calculation over recent sales; the last sale is one real
# number. They disagree, and the gap is the point -- a PSA 10 guided at $200
# whose last one fetched $275 is telling you something a single figure cannot.
COL_OF = {"raw": ("Raw price", "Raw last sold", "Raw last sale"),
          "psa9": ("PSA 9 price", "PSA 9 last sold", "PSA 9 last sale"),
          "psa10": ("PSA 10 price", "PSA 10 last sold", "PSA 10 last sale")}
HISTORY = "price_history.csv"

# Which console page a card lives on. Base cards and their parallels share one;
# named inserts get their own.
SET_SLUG = "football-cards-2025-panini-prizm-draft-picks"
INSERT_SLUG = {
    "student orientation": SET_SLUG + "-student-orientation",
    "fearless": SET_SLUG + "-fearless",
    "instant impact": SET_SLUG + "-instant-impact",
    "new recruits": SET_SLUG + "-new-recruits",
    "signing day": SET_SLUG + "-signing-day",
    "on campus": SET_SLUG + "-on-campus",
    "trophy hunting": SET_SLUG + "-trophy-hunting",
    "manga": SET_SLUG + "-manga",
}

# Two price guides, one company. sportscardspro carries the sports sets and
# pricecharting carries the rest, Pokemon included. Which one a card belongs
# to follows from its Brand / set, so one pass prices a mixed workbook.
SITES = {"sportscardspro": "https://www.sportscardspro.com",
         "pricecharting": "https://www.pricecharting.com"}

POKEMON_SETS = {
    "pitch black": "pokemon-pitch-black",
}

# Baseball. 2026 Topps Series 2, which is also what the All-Star Game mega
# box is filled with.
MLB_SET = "baseball-cards-2026-topps-series-2"
MLB_INSERTS = {
    "stars of mlb": MLB_SET + "-stars-of-mlb",
    "titans of the game": MLB_SET + "-titans-of-the-game",
    "titans": MLB_SET + "-titans-of-the-game",
    "glove work": MLB_SET + "-glove-work",
    "1991 topps": MLB_SET + "-1991",
    "1991": MLB_SET + "-1991",
    "1991 topps all-stars": MLB_SET + "-1991-all-stars",
    "1991 all-stars": MLB_SET + "-1991-all-stars",
}

# Brand / set -> (its base page, the insert pages hanging off it). Anything
# unrecognised falls through to the football set, which is where every card
# went before there was more than one game.
# Keys are what norm() makes of the Brand / set cell. Spelled out rather than
# written as norm("Panini Prizm Draft Picks"): norm is defined further down,
# and calling it up here is a NameError at import.
SETS = {
    "panini prizm draft picks": (SET_SLUG, INSERT_SLUG),
    "topps series 2": (MLB_SET, MLB_INSERTS),
}

# The Parallel column holds two different kinds of thing for a Pokemon card.
# "Reverse Holo" is a physically different card with its own page and its own
# price -- often ten times the plain one -- and belongs in the URL. "Secret
# Rare" only says how hard the card was to pull, and putting it in the URL
# asks for a page that does not exist.
RARITY_WORDS = {
    "common", "uncommon", "rare", "double rare", "ultra rare", "secret rare",
    "illustration rare", "special illustration rare", "hyper rare",
    "shiny rare", "shiny ultra rare", "ace spec rare", "radiant rare",
    "amazing rare", "promo",
}

# "Arch Manning [Gold Ice] #166", "Jaxson Dart #90 [RC]"
LABEL = re.compile(
    r"^(?P<name>.+?)\s*(?:\[(?P<par>[^\]]+)\]\s*)?#(?P<num>[\w.-]+)\s*"
    r"(?:\[RC\])?\s*$")
DATE = re.compile(r"^\d{4}-\d{2}-\d{2}$")
MONEY = re.compile(r"\$\s*([\d,]+(?:\.\d{1,2})?)")

SUFFIX = re.compile(r"\b(jr|sr|ii|iii|iv|v)\.?$", re.I)


def norm(s):
    """A name two sources will agree on: no case, no punctuation, no suffix.

    Dots and apostrophes are DELETED, not turned into spaces. One source
    writes "T.J. Sanders" and the other "TJ Sanders"; spacing the dots gives
    "t j sanders" against "tj sanders", which do not match -- and the card
    then has to be found by number instead, or not at all. Everything else
    becomes a space, so a hyphenated name still splits into words.
    """
    s = str(s or "").strip().lower()
    s = s.replace("&", "and")
    s = re.sub(r"[.'’]", "", s)
    s = re.sub(r"[^\w\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    s = SUFFIX.sub("", s).strip()
    return s


def blank_cell(v):
    return v is None or (isinstance(v, str) and not v.strip())


def key(name, parallel, num):
    return (norm(name), norm(parallel), str(num or "").strip().lstrip("#"))


def money(txt):
    m = MONEY.search(str(txt or ""))
    if not m:
        return None
    return float(m.group(1).replace(",", ""))


def grade_of(title):
    """Which price tier a completed listing belongs to, from its own words.

    Read off the listing rather than off which table it sat in: the tables
    carry no ids, and a new grade tier appearing on the site would shift every
    index by one without anything failing."""
    t = " " + re.sub(r"\s+", " ", str(title or "")).lower() + " "
    if re.search(r"\b(psa|bgs|sgc|cgc|hga|tag)\s*10\b", t):
        return "psa10"
    if re.search(r"\bpsa\s*9(?!\s*\.)\b", t):
        return "psa9"
    # any other stated grade is a tier we do not track; not raw
    if re.search(r"\b(psa|bgs|sgc|cgc|hga|tag)\s*\d", t):
        return None
    if re.search(r"\b(graded|slab)\b", t):
        return None
    return "raw"


# --- the site ---------------------------------------------------------------

def harvest_console(pg, slug, settle=3):
    """Every card on a set page: label -> (link, ungraded, g9, psa10).

    The table lazy-loads, so scroll until the row count stops moving."""
    pg.goto("%s/console/%s" % (BASE, slug), wait_until="domcontentloaded",
            timeout=60000)
    pg.wait_for_timeout(2000)
    if not pg.locator("#games_table").count():
        return {}
    # The table appends as you reach the bottom. Stopping early is silent --
    # you get a partial index and cards "miss" for no visible reason, and not
    # the same ones twice. So: keep going until the count has held still for
    # several checks, with a hard ceiling so a broken page cannot spin here.
    seen, stable, spins = -1, 0, 0
    while stable < settle and spins < 400:
        spins += 1
        pg.keyboard.press("End")
        pg.mouse.wheel(0, 40000)
        pg.wait_for_timeout(650)
        n = pg.locator("#games_table tbody tr").count()
        if n == seen:
            stable += 1
        else:
            stable, seen = 0, n

    rows = pg.evaluate("""() => {
      const out = [];
      document.querySelectorAll('#games_table tbody tr').forEach(tr => {
        const td = tr.querySelectorAll('td');
        if (td.length < 5) return;
        const a = tr.querySelector('a');
        out.push({label: td[1].innerText.trim(),
                  href: a ? a.getAttribute('href') : '',
                  ungraded: td[2].innerText.trim(),
                  g9: td[3].innerText.trim(),
                  psa10: td[4].innerText.trim()});
      });
      return out;
    }""")

    idx, bynum, noname = {}, {}, {}
    for r in rows:
        m = LABEL.match(r["label"])
        if not m:
            continue
        nm, par, num = m.group("name"), m.group("par") or "", m.group("num")
        r["site_name"] = nm.strip()
        idx.setdefault(key(nm, par, num), r)
        # Insert sets number their cards differently from the base set
        # (#II-SSS, not #18), so keep a way in that ignores the number too.
        noname.setdefault((norm(nm), norm(par)), r)
        # Card number and parallel identify a card on their own inside a set.
        # The NAME is the part that gets mistyped, so keep a way in that does
        # not depend on it.
        bynum.setdefault((norm(par), str(num).strip().lstrip("#")), r)
    return {"byname": idx, "bynum": bynum, "noname": noname}


def slug(s):
    s = norm(s).replace(" ", "-")
    return re.sub(r"-+", "-", s).strip("-")


def pokemon_num(num):
    """013/084 is how a Pokemon card is printed; 13 is how it is addressed."""
    n = str(num or "").strip().lstrip("#").split("/")[0].strip()
    return n.lstrip("0") or n


def name_forms(name, apostrophe_first):
    """The ways a price guide might spell one name in a URL.

    sportscardspro deletes the apostrophe (jamarr-chase); pricecharting keeps
    it, url-encoded (misty%27s-vitality). Neither reliably accepts the other's
    spelling, and which site wants which is not worth remembering at the call
    site -- so offer both and put the likely one first.
    """
    n = str(name or "").strip().replace("\u2019", "'")
    plain = [slug(n)]
    if "'" not in n:
        return plain
    raw = re.sub(r"-+", "-", re.sub(r"[^a-z0-9']+", "-", n.lower())).strip("-")
    both = [raw.replace("'", "%27"), raw.replace("'", "-")]
    forms = both + plain if apostrophe_first else plain + both
    return list(dict.fromkeys(forms))


def on_default_set(brand):
    """True when a card's Brand / set is the one the fallback set page covers.

    An unrecognised brand still routes to the default set so that nothing
    quietly stops working, but harvesting two thousand football rows to hunt
    for a card that was never in that set costs two minutes and finds nothing.
    """
    b = norm(brand)
    return not b or b in SETS


def route(card):
    """Where a card's prices live, and how it is addressed there.

    Returns (site, set pages to try, number, parallel, is_pokemon).
    """
    brand = norm(card.get("brand", ""))
    if brand in POKEMON_SETS:
        par = card.get("parallel") or ""
        if norm(par) in RARITY_WORDS:
            par = ""
        return (SITES["pricecharting"], [POKEMON_SETS[brand]],
                pokemon_num(card.get("num")), par, True)
    base, inserts = SETS.get(brand, (SET_SLUG, INSERT_SLUG))
    ins = (card.get("insert") or "").strip().lower()
    first = inserts.get(ins, base) if ins else base
    order = [first] + ([base] if first != base else [])
    return (SITES["sportscardspro"], order,
            str(card.get("num") or "").strip().lstrip("#").lower(),
            card.get("parallel") or "", False)


def card_url(set_slug, name, parallel, num, base=None, name_form=None):
    """The card page URL, built rather than looked up.

    The set page appends rows as you scroll and does not always finish, so an
    index built from it is missing a different handful every run. This does
    not depend on it."""
    bits = [name_form or slug(name)]
    if parallel:
        bits.append(slug(parallel))
    bits.append(str(num).strip().lstrip("#").lower())
    return "%s/game/%s/%s" % (base or BASE, set_slug, "-".join(bits))


def card_page(pg, url, delay):
    """Prices and the most recent sale date per tier, from one card page.

    Prices come from here rather than the set page so that a card missing from
    the index still gets its numbers, and so the two can never disagree."""
    r = pg.goto(url, wait_until="domcontentloaded", timeout=60000)
    if r is not None and r.status >= 400:
        time.sleep(delay)
        # Four values, like the success path. Returning three worked only for
        # as long as no URL ever 404'd; the moment one did -- which trying a
        # second spelling of a name makes routine -- the caller unpacked a
        # short tuple and the whole run died on one missing page.
        return None, {}, {}, []
    pg.wait_for_timeout(1200)

    # PriceCharting inherits its field names from video games; these are the
    # ids the card grades actually land in.
    prices = pg.evaluate("""() => {
      const get = id => {
        const el = document.querySelector(id);
        return el ? el.innerText.trim() : '';
      };
      return {raw: get('#used_price'), psa9: get('#graded_price'),
              psa10: get('#manual_only_price')};
    }""")
    prices = {k: money(v) for k, v in prices.items()}

    rows = pg.evaluate("""() => {
      const out = [];
      document.querySelectorAll('#price_comparison table tr').forEach(tr => {
        const td = tr.querySelectorAll('td');
        if (td.length < 4) return;
        out.push({date: td[0].innerText.trim(),
                  title: td[2].innerText.trim(),
                  price: td[3].innerText.trim()});
      });
      return out;
    }""")

    best, last, titles = {}, {}, []
    for r in rows:
        titles.append(r["title"])
        d = r["date"].split("\n")[0].strip()
        if not DATE.match(d):
            continue
        g = grade_of(r["title"])
        if g and (g not in best or d > best[g]):
            best[g] = d
            # The cell can hold two figures -- what the item made, then the
            # same with postage. Take the first: it is the number the guide
            # price is comparable with.
            last[g] = money(r["price"])
    time.sleep(delay)
    return prices, best, last, titles


# --- the workbook -----------------------------------------------------------

def ensure_columns(ws):
    """Append any missing price column. Appending matters: workbook_extra.py
    addresses Inventory by letter, so an insert moves every formula."""
    hdr = [c.value for c in ws[1]]
    added = []
    for name, width in PRICE_COLS:
        if name in hdr:
            continue
        col = len(hdr) + 1
        c = ws.cell(row=1, column=col, value=name)
        base = ws.cell(row=1, column=1)
        # openpyxl hands back a StyleProxy, which cannot be assigned onward
        c.font = copy(base.font)
        c.fill = copy(base.fill)
        c.alignment = copy(base.alignment)
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col)].width = width
        hdr.append(name)
        added.append(name)
    return hdr, added


def wanted_skus(values):
    """The SKUs asked for on the command line.

    --sku CRH-0062, or repeated, or comma-separated, or any mix of those,
    because all three are what somebody reaches for and none of them is wrong.
    Case is levelled: a workbook holds CRH-0062 and a person types crh-62 at
    their own risk, but not crh-0062."""
    out = set()
    for v in (values or []):
        out.update(x.strip().upper() for x in str(v).split(",") if x.strip())
    return out


def read_cards(ws, hdr, sport):
    """The cards to price.

    sport may name several, comma separated, so that one daily run covers a
    workbook holding football and Pokemon both.
    """
    want = {x.strip().lower() for x in str(sport or "").split(",") if x.strip()}
    g = {n: i for i, n in enumerate(hdr)}
    out = []
    for r in range(2, ws.max_row + 1):
        vals = [c.value for c in ws[r]]
        vals += [None] * (len(hdr) - len(vals))
        name = vals[g["Player or card name"]]
        if not name or not str(name).strip():
            continue
        if want and str(vals[g.get("Sport or game", 0)] or "").strip() \
                .lower() not in want:
            continue
        out.append({
            "row": r, "name": str(name).strip(),
            "num": str(vals[g["Card #"]] or "").strip(),
            "parallel": str(vals[g["Parallel"]] or "").strip(),
            "insert": str(vals[g["Insert set"]] or "").strip(),
            "brand": (str(vals[g["Brand / set"]] or "").strip()
                      if "Brand / set" in g else ""),
            "sku": vals[g["SKU"]],
        })
    return out


def main():
    p = argparse.ArgumentParser(
        description="Fill raw / PSA 9 / PSA 10 prices and last-sold dates.")
    p.add_argument("--workbook", default=WORKBOOK)
    p.add_argument("--sport", default="Football,Pokemon",
                   help="which sports/games to price; comma-separate for "
                        "several (default Football,Pokemon)")
    p.add_argument("--sku", action="append", metavar="CRH-0062",
                   help="just this card. Repeat it, or comma-separate, for a "
                        "few. Skips the sport filter, so any card works.")
    p.add_argument("--limit", type=int, help="only the first N cards")
    p.add_argument("--delay", type=float, default=1.5,
                   help="seconds between card pages (default 1.5)")
    p.add_argument("--report", action="store_true",
                   help="match against the site and say what it found; no dates, no writing")
    p.add_argument("--go", action="store_true", help="write the values in")
    p.add_argument("--overwrite", action="store_true",
                   help="replace values already in the price columns")
    p.add_argument("--daily", action="store_true",
                   help="a day's check: refresh every price and audit what "
                        "moved (same as --go --overwrite --teams)")
    p.add_argument("--audit-pct", type=float, default=5.0, metavar="N",
                   help="call out anything that moved N%% or more (default 5)")
    p.add_argument("--history", default=HISTORY,
                   help="CSV each run is appended to (default %s)" % HISTORY)
    p.add_argument("--log", metavar="FILE",
                   help="append the audit to FILE as well as printing it")
    p.add_argument("--teams", action="store_true",
                   help="also read each card's school out of its listings")
    p.add_argument("--team-min", type=int, default=2, metavar="N",
                   help="how many listings must agree on the school "
                        "(default 2; use 1 for cards with few sales)")
    p.add_argument("--fix-names", action="store_true",
                   help="rewrite mistyped player names to the site's spelling")
    a = p.parse_args()

    if a.daily:
        a.go = a.overwrite = a.teams = True
        a.report = False
        if not a.log:
            a.log = "price-check.log"

    inuse.refuse_if_open(a.workbook)

    wb = load_workbook(a.workbook)
    ws = wb["Inventory"]
    hdr, added = ensure_columns(ws)
    # Naming a card means you want THAT card, whatever sport it is. Applying
    # the sport filter as well would answer "no Football cards found" for a
    # Pokemon SKU, which is true and useless.
    wanted = wanted_skus(a.sku)

    cards = read_cards(ws, hdr, None if wanted else a.sport)
    if wanted:
        cards = [c for c in cards
                 if str(c["sku"] or "").strip().upper() in wanted]
        missing = wanted - {str(c["sku"] or "").strip().upper() for c in cards}
        if missing:
            sys.exit("no card with SKU %s in %s"
                     % (", ".join(sorted(missing)), a.workbook))
    if a.limit:
        cards = cards[:a.limit]
    if not cards:
        sys.exit("no %s cards found in %s" % (a.sport, a.workbook))

    what = ", ".join(c["sku"] for c in cards) if wanted \
        else "%d %s card(s)" % (len(cards), a.sport)
    print("%s%s" % (what, "; added columns: " + ", ".join(added)
                    if added else ""))

    from playwright.sync_api import sync_playwright
    g = {n: i for i, n in enumerate(hdr)}
    found = missed = 0
    typos = []          # matched by number, but the name you typed differs
    teams, weak = [], []    # schools read off the listings, confident and not
    moves = []              # guide prices that changed since the last run
    snapshot = []           # every figure this run, for price_history.csv

    with sync_playwright() as pw:
        b = pw.chromium.launch()
        pg = b.new_page(user_agent=UA)

        # The set page is only needed when a card cannot be found by its own
        # URL -- which in practice means the name is mistyped. Harvesting it
        # costs a couple of minutes and two thousand rows, so it is fetched on
        # demand rather than up front. A run where every name is right never
        # touches it at all.
        consoles = {}

        def console(slug):
            if slug not in consoles:
                print("   (looking a card up the slow way -- fetching %s)"
                      % slug)
                consoles[slug] = harvest_console(pg, slug)
                print("   set page %-56s %4d card(s)"
                      % (slug, len(consoles[slug].get("byname", {}))))
            return consoles[slug]

        for c in cards:
            site, order, cnum, cpar, is_poke = route(c)
            nk = key(c["name"], cpar, cnum)
            numk = (norm(cpar), cnum)

            def find():
                """The set page, consulted only because the URL did not work."""
                for s in order:
                    h = console(s).get("byname", {}).get(nk)
                    if h:
                        return h
                for s in order:                   # name mistyped: by number
                    h = console(s).get("bynum", {}).get(numk)
                    if h:
                        typos.append((c["sku"], c["name"], h["site_name"],
                                      c["parallel"], c["num"], c["row"]))
                        return h
                for s in order:                   # inserts number differently
                    h = console(s).get("noname", {}).get(
                        (norm(c["name"]), norm(c["parallel"])))
                    if h:
                        return h
                return None

            hit = find() if a.report else None

            if a.report:
                # fast sanity check: index only, no card pages
                if not hit:
                    missed += 1
                    print("   MISS %-9s %-22s %-10s #%s"
                          % (c["sku"], c["name"], c["parallel"], c["num"]))
                    continue
                found += 1
                print("   %-9s %-22s %-9s raw %-8s 9 %-8s 10 %s"
                      % (c["sku"], c["name"][:22], c["parallel"],
                         money(hit["ungraded"]), money(hit["g9"]),
                         money(hit["psa10"])))
                continue

            # The card page is the source of truth: it carries the dates, and
            # taking the prices from the same page means the two can never
            # disagree. The index is only consulted for a URL it may not have.
            def usable(pr):
                return bool(pr) and any(v is not None for v in pr.values())

            urls = [card_url(s, c["name"], cpar, cnum, base=site,
                             name_form=nf)
                    for s in order
                    for nf in name_forms(c["name"], is_poke)]

            prices, dates, lasts, titles = None, {}, {}, []
            for u in urls:
                prices, dates, lasts, titles = card_page(pg, u, a.delay)
                if usable(prices):
                    break

            if not usable(prices) and not is_poke \
                    and on_default_set(c.get("brand")):
                # The built URL got nowhere. NOW the set page is worth its two
                # minutes: the name is probably spelled differently there.
                # Only for the sports sets -- harvest_console reaches for
                # sportscardspro, so running it for a Pokemon card would
                # search the wrong site and answer confidently.
                hit = find()
                if hit:
                    h = hit["href"]
                    u = h if h.startswith("http") else site + h
                    prices, dates, lasts, titles = card_page(pg, u, a.delay)
            if not prices or not any(v is not None for v in prices.values()):
                missed += 1
                print("   MISS %-9s %-22s %-10s #%s"
                      % (c["sku"], c["name"], c["parallel"], c["num"]))
                continue

            found += 1
            print("   %-9s %-22s %-9s raw %-8s 9 %-8s 10 %-8s  %s"
                  % (c["sku"], c["name"][:22], c["parallel"],
                     prices["raw"], prices["psa9"], prices["psa10"],
                     " ".join("%s=%s" % (t, dates.get(t, "-")) for t in TIERS)))

            # The school is not a field on any price guide, but sellers put it
            # in the title, so read it from there and take the commonest.
            # Only ever fills a blank. --overwrite is about prices, which move;
            # a player's college does not, so a daily run re-voting on 60
            # already-correct schools is noise that buries the prices.
            if a.teams and not is_poke and "Team" in g:
                cell = ws.cell(row=c["row"], column=g["Team"] + 1)
                if blank_cell(cell.value):
                    school, n = colleges.vote(titles)
                    if school and n >= a.team_min:
                        teams.append((c["sku"], c["name"], school, n))
                        if a.go:
                            cell.value = school
                    else:
                        weak.append((c["sku"], c["name"], school, n))

            # What the workbook said before this run, so the audit can say what
            # moved rather than only what it is now.
            for tier in TIERS:
                pcol, _dcol, _scol = COL_OF[tier]
                was = ws.cell(row=c["row"], column=g[pcol] + 1).value
                now = prices[tier]
                if isinstance(was, (int, float)) and isinstance(now, (int, float)):
                    if abs(now - was) >= 0.005:
                        moves.append((c["sku"], c["name"], tier, float(was),
                                      float(now)))
                snapshot.append((c["sku"], tier, prices[tier],
                                 dates.get(tier) or "", lasts.get(tier)))

            if not a.go:
                continue
            for tier in TIERS:
                pcol, dcol, scol = COL_OF[tier]

                cell = ws.cell(row=c["row"], column=g[pcol] + 1)
                if prices[tier] is not None and (a.overwrite or cell.value in (None, "")):
                    cell.value = prices[tier]
                    cell.number_format = '"$"#,##0.00'

                d = dates.get(tier)
                cell = ws.cell(row=c["row"], column=g[dcol] + 1)
                if d and (a.overwrite or cell.value in (None, "")):
                    cell.value = dt.date.fromisoformat(d)
                    cell.number_format = "yyyy-mm-dd"

                s = lasts.get(tier)
                cell = ws.cell(row=c["row"], column=g[scol] + 1)
                if s is not None and (a.overwrite or cell.value in (None, "")):
                    cell.value = s
                    cell.number_format = '"$"#,##0.00'
        b.close()

    print("\nmatched %d, missed %d" % (found, missed))

    # --- the audit: what actually moved -------------------------------------
    def pct(w, n):
        return 100.0 * (n - w) / w if w else 0.0

    audit = []
    if moves:
        moves.sort(key=lambda m: abs(pct(m[3], m[4])), reverse=True)
        up = [m for m in moves if m[4] > m[3]]
        down = [m for m in moves if m[4] < m[3]]
        net = sum(m[4] - m[3] for m in moves)

        audit.append("%d price(s) moved since the last check -- %d up, %d "
                     "down, %+.2f net across the lot."
                     % (len(moves), len(up), len(down), net))
        # The trap: --go fills BLANKS only, so a card that already had a price
        # keeps it while the audit cheerfully reports the change. It reads
        # exactly like a successful update and is not one.
        if not a.overwrite:
            audit.append("NOT written -- --go only fills empty cells. Add "
                         "--overwrite to replace prices already there.")
        big = [m for m in moves if abs(pct(m[3], m[4])) >= a.audit_pct]
        if big:
            audit.append("")
            audit.append("Moved %g%% or more:" % a.audit_pct)
            audit.append("   %-9s %-22s %-6s %9s %9s %8s"
                         % ("SKU", "card", "grade", "was", "now", "change"))
            for sku, name, tier, was, now in big:
                audit.append("   %-9s %-22s %-6s %9.2f %9.2f %+7.1f%%"
                             % (sku, name[:22], tier, was, now, pct(was, now)))
        else:
            audit.append("Nothing moved by %g%% or more -- a quiet day."
                         % a.audit_pct)
    elif found:
        audit.append("No price changed since the last check.")

    if audit:
        print("\n" + "\n".join(audit))

    # The audit is the part worth keeping. Written by this script rather than
    # by redirecting the whole run, so the log holds the answer and not four
    # minutes of scrolling -- and so the site is asked once, not twice.
    if a.log and audit:
        with open(a.log, "a", encoding="utf-8") as fh:
            fh.write("\n===== %s  (%d card(s) checked, %d missed)\n"
                     % (dt.datetime.now().strftime("%Y-%m-%d %H:%M"),
                        found, missed))
            fh.write("\n".join(audit) + "\n")

    # --- history, so a week from now this is a trend and not a snapshot -----
    if a.go and snapshot:
        new = not os.path.exists(a.history)
        with open(a.history, "a", newline="", encoding="utf-8") as fh:
            w = csv.writer(fh)
            if new:
                w.writerow(["checked", "sku", "grade", "guide price",
                            "last sold", "last sale price"])
            today = dt.date.today().isoformat()
            for sku, tier, guide, sold, last in snapshot:
                w.writerow([today, sku, tier,
                            "" if guide is None else guide, sold,
                            "" if last is None else last])
        print("\nAppended %d row(s) to %s." % (len(snapshot), a.history))

    if teams:
        print("\nschool read off the listings for %d card(s)%s."
              % (len(teams), "" if a.go else " -- not written, add --go"))
    if weak:
        print("\n%d card(s) whose listings never name a school. Sellers just "
              "did not\ntype it; fill these by hand if you want them in the "
              "title:" % len(weak))
        for sku, name, school, n in weak:
            print("   %-9s %-24s %s"
                  % (sku, name, "%s, but only %d listing said so" % (school, n)
                     if school else "nothing found"))

    if typos:
        print("\n%d card(s) matched on number and parallel but the name you "
              "typed does not match the site.\nThe price is right -- the "
              "spelling is what matters, because an eBay title nobody\nsearches "
              "for is a card nobody finds:" % len(typos))
        print("   %-9s %-24s %-24s %s" % ("SKU", "you typed", "site has", "card"))
        for sku, mine, theirs, par, num, _row in typos:
            print("   %-9s %-24s %-24s %s #%s" % (sku, mine, theirs, par, num))
        if a.fix_names and a.go:
            col = g["Player or card name"] + 1
            for sku, mine, theirs, _par, _num, row in typos:
                ws.cell(row=row, column=col, value=theirs)
            print("   -> rewritten to the site's spelling.")
        elif not a.fix_names:
            print("   Add --fix-names (with --go) to correct them in place.")
    if a.go:
        wb.save(a.workbook)
        print("Saved %s. Run sport_tabs.py to push these onto the game tabs."
              % a.workbook)
    else:
        print("Nothing written. Add --go to fill the columns in.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
